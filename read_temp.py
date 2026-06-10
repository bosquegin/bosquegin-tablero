import openpyxl, sys

# Read Bosque salidas
print("=== Bosque salidas.xlsx ===")
wb = openpyxl.load_workbook(r'C:\Users\SupplyDestileria\Documents\Bosque\Claude\Tablero operativo\Salidas\Bosque salidas.xlsx', read_only=True, data_only=True)
print('Sheets:', wb.sheetnames)
ws = wb.active
rows_data = []
for row in ws.iter_rows(values_only=True):
    if any(c is not None and str(c).strip() != '' for c in row):
        rows_data.append(row)
    if len(rows_data) >= 12:
        break
for i, r in enumerate(rows_data):
    print(f'Row {i+1}: {r}')
wb.close()

wb2 = openpyxl.load_workbook(r'C:\Users\SupplyDestileria\Documents\Bosque\Claude\Tablero operativo\Salidas\Bosque salidas.xlsx', read_only=True, data_only=True)
ws2 = wb2.active
count = 0
deps = set()
last_row = None
dates_s = []
for row in ws2.iter_rows(values_only=True):
    c0 = str(row[0]).strip() if row[0] is not None else ''
    if c0 and c0 not in ('FECHA','fecha'):
        count += 1
        last_row = row
        if row[4] is not None and str(row[4]).strip():
            deps.add(str(row[4]).strip())
        if len(dates_s) < 3: dates_s.append(c0)
wb2.close()
print(f'Total rows: {count}, Last: {last_row}, Deps: {sorted(deps)}, Dates: {dates_s}')

print("\n=== PRODUCTOS.xlsx ===")
wb3 = openpyxl.load_workbook(r'C:\Users\SupplyDestileria\Documents\Bosque\Claude\Tablero operativo\Productos\PRODUCTOS.xlsx', read_only=True, data_only=True)
ws3 = wb3.active
for i, row in enumerate(ws3.iter_rows(values_only=True)):
    if any(c is not None and str(c).strip() != '' for c in row):
        print(f'Row {i+1}: {row}')
    if i >= 7: break
wb3.close()
