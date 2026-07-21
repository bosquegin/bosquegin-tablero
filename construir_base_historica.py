#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
construir_base_historica.py

Consolida TODA la información histórica del tablero en una única base de
datos SQLite (Data/Base_historica.db), para poder consultarla sin abrir
Excels ni releer Google Sheets cada vez.

Fuentes y tablas:
  productos          <- Data/Productos/PRODUCTOS.xlsx
  ventas             <- Data/Salidas/Salidas_consolidado.xlsx        (~33k filas)
  stock_diario       <- Data/Inventario/Stock_consolidado_por_deposito_y_dia.xlsx
  stock_cierre_mes   <- data_stock_cierre.js  (stock KLOZER+OFI al cierre de cada mes)
  costos             <- data_costos.js         (snapshot actual por producto)
  costos_anual       <- data_costos.js         (costo/pvp/mb/cmv por año)
  costos_mensual     <- data_costos.js         (costo/pvp/mb por mes)
  proyeccion         <- data_proyeccion.js  (7 trimestres, valores YA CORREGIDOS/publicados)
  proyeccion_mensual <- idem, desglose mes a mes por producto

Es idempotente: cada corrida reconstruye la base desde cero a partir de las
fuentes. Correr con:  python_embed\python.exe construir_base_historica.py
"""
import os, re, json, sqlite3
from datetime import datetime, date

BASE     = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE, "Data")
DB_PATH  = os.path.join(DATA_DIR, "Base_historica.db")


def _openpyxl():
    try:
        import openpyxl
        return openpyxl
    except ImportError:
        import subprocess, sys
        subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
        import openpyxl
        return openpyxl


def _fecha_iso(v):
    """Normaliza una fecha (datetime o texto) a 'YYYY-MM-DD', o None."""
    if v is None:
        return None
    if hasattr(v, "year"):                       # datetime/date
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    m = re.match(r"(\d{4})[-/](\d{1,2})[-/](\d{1,2})", s)
    if m:
        return "%04d-%02d-%02d" % (int(m.group(1)), int(m.group(2)), int(m.group(3)))
    m = re.match(r"(\d{1,2})[-/](\d{1,2})[-/](\d{4})", s)   # dd/mm/yyyy
    if m:
        return "%04d-%02d-%02d" % (int(m.group(3)), int(m.group(2)), int(m.group(1)))
    return s[:10] if len(s) >= 10 else None


def _num(v):
    if v is None:
        return None
    try:
        return float(str(v).replace(",", "."))
    except Exception:
        return None


def _cod(v):
    if v is None:
        return None
    s = str(v).strip()
    try:
        return str(int(float(s)))
    except Exception:
        return s or None


def _leer_costos():
    path = os.path.join(BASE, "data_costos.js")
    if not os.path.exists(path):
        return {}
    txt = open(path, encoding="utf-8").read()
    m = re.search(r"=\s*(\{.*\});?\s*$", txt, re.DOTALL)
    return json.loads(m.group(1)) if m else {}


def _leer_stock_cierre():
    path = os.path.join(BASE, "data_stock_cierre.js")
    if not os.path.exists(path):
        return {}
    js = open(path, encoding="utf-8").read()
    out = {}
    for mes, body in re.findall(r'"(\d{4}_\d{1,2})"\s*:\s*(\{[^}]*\})', js):
        out[mes] = {c: int(q) for c, q in re.findall(r'"(\d+)"\s*:\s*(-?\d+)', body)}
    return out


def _leer_proyeccion():
    # data_proyeccion.js (no el caché Analisis_proyeccion_cache.json): es lo
    # PUBLICADO, con los overrides del mes en curso ya aplicados (venta real,
    # saldo_stock correcto). El caché es el forecast crudo sin esos ajustes.
    path = os.path.join(BASE, "data_proyeccion.js")
    if not os.path.exists(path):
        return {}
    txt = open(path, encoding="utf-8").read()
    m = re.search(r"=\s*(\{.*\});?\s*$", txt, re.DOTALL)
    return json.loads(m.group(1)).get("trimestres", {}) if m else {}


def construir():
    if os.path.exists(DB_PATH):
        os.remove(DB_PATH)
    con = sqlite3.connect(DB_PATH)
    cur = con.cursor()
    openpyxl = _openpyxl()
    resumen = []

    def _rows(tabla):
        return cur.execute(f"SELECT COUNT(*) FROM {tabla}").fetchone()[0]

    # ── productos ──────────────────────────────────────────────────────────
    cur.execute("""CREATE TABLE productos (
        codigo TEXT PRIMARY KEY, rubro TEXT, sub_rubro TEXT,
        nombre TEXT, especificacion TEXT)""")
    try:
        p = os.path.join(DATA_DIR, "Productos", "PRODUCTOS.xlsx")
        wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
        ws = wb.active
        vistos = set()
        for r in ws.iter_rows(values_only=True):
            if not r or r[2] is None:
                continue
            if str(r[0]).strip().upper() == "RUBRO":      # fila de encabezado
                continue
            cod = _cod(r[2])
            if not cod or cod in vistos:
                continue
            vistos.add(cod)
            cur.execute("INSERT INTO productos VALUES (?,?,?,?,?)",
                        (cod, r[0], r[1], r[3], r[4] if len(r) > 4 else None))
        wb.close()
        resumen.append(("productos", _rows("productos"), "PRODUCTOS.xlsx"))
    except Exception as e:
        resumen.append(("productos", "ERROR", str(e)[:80]))

    # ── ventas (salidas) ───────────────────────────────────────────────────
    cur.execute("""CREATE TABLE ventas (
        fecha TEXT, razon_social TEXT, codigo TEXT, cantidad REAL,
        deposito TEXT, nombre_fantasia TEXT, fuente TEXT)""")
    try:
        p = os.path.join(DATA_DIR, "Salidas", "Salidas_consolidado.xlsx")
        wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
        ws = wb["SALIDAS"] if "SALIDAS" in wb.sheetnames else wb.active
        it = ws.iter_rows(values_only=True)
        next(it, None)                                    # encabezado
        buf = []
        for r in it:
            if not r or r[0] is None or r[2] is None:
                continue
            buf.append((_fecha_iso(r[0]), r[1], _cod(r[2]), _num(r[3]),
                        r[4], r[5] if len(r) > 5 else None, r[6] if len(r) > 6 else None))
            if len(buf) >= 5000:
                cur.executemany("INSERT INTO ventas VALUES (?,?,?,?,?,?,?)", buf); buf = []
        if buf:
            cur.executemany("INSERT INTO ventas VALUES (?,?,?,?,?,?,?)", buf)
        wb.close()
        resumen.append(("ventas", _rows("ventas"), "Salidas_consolidado.xlsx"))
    except Exception as e:
        resumen.append(("ventas", "ERROR", str(e)[:80]))

    # ── stock_diario ───────────────────────────────────────────────────────
    cur.execute("""CREATE TABLE stock_diario (
        fecha TEXT, deposito TEXT, articulo TEXT, codigo TEXT,
        proveedor TEXT, cantidad REAL, rubro TEXT, subrubro TEXT)""")
    try:
        p = os.path.join(DATA_DIR, "Inventario", "Stock_consolidado_por_deposito_y_dia.xlsx")
        wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
        ws = wb.active
        buf = []
        for r in ws.iter_rows(values_only=True):
            if not r or r[0] is None or str(r[0]).strip().lower() == "fecha":
                continue
            buf.append((_fecha_iso(r[0]), r[1], r[2], _cod(r[3]),
                        r[4] if len(r) > 4 else None, _num(r[5]) if len(r) > 5 else None,
                        r[6] if len(r) > 6 else None, r[7] if len(r) > 7 else None))
            if len(buf) >= 5000:
                cur.executemany("INSERT INTO stock_diario VALUES (?,?,?,?,?,?,?,?)", buf); buf = []
        if buf:
            cur.executemany("INSERT INTO stock_diario VALUES (?,?,?,?,?,?,?,?)", buf)
        wb.close()
        resumen.append(("stock_diario", _rows("stock_diario"), "Stock_consolidado_por_deposito_y_dia.xlsx"))
    except Exception as e:
        resumen.append(("stock_diario", "ERROR", str(e)[:80]))

    # ── stock_cierre_mes ───────────────────────────────────────────────────
    cur.execute("""CREATE TABLE stock_cierre_mes (
        mes TEXT, codigo TEXT, cantidad REAL)""")
    try:
        for mes, prods in _leer_stock_cierre().items():
            cur.executemany("INSERT INTO stock_cierre_mes VALUES (?,?,?)",
                            [(mes, c, q) for c, q in prods.items()])
        resumen.append(("stock_cierre_mes", _rows("stock_cierre_mes"), "data_stock_cierre.js"))
    except Exception as e:
        resumen.append(("stock_cierre_mes", "ERROR", str(e)[:80]))

    # ── costos (snapshot + anual + mensual) ────────────────────────────────
    cur.execute("""CREATE TABLE costos (
        codigo TEXT PRIMARY KEY, desc TEXT, rubro TEXT, sub TEXT, linea TEXT,
        estado TEXT, costo REAL, pvp REAL, mb REAL)""")
    cur.execute("""CREATE TABLE costos_anual (
        codigo TEXT, anio TEXT, costo REAL, pvp REAL, mb REAL, cmv REAL)""")
    cur.execute("""CREATE TABLE costos_mensual (
        codigo TEXT, mes TEXT, costo REAL, pvp REAL, mb REAL)""")
    try:
        # data_costos.js repite el mismo código varias veces (17 códigos con
        # duplicados; ej. 100001 aparece 5x). Se deduplica quedándose con la
        # entrada más completa por código: la que tiene costo no nulo y más
        # meses cargados. Si no se dedupe, costos_mensual/anual quedan inflados.
        por_cod = {}
        for pr in _leer_costos().get("productos", []):
            cod = _cod(pr.get("cod"))
            if cod:
                por_cod.setdefault(cod, []).append(pr)

        def _completitud(pr):
            return (pr.get("costo") is not None, len(pr.get("meses") or {}), len(pr.get("periodos") or {}))

        for cod, candidatos in por_cod.items():
            pr = max(candidatos, key=_completitud)   # el más completo
            cur.execute("INSERT OR REPLACE INTO costos VALUES (?,?,?,?,?,?,?,?,?)",
                        (cod, pr.get("desc"), pr.get("rubro"), pr.get("sub"), pr.get("linea"),
                         pr.get("estado"), pr.get("costo"), pr.get("pvp"), pr.get("mb")))
            for anio, d in (pr.get("periodos") or {}).items():
                cur.execute("INSERT INTO costos_anual VALUES (?,?,?,?,?,?)",
                            (cod, anio, d.get("costo"), d.get("pvp"), d.get("mb"), d.get("cmv")))
            for mes, d in (pr.get("meses") or {}).items():
                cur.execute("INSERT INTO costos_mensual VALUES (?,?,?,?,?)",
                            (cod, mes, d.get("costo"), d.get("pvp"), d.get("mb")))
        resumen.append(("costos",         _rows("costos"),         "data_costos.js"))
        resumen.append(("costos_anual",   _rows("costos_anual"),   "data_costos.js"))
        resumen.append(("costos_mensual", _rows("costos_mensual"), "data_costos.js"))
    except Exception as e:
        resumen.append(("costos", "ERROR", str(e)[:80]))

    # ── proyeccion (+ mensual) ─────────────────────────────────────────────
    cur.execute("""CREATE TABLE proyeccion (
        trimestre TEXT, codigo TEXT, articulo TEXT, stock_actual REAL, stock_total REAL,
        venta_prom_mensual_anterior REAL, venta_prom_6m REAL, venta_prom_12m REAL,
        total_objetivo_ventas REAL, meses_stock REAL,
        alerta TEXT, comprar REAL, pallet REAL, cantidad_pallets REAL)""")
    cur.execute("""CREATE TABLE proyeccion_mensual (
        trimestre TEXT, codigo TEXT, mes_idx TEXT, mes_label TEXT,
        proyeccion_abastecimiento REAL, pendiente_retiro REAL, venta_objetivo REAL,
        venta_actual REAL, proyeccion_mensual REAL, objetivo_cumplido_pct REAL, saldo_stock REAL)""")
    try:
        for trimestre, T in _leer_proyeccion().items():
            for pr in T.get("productos", []):
                cod = _cod(pr.get("cod"))
                cur.execute("INSERT INTO proyeccion VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                            (trimestre, cod, pr.get("art"), pr.get("stock_actual"), pr.get("stock_total"),
                             pr.get("venta_prom_mensual_anterior"), pr.get("venta_prom_6m"), pr.get("venta_prom_12m"),
                             pr.get("total_objetivo_ventas"), pr.get("meses_stock"), pr.get("alerta"),
                             pr.get("comprar"), pr.get("pallet"), pr.get("cantidad_pallets")))
                for mk, m in (pr.get("mensual") or {}).items():
                    cur.execute("INSERT INTO proyeccion_mensual VALUES (?,?,?,?,?,?,?,?,?,?,?)",
                                (trimestre, cod, mk, m.get("label"), m.get("proyeccion_abastecimiento"),
                                 m.get("pendiente_retiro"), m.get("venta_objetivo"), m.get("venta_actual"),
                                 m.get("proyeccion_mensual"), m.get("objetivo_cumplido_pct"), m.get("saldo_stock")))
        resumen.append(("proyeccion",         _rows("proyeccion"),         "data_proyeccion.js"))
        resumen.append(("proyeccion_mensual", _rows("proyeccion_mensual"), "data_proyeccion.js"))
    except Exception as e:
        resumen.append(("proyeccion", "ERROR", str(e)[:80]))

    # ── índices para consultas rápidas ─────────────────────────────────────
    for idx in [
        "CREATE INDEX ix_ventas_cod ON ventas(codigo)",
        "CREATE INDEX ix_ventas_fecha ON ventas(fecha)",
        "CREATE INDEX ix_stock_cod ON stock_diario(codigo)",
        "CREATE INDEX ix_stock_fecha ON stock_diario(fecha)",
        "CREATE INDEX ix_costos_mensual_cod ON costos_mensual(codigo)",
        "CREATE INDEX ix_proy_cod ON proyeccion(codigo)",
        "CREATE INDEX ix_scm_mes ON stock_cierre_mes(mes)",
    ]:
        try:
            cur.execute(idx)
        except Exception:
            pass

    # ── metadatos (autodescriptiva) ────────────────────────────────────────
    cur.execute("CREATE TABLE _meta (clave TEXT PRIMARY KEY, valor TEXT)")
    cur.execute("INSERT INTO _meta VALUES ('generado', ?)", (datetime.now().strftime("%Y-%m-%d %H:%M"),))
    cur.execute("INSERT INTO _meta VALUES ('descripcion', ?)",
                ("Base histórica consolidada del tablero Bosque Gin. Reconstruir con construir_base_historica.py",))
    for tabla, filas, fuente in resumen:
        cur.execute("INSERT OR REPLACE INTO _meta VALUES (?, ?)", (f"tabla:{tabla}", f"{filas} filas <- {fuente}"))

    con.commit()
    con.close()
    return resumen


if __name__ == "__main__":
    import sys
    sys.stdout.reconfigure(encoding="utf-8")
    print("Construyendo base histórica ->", DB_PATH)
    resumen = construir()
    print("\n%-20s %12s   %s" % ("TABLA", "FILAS", "FUENTE"))
    print("-" * 72)
    for tabla, filas, fuente in resumen:
        print("%-20s %12s   %s" % (tabla, filas, fuente))
    mb = os.path.getsize(DB_PATH) / 1024 / 1024
    print("-" * 72)
    print("Base creada: %s  (%.1f MB)" % (DB_PATH, mb))
