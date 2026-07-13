#!/usr/bin/env python3
"""
actualizar_bosquegin.py  —  v3.1  (2026-07-13)
Reconstruye bosquegin_data.js leyendo:
  - Stock:  EN VIVO desde la API REST de Contabilium (contabilium_api.py)
  - Salidas: Salidas_consolidado.xlsx (Bosque salidas + GC histórico + Contabilium)
             + correcciones manuales (Salidas_consolidado correcciones.xlsx)
  - Destileria: conserva la seccion existente en bosquegin_data.js

Cambios v3.1 (2026-07-13):
  - Stock en vivo vía API REST de Contabilium (contabilium_api.py), filtra inactivos
  - Historial de salidas de Contabilium automatizado vía CDP (contabilium_downloader.py)
  - Correcciones manuales de Salidas_consolidado persisten a través de cada reconstrucción
  - Descarga de Gestión Cervecera deshabilitada (obsoleta, reemplazada por Contabilium)
  - Fix de encoding UTF-8 que crasheaba secciones silenciosamente (cervezas, etc.)
  - Guardado seguro de Salidas_consolidado.xlsx (reintentos si está abierto en Excel)
  - Insumos vía CDP (fallback cuando la hoja no está publicada públicamente)
  - Resumen de salud al final de cada corrida (_ok/_fail por paso)
  - CLIENT_DATA separado en data_clientes.js (data_ventas.js: 4.9MB -> 631KB)
  - Cache-busting del dashboard usa meta.generado en vez de Date.now()

Cambios v2.0:
  - Stock se toma automaticamente de GC (reemplaza archivos manuales diarios)
  - update_consolidado() reconoce formato GC Simplificado y GC Comp
"""
import os, json, re, glob, time, sys
from datetime import datetime, date, timedelta, timezone

# Evita que un print con tildes/emojis/flechas (p.ej. "→") crashee la corrida
# cuando la consola de Windows no está en UTF-8 (cp1252 por defecto).
try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

# Zona horaria Argentina (UTC-3, sin DST)
_AR = timezone(timedelta(hours=-3))

import platform
if platform.system() == "Windows":
    BASE = r"C:\Users\SupplyDestileria\Documents\Bosque\Claude\Tablero operativo"
else:
    # En cloud (Render/Linux): usar variable de entorno o directorio del script
    BASE = os.environ.get("BOSQUEGIN_BASE",
                          os.path.dirname(os.path.abspath(__file__)))

# Modo cloud: cuando True, omite el git push (lo maneja actualizar_cloud.py)
_SKIP_GIT_PUSH = False

DATA_DIR        = os.path.join(BASE, "Data")
INV_DIR         = os.path.join(DATA_DIR, "Inventario")
VENTAS_F        = os.path.join(DATA_DIR, "Salidas", "Bosque salidas.xlsx")
GC_SALIDAS_DIR       = os.path.join(DATA_DIR, "Salidas", "GC")
CONTABILIUM_DIR      = os.path.join(DATA_DIR, "Salidas", "Contabilium")
CONTABILIUM_STOCK_DIR = os.path.join(INV_DIR, "Contabilium")
CONTABILIUM_CUTOVER  = "2026-06-29"   # desde esta fecha, salidas y stock vienen de Contabilium
COSTOS_CSV      = os.path.join(DATA_DIR, "Costos y PVP", "Analisis de costos y PVP - COSTOS.csv")
INSUMOS_CSV     = os.path.join(DATA_DIR, "Insumos", "Stock insumos.csv")
SALIDAS_CONS    = os.path.join(DATA_DIR, "Salidas", "Salidas_consolidado.xlsx")
SALIDAS_CORR    = os.path.join(DATA_DIR, "Salidas", "Salidas_consolidado correcciones.xlsx")
SALIDAS_CONS_HEADERS = ["FECHA", "RAZON SOCIAL", "CODIGO", "CANTIDAD", "DEPOSITO", "NOMBRE FANTASIA", "FUENTE"]
INSUMOS_GVIZ    = (
    "https://docs.google.com/spreadsheets/d/"
    "16_Ri4Rspy2s3pno-zn61Yvo3y3USeELHAbO72Hjm6SM/"
    "gviz/tq?tqx=out:csv&gid=805030698"
)
OUT_JS          = os.path.join(BASE, "bosquegin_data.js")

MESES_NOM = ['','Ene','Feb','Mar','Abr','May','Jun','Jul','Ago','Sep','Oct','Nov','Dic']

def do_mg(m, c):
    return round((m - c) / m * 100, 1) if m > 0 else 0.0

def parse_file_date(fname):
    """
    Extrae fecha del nombre. Formatos soportados:
      - d-m-yyyy  o  d/m/yyyy  (ej: 8-5-2026)
      - yyyy-mm-dd              (ej: 2026-05-08)
    Devuelve string 'yyyy-mm-dd' o date.today() como fallback.
    """
    # Intenta d-m-yyyy (formato local argentino)
    m = re.search(r'(\d{1,2})[-./](\d{1,2})[-./](\d{4})', fname)
    if m:
        d_n, mo_n, yr_n = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if 1 <= mo_n <= 12 and 1 <= d_n <= 31 and 2020 <= yr_n <= 2030:
            return f"{yr_n}-{mo_n:02d}-{d_n:02d}"
    # Intenta yyyy-mm-dd
    m = re.search(r'(\d{4})[-./](\d{2})[-./](\d{2})', fname)
    if m:
        return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
    return date.today().strftime("%Y-%m-%d")


# ─── 1. ACTUALIZAR CONSOLIDADO ────────────────────────────────────────────────
CONS_FILE = os.path.join(INV_DIR, "Stock_consolidado_por_deposito_y_dia.xlsx")

def update_consolidado():
    """
    Agrega al archivo consolidado las filas de los archivos diarios
    que sean mas recientes que la ultima fecha ya registrada.
    """
    try:
        import openpyxl
    except ImportError:
        import subprocess, sys
        subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl",
                        "--break-system-packages", "-q"])
        import openpyxl

    # Leer ultima fecha, lookup y fechas ya presentes (para evitar duplicados)
    wb0 = openpyxl.load_workbook(CONS_FILE, read_only=True, data_only=True)
    ws0 = wb0.active
    lookup        = {}
    last_date     = "0"
    existing_keys = set()   # (fecha, dep, cod) ya en el consolidado
    for r in ws0.iter_rows(values_only=True):
        if not r or r[0] == "fecha": continue
        if r[3] is not None:
            try:
                cod = str(int(float(str(r[3]))))
                if cod not in lookup:
                    lookup[cod] = (r[6] or "", r[7] or "", r[2] or "")
                fecha_k = str(r[0])[:10] if r[0] else ""
                dep_k   = str(r[1]).strip().upper() if r[1] else ""
                if fecha_k and dep_k:
                    existing_keys.add((fecha_k, dep_k, cod))
            except Exception:
                pass
        if r[0]:
            d = str(r[0])[:10]
            if d > last_date: last_date = d
    wb0.close()
    print(f"    Consolidado al {last_date} | {len(lookup)} codigos conocidos")

    # Si el consolidado ya tiene datos de hoy, eliminarlos para poder refrescarlos
    today_str = date.today().strftime("%Y-%m-%d")
    if last_date >= today_str:
        print(f"    Eliminando filas de {today_str} para refrescar...")
        wb_r = openpyxl.load_workbook(CONS_FILE)
        ws_r = wb_r.active
        all_rows_r = list(ws_r.iter_rows(values_only=True))
        keep_rows  = [list(r) for r in all_rows_r
                      if not r or r[0] == "fecha" or
                      not r[0] or str(r[0])[:10] != today_str]
        ws_r.delete_rows(1, ws_r.max_row + 1)
        for r in keep_rows:
            ws_r.append(r)
        wb_r.save(CONS_FILE)
        wb_r.close()
        # Actualizar variables en memoria
        existing_keys = {k for k in existing_keys if k[0] != today_str}
        dates_left = [str(r[0])[:10] for r in keep_rows if r and r[0] and r[0] != "fecha"]
        last_date  = max(dates_left, default="0")
        print(f"    Consolidado reseteado al {last_date}")

    # Archivos diarios mas nuevos que last_date
    # Nota: en Windows glob es case-insensitive, por eso se filtra y deduplica
    _seen = set()
    all_daily = []
    for fp in (
        sorted(glob.glob(os.path.join(INV_DIR, "Stock productos*.xlsx"))) +
        sorted(glob.glob(os.path.join(INV_DIR, "Stock Productos Comp*.xlsx"))) +
        sorted(glob.glob(os.path.join(INV_DIR, "Stock de Productos Simplificado*.xlsx")))
    ):
        key = os.path.normcase(fp)
        if key not in _seen:
            _seen.add(key)
            all_daily.append(fp)
    new_files = []
    for fp in all_daily:
        fd = parse_file_date(os.path.basename(fp))
        if fd and fd > last_date:
            new_files.append((fd, fp))
    new_files.sort()

    if not new_files:
        print("    Consolidado ya esta al dia.")
        return last_date

    print(f"    Agregando {len(new_files)} archivo(s) nuevos...")
    new_rows = []

    for file_date, fpath in new_files:
        fname = os.path.basename(fpath)
        try:
            wb2 = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
        except Exception as e:
            print(f"      Salteo {fname}: {e}")
            continue

        # ── Formato GC Simplificado (Stock de Productos Simplificado*.xlsx) ──
        # Hoja "Produccion terminada", seccion "Stock de productos en envases"
        # Columnas: Deposito[0] | Producto[1] | Codigo[2] | Envase[3] | Caja[4] | Lote[5] | Cantidad[6]
        # Deposito en cada fila; multiples filas por lote → sumar por (dep, cod)
        if "Simplificado" in fname or "simplificado" in fname:
            try:
                ws2 = wb2["Producción terminada"]
            except KeyError:
                try:
                    ws2 = wb2["Produccion terminada"]
                except KeyError:
                    print(f"      Salteo {fname}: hoja no encontrada"); wb2.close(); continue
            rows = list(ws2.iter_rows(values_only=True))
            env_start = None
            for i, r in enumerate(rows):
                if r and r[0] and "envase" in str(r[0]).lower():
                    env_start = i; break
            if env_start is None:
                print(f"      Salteo {fname}: seccion envases no encontrada"); wb2.close(); continue
            # Pre-agregar: (dep, cod) -> [qty_total, art]
            agg = {}
            for row in rows[env_start + 2:]:   # +1 titulo seccion, +1 cabecera
                if not row or not row[0]: continue
                dep_raw = str(row[0]).strip().upper()
                art_raw = str(row[1]).strip() if row[1] else ""
                cod_raw = row[2]; qty_raw = row[6]
                if cod_raw is None or qty_raw is None: continue
                try: cod = str(int(float(str(cod_raw).strip())))
                except: continue
                try: qty = float(str(qty_raw).replace(",", "."))
                except: continue
                if qty <= 0: continue
                k = (dep_raw, cod)
                if k not in agg: agg[k] = [0.0, art_raw]
                agg[k][0] += qty
                if art_raw and len(art_raw) > len(agg[k][1]): agg[k][1] = art_raw
            cnt = 0
            for (dep_raw, cod), (qty, art) in agg.items():
                if not dep_raw or not cod: continue
                if art in ("nan", "None", ""): art = lookup.get(cod, ("","",""))[2]
                ri = lookup.get(cod, ("","",art))
                key = (file_date, dep_raw, cod)
                if key not in existing_keys:
                    new_rows.append((file_date, dep_raw, art or ri[2], int(cod), None, qty, ri[0], ri[1]))
                    existing_keys.add(key); cnt += 1
            print(f"      {fname}: {cnt} registros (GC simplificado)")
            wb2.close(); continue

        # ── Formato clásico / GC Comp ────────────────────────────────────────
        ws2 = wb2.active
        rows = list(ws2.iter_rows(values_only=True))

        def _is_dep_hdr(r):
            return r and r[0] and str(r[0]).strip().upper() in ("DEPOSITO", "DEP\xd3SITO", "DEPÓSITO")

        def _norm_sec(s):
            return (str(s).strip().lower()
                    .replace('ó','o').replace('é','e').replace('á','a').replace('í','i').replace('ú','u'))

        TARGET_SECS   = {"envases en depositos", "barriles en depositos"}
        BARRILES_ONLY = {"100017"}

        # Solo estas filas actúan como delimitadores — no los nombres de depósito
        MAJOR_SECS = {
            "stock de producto en tanques", "stock de producto en barriles",
            "stock de producto en envases", "stock de producto en envses",
            "barriles en depositos", "envases en depositos", "total",
        }

        major_starts = []
        for i, r in enumerate(rows):
            if not r or not r[0]: continue
            nm = _norm_sec(str(r[0]))
            if nm in MAJOR_SECS:
                major_starts.append((i, nm))
        major_starts.append((len(rows), ""))

        all_secs = {}
        for k in range(len(major_starts) - 1):
            si, nm = major_starts[k]
            all_secs[nm] = (si, major_starts[k + 1][0])

        has_targets = any(s in all_secs for s in TARGET_SECS)
        if has_targets:
            ranges_to_process = [(nm, s, e) for nm, (s, e) in all_secs.items() if nm in TARGET_SECS]
        else:
            # Fallback: buscar primer DEPOSITO header (archivos Comp u otros formatos)
            hdr_idx = None
            for i, r in enumerate(rows[:5]):
                if _is_dep_hdr(r): hdr_idx = i; break
            if hdr_idx is None and "Comp" in fname:
                for i, r in enumerate(rows):
                    if _is_dep_hdr(r): hdr_idx = i; break
            if hdr_idx is None:
                wb2.close(); continue
            ranges_to_process = [("_fallback", hdr_idx, len(rows))]

        cnt = 0
        for sec_nm, sec_start, sec_end in ranges_to_process:
            sec_rows = rows[sec_start:sec_end]
            current_dep = None

            # ── Barriles: col0=depósito/None, col2=nombre, col3=código, col4=cantidad ──
            if sec_nm == "barriles en depositos":
                for row in sec_rows[1:]:   # saltar fila de título de sección
                    if not row or not any(row): continue
                    if row[0] and str(row[0]).strip():
                        raw = str(row[0]).strip().upper()
                        if not re.match(r'^\d', raw):
                            current_dep = raw
                    cod_raw = row[3] if len(row) > 3 else None
                    qty_raw = row[4] if len(row) > 4 else None
                    if cod_raw is None or qty_raw is None: continue
                    try:
                        cod = str(int(float(str(cod_raw).strip())))
                    except Exception:
                        continue
                    if cod not in BARRILES_ONLY: continue
                    try:
                        qty = float(str(qty_raw).replace(",", "."))
                    except Exception:
                        continue
                    if qty <= 0: continue
                    art_raw = str(row[2]).strip() if len(row) > 2 and row[2] else ""
                    if art_raw in ("nan", "None", ""): art_raw = lookup.get(cod, ("", "", ""))[2]
                    ri = lookup.get(cod, ("", "", art_raw))
                    key = (file_date, current_dep, cod)
                    if key not in existing_keys:
                        new_rows.append((file_date, current_dep, art_raw or ri[2],
                                         int(cod), None, qty, ri[0], ri[1]))
                        existing_keys.add(key)
                        cnt += 1
                continue

            # ── Envases / Formato estándar / Comp ──
            hdr_local = None
            for i, r in enumerate(sec_rows):
                if _is_dep_hdr(r): hdr_local = i; break
            if hdr_local is None:
                continue
            hdr_str = " ".join(str(c).upper() if c else "" for c in sec_rows[hdr_local])
            is_main = "CANTIDAD" in hdr_str or "ENVASE" in hdr_str

            for row in sec_rows[hdr_local + 1:]:
                if not row: continue
                dep_cell = row[0] if len(row) > 0 else None
                if dep_cell and str(dep_cell).strip() not in ("", "None"):
                    raw = str(dep_cell).strip().upper()
                    if not re.match(r'^[WwLl]\d', raw) and not re.match(r'^\d+\.?\d*$', raw):
                        current_dep = raw

                if is_main:
                    cod_cell = row[2] if len(row) > 2 else None
                    qty_cell = row[6] if len(row) > 6 else None
                    art_cell = row[1] if len(row) > 1 else None
                    prv_cell = None
                else:
                    cod_cell = row[3] if len(row) > 3 else None
                    qty_cell = row[4] if len(row) > 4 else None
                    art_cell = row[1] if len(row) > 1 else None
                    prv_cell = row[2] if len(row) > 2 else None

                if cod_cell is None or qty_cell is None: continue
                try:
                    cod = str(int(float(str(cod_cell).strip())))
                except Exception:
                    continue
                try:
                    qty = float(str(qty_cell).replace(",", "."))
                except Exception:
                    continue
                if qty <= 0: continue

                art = str(art_cell).strip() if art_cell else ""
                if art in ("nan", "None", ""): art = lookup.get(cod, ("", "", ""))[2]
                prv = str(prv_cell).strip() if prv_cell else None
                if prv in ("nan", "None", ""): prv = None

                ri = lookup.get(cod, ("", "", art))
                rub = ri[0]; subrub = ri[1]
                if not art: art = ri[2]

                key = (file_date, current_dep, cod)
                if key not in existing_keys:
                    new_rows.append((file_date, current_dep, art, int(cod), prv, qty, rub, subrub))
                    existing_keys.add(key)
                    cnt += 1

        sec_label = ", ".join(nm for nm in all_secs if nm in TARGET_SECS) if has_targets else "todo el archivo"
        print(f"      {fname}: {cnt} registros ({sec_label})")
        wb2.close()

    if new_rows:
        wb3 = openpyxl.load_workbook(CONS_FILE)
        ws3 = wb3.active
        for row in new_rows:
            ws3.append(list(row))
        wb3.save(CONS_FILE)
        wb3.close()
        print(f"    +{len(new_rows)} filas agregadas al consolidado.")
        return max(r[0] for r in new_rows)
    return last_date


# ─── 2. STOCK desde consolidado ───────────────────────────────────────────────
def parse_stock(inv_dir):
    """
    Lee el archivo Stock_consolidado_por_deposito_y_dia.xlsx.
    Para cada producto toma el stock del dia mas reciente por deposito.
    Columnas: fecha | deposito | articulo | codigo | proveedor | cantidad | rubro | subrubro
    Devuelve (inv_data_list, last_date_str).
    """
    try:
        import openpyxl
    except ImportError:
        import subprocess, sys
        subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl",
                        "--break-system-packages", "-q"])
        import openpyxl

    K_DEPS  = {"KLOZER"}
    KM_DEPS = {"KLOZER MKT"}
    O_DEPS  = {"OFI", "OFICINA"}
    SG_DEPS = {"SHOP GALLERY"}
    AV_DEPS = {"AVOLTA"}

    prod_info = {}   # cod -> {art, rub, sub}
    all_rows  = []   # lista de (file_date, dep_grp, cod, qty)
    dep_max   = {}   # dep_grp -> fecha maxima disponible

    wb = openpyxl.load_workbook(CONS_FILE, read_only=True, data_only=True)
    ws = wb.active

    for row in ws.iter_rows(values_only=True):
        if not row or row[0] == "fecha": continue
        # Columnas: fecha(0) deposito(1) articulo(2) codigo(3) proveedor(4) cantidad(5) rubro(6) subrubro(7)
        fecha_raw = row[0]
        dep_raw   = str(row[1]).strip().upper() if row[1] else ""
        art_raw   = str(row[2]).strip()         if row[2] else ""
        cod_raw   = row[3]
        qty_raw   = row[5]
        rub_raw   = str(row[6]).strip()         if row[6] else ""
        sub_raw   = str(row[7]).strip()         if row[7] else ""

        if cod_raw is None or qty_raw is None: continue
        try:
            cod = str(int(float(str(cod_raw))))
        except Exception:
            continue
        try:
            qty = float(str(qty_raw).replace(",", "."))
        except Exception:
            continue
        if qty < 0: qty = 0

        file_date = str(fecha_raw)[:10] if fecha_raw else ""
        if not file_date: continue

        if cod not in prod_info:
            prod_info[cod] = {"art": art_raw, "rub": rub_raw or "OTROS", "sub": sub_raw}
        elif art_raw and len(art_raw) > len(prod_info[cod]["art"]):
            prod_info[cod]["art"] = art_raw

        dep_grp = None
        if dep_raw in K_DEPS:    dep_grp = "klozer"
        elif dep_raw in KM_DEPS: dep_grp = "klozer_mkt"
        elif dep_raw in O_DEPS:  dep_grp = "ofi"
        elif dep_raw in SG_DEPS: dep_grp = "shop_gallery"
        elif dep_raw in AV_DEPS: dep_grp = "avolta"

        if dep_grp:
            all_rows.append((file_date, dep_grp, cod, qty))
            if dep_grp not in dep_max or file_date > dep_max[dep_grp]:
                dep_max[dep_grp] = file_date

    wb.close()

    # Solo incluir productos del archivo mas reciente de cada deposito.
    # Se compara file_date contra dep_max[dep_grp] (no contra hoy),
    # así funciona aunque hoy no haya descarga nueva todavía.
    from datetime import date as _dt
    def _gap(file_date, ref_date):
        try: return (_dt.fromisoformat(ref_date) - _dt.fromisoformat(file_date)).days
        except: return 9999

    prod_latest = {}  # (dep_grp, cod) -> fecha
    prod_best   = {}  # cod -> {dep_grp -> qty}
    for file_date, dep_grp, cod, qty in all_rows:
        ref = dep_max.get(dep_grp, file_date)
        if _gap(file_date, ref) > 0:
            continue                # no es el archivo mas reciente de este deposito
        key = (dep_grp, cod)
        prev = prod_latest.get(key, "")
        if file_date < prev: continue
        if file_date == prev:
            prod_best.setdefault(cod, {})[dep_grp] = max(
                prod_best.get(cod, {}).get(dep_grp, 0), qty)
        else:
            prod_latest[key] = file_date
            prod_best.setdefault(cod, {})[dep_grp] = qty

    print(f"    Fecha max por deposito: { {k: v for k,v in sorted(dep_max.items())} }")

    # Build INV_DATA (sin velocidad — se aplica despues desde ventas)
    # Construir lista de todos los codigos conocidos (con o sin stock hoy)
    all_cods = set(prod_info.keys())
    # También incluir productos que aparecieron en prod_best
    all_cods |= set(prod_best.keys())

    inv_data = []
    for cod in all_cods:
        deps    = prod_best.get(cod, {})
        klozer       = round(deps.get("klozer",       0))
        klozer_mkt   = round(deps.get("klozer_mkt",   0))
        ofi          = round(deps.get("ofi",           0))
        shop_gallery = round(deps.get("shop_gallery",  0))
        avolta       = round(deps.get("avolta",        0))
        both         = klozer + klozer_mkt + ofi + shop_gallery + avolta

        # Ignorar articulos sin stock y sin codigo numerico
        if klozer == 0 and klozer_mkt == 0 and ofi == 0 and shop_gallery == 0 and avolta == 0:
            try:
                int(cod)
            except (ValueError, TypeError):
                continue

        # Fecha del deposito = la fecha maxima de ese deposito (si tiene datos)
        k_fecha  = dep_max.get("klozer",       "") if klozer       > 0 else ""
        km_fecha = dep_max.get("klozer_mkt",   "") if klozer_mkt   > 0 else ""
        o_fecha  = dep_max.get("ofi",          "") if ofi          > 0 else ""
        sg_fecha = dep_max.get("shop_gallery", "") if shop_gallery > 0 else ""
        av_fecha = dep_max.get("avolta",       "") if avolta       > 0 else ""

        info = prod_info.get(cod, {})
        inv_data.append({
            "cod": cod, "art": info.get("art", ""), "rub": info.get("rub", ""),
            "sub": info.get("sub", ""),
            "klozer": klozer, "klozer_mkt": klozer_mkt, "ofi": ofi,
            "shop_gallery": shop_gallery, "avolta": avolta, "both": both,
            "k_fecha": k_fecha, "km_fecha": km_fecha, "o_fecha": o_fecha,
            "sg_fecha": sg_fecha, "av_fecha": av_fecha,
            # velocidad se rellena luego
            "k90": 0.0, "k365": 0.0, "min_k": 0.0, "min_k90": 0.0, "min_k365": 0.0,
            "o90": 0.0, "o365": 0.0, "min_o": 0.0, "min_o90": 0.0, "min_o365": 0.0,
            "b90": 0.0, "b365": 0.0, "min_b": 0.0,
            "pk": 0.0, "pk90": 0.0, "pk365": 0.0, "pk_dep": 0.0, "pk_mkt": 0.0, "pk_mkt90": 0.0, "pk_mkt365": 0.0,
            "po": 0.0, "po90": 0.0, "po365": 0.0, "pb": 0.0,
        })

    stock_hasta = max(dep_max.values()) if dep_max else date.today().strftime("%Y-%m-%d")
    inv_data.sort(key=lambda x: x["art"])
    return inv_data, stock_hasta


# Mapeo nombre de depósito en Contabilium -> clave de dep_grp usada en INV_DATA
_CONT_DEP_GRP = {
    "KLOZER":       "klozer",
    "KLOZER MKT":   "klozer_mkt",
    "OFICINA":      "ofi",
    "SHOP GALLERY": "shop_gallery",
    "AVOLTA":       "avolta",
}


def apply_stock_contabilium_vivo(inv_data, inv_gen=None):
    """
    Sobreescribe el stock de INV_DATA con los valores EN VIVO de la API de
    Contabilium para los depósitos mapeados en _CONT_DEP_GRP. Agrega productos
    nuevos que no estaban en el consolidado. Devuelve (inv_data, stock_hasta).
    """
    import contabilium_api as _cont_api
    inv_gen = inv_gen or {}
    hoy = date.today().strftime("%Y-%m-%d")

    stock_vivo = _cont_api.get_stock_todos_depositos(list(_CONT_DEP_GRP.keys()))
    activos = _cont_api.get_active_codigos()

    # cod -> {dep_grp: qty}   (solo productos activos en Contabilium)
    por_cod = {}
    for dep_nombre, items in stock_vivo.items():
        dep_grp = _CONT_DEP_GRP.get(dep_nombre)
        if not dep_grp: continue
        for cod, qty in items.items():
            if cod not in activos: continue
            por_cod.setdefault(cod, {})[dep_grp] = qty

    # Sacar del inventario cualquier producto que Contabilium marca inactivo
    inv_data = [it for it in inv_data if it["cod"] in activos]

    by_cod_idx = {item["cod"]: item for item in inv_data}

    for cod, deps_qty in por_cod.items():
        item = by_cod_idx.get(cod)
        if item is None:
            if not any(q > 0 for q in deps_qty.values()):
                continue   # no crear filas nuevas para conceptos/servicios sin stock
            info = inv_gen.get(cod, {})
            item = {
                "cod": cod, "art": info.get("art", ""), "rub": info.get("rub", "OTROS"),
                "sub": info.get("sub", ""),
                "klozer": 0, "klozer_mkt": 0, "ofi": 0, "shop_gallery": 0, "avolta": 0, "both": 0,
                "k_fecha": "", "km_fecha": "", "o_fecha": "", "sg_fecha": "", "av_fecha": "",
                "k90": 0.0, "k365": 0.0, "min_k": 0.0, "min_k90": 0.0, "min_k365": 0.0,
                "o90": 0.0, "o365": 0.0, "min_o": 0.0, "min_o90": 0.0, "min_o365": 0.0,
                "b90": 0.0, "b365": 0.0, "min_b": 0.0,
                "pk": 0.0, "pk90": 0.0, "pk365": 0.0, "pk_dep": 0.0, "pk_mkt": 0.0, "pk_mkt90": 0.0, "pk_mkt365": 0.0,
                "po": 0.0, "po90": 0.0, "po365": 0.0, "pb": 0.0,
            }
            inv_data.append(item)
            by_cod_idx[cod] = item

        for dep_grp, qty in deps_qty.items():
            item[dep_grp] = round(qty)
            fecha_key = {"klozer": "k_fecha", "klozer_mkt": "km_fecha", "ofi": "o_fecha",
                         "shop_gallery": "sg_fecha", "avolta": "av_fecha"}[dep_grp]
            item[fecha_key] = hoy

        item["both"] = (item["klozer"] + item["klozer_mkt"] + item["ofi"]
                         + item["shop_gallery"] + item["avolta"])

    inv_data.sort(key=lambda x: x["art"])
    return inv_data, hoy


# ─── 2. VENTAS ────────────────────────────────────────────────────────────────
DEP_MAP = {
    "KLOZER": "KLOZER", "KLOZER MKT": "KLOZER_MKT",
    "OFI": "OFICINA", "OFICINA": "OFICINA",
    "AVOLTA": "AVOLTA", "FILIDORO": "FILIDORO",
    "SHOP GALLERY": "SHOP GALLERY",
    "DIRECTORO EXPORTACIONES": "DIRECTORO EXPORTACIONES",
    "SUR BONIFICACIONES": "SUR BONIFICACIONES",
    "BIERHAUS (FRIO)": "BIERHAUS",
    "QUIRINO": "QUIRINO",
    "POLDITOS/MERLE": "POLDITOS/MERLE",
}


def _last_ventas_date():
    """Escanea Bosque salidas.xlsx y retorna la fecha maxima con datos reales
    (cod + cantidad no nulos) como string 'YYYY-MM-DD', o None.
    Ignora filas con solo fecha (marcadores de posición vacíos)."""
    if not os.path.exists(VENTAS_F):
        return None
    try:
        import openpyxl
        wb   = openpyxl.load_workbook(VENTAS_F, read_only=True, data_only=True)
        ws   = wb["SALIDAS"] if "SALIDAS" in wb.sheetnames else wb.active
        last = "0"
        for row in ws.iter_rows(min_row=3, values_only=True):
            if not row or not row[0]: continue
            # Ignorar filas sin código o sin cantidad (son marcadores de posición)
            cod_raw = row[2] if len(row) > 2 else None
            qty_raw = row[3] if len(row) > 3 else None
            if cod_raw is None or qty_raw is None: continue
            try:
                qty = float(str(qty_raw).replace(",", "."))
                if qty <= 0: continue
            except: continue
            raw = row[0]
            if hasattr(raw, "year"):
                d = f"{raw.year}-{raw.month:02d}-{raw.day:02d}"
            else:
                s = str(raw).strip()
                parts = s.replace("/", "-").split("-")
                try:
                    if len(parts) == 3:
                        if len(parts[0]) == 4:
                            d = f"{parts[0]}-{int(parts[1]):02d}-{int(parts[2]):02d}"
                        else:
                            d = f"{parts[2]}-{int(parts[1]):02d}-{int(parts[0]):02d}"
                    else: continue
                except: continue
            if d > last: last = d
        wb.close()
        return last if last != "0" else None
    except Exception:
        return None


def _aplicar_correcciones_salidas(rows_out):
    """
    Sobreescribe filas de rows_out con los valores de
    Data/Salidas/Salidas_consolidado correcciones.xlsx (hoja "Correcciones"),
    matcheando por (FECHA exacta, CODIGO). Se llama en cada corrida de
    update_salidas_consolidado() ANTES de guardar — así las correcciones
    sobreviven a la reconstrucción completa del consolidado en cada Actualizar.
    Columnas esperadas: FECHA | RAZON SOCIAL | CODIGO | CANTIDAD | DEPOSITO | NOMBRE FANTASIA
    """
    if not os.path.exists(SALIDAS_CORR):
        return 0
    try:
        import openpyxl
        wb_c = openpyxl.load_workbook(SALIDAS_CORR, read_only=True, data_only=True)
        ws_c = wb_c["Correcciones"] if "Correcciones" in wb_c.sheetnames else wb_c.active
        corr_rows = list(ws_c.iter_rows(values_only=True))[1:]
        wb_c.close()
    except Exception as e:
        print(f"    Advertencia leyendo correcciones: {e}")
        return 0

    # Índice rows_out por (fecha, codigo) -> lista de indices
    idx = {}
    for i, r in enumerate(rows_out):
        idx.setdefault((r[0], str(r[2]).strip()), []).append(i)

    aplicadas = 0
    for cr in corr_rows:
        if not cr or cr[0] is None: continue
        fecha, rs, cod, cant, dep, fan = cr[0], cr[1], cr[2], cr[3], cr[4], cr[5]
        cod = str(cod).strip()
        for i in idx.get((fecha, cod), []):
            r = rows_out[i]
            if r[1] != (rs or "") or r[3] != cant or r[4] != dep or r[5] != (fan or ""):
                rows_out[i] = [fecha, rs or "", cod, cant, dep, fan or "", r[6]]
                aplicadas += 1
    return aplicadas


def update_salidas_consolidado():
    """
    Reconstruye Data/Salidas/Salidas_consolidado.xlsx combinando:
      1. Todas las filas de Bosque salidas.xlsx  (fuente manual, NOMBRE FANTASIA en MAYUSCULAS)
      2. Filas de Remitos GC con fecha posterior a la ultima fecha de Bosque salidas
         (GC cubre el periodo reciente que el archivo manual todavia no tiene)
    El resultado es la unica fuente que lee parse_ventas().
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        import subprocess, sys
        subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl",
                        "--break-system-packages", "-q"])
        import openpyxl
        from openpyxl.styles import Font, PatternFill

    def _fantasia(fan_raw, rs_raw):
        f = str(fan_raw).strip().upper() if fan_raw not in (None, "", "nan", "None") else ""
        if not f or f in ("NAN", "NONE"):
            f = str(rs_raw).strip().upper() if rs_raw not in (None, "", "nan", "None") else "SIN CLIENTE"
        if not f or f in ("NAN", "NONE", ""):
            f = "SIN CLIENTE"
        return f

    def _fantasia_obs(obs):
        """Extrae nombre de cliente del campo Observaciones de Contabilium.
        'FCA 00000012 LOS TEMPLOS S.R.L.' -> 'LOS TEMPLOS S.R.L.'
        """
        if not obs:
            return "SIN CLIENTE"
        s = str(obs).strip()
        import re as _re
        m = _re.match(r'^FCA\s+\S+\s+(.+)', s, _re.IGNORECASE)
        if m:
            return m.group(1).strip().upper()
        return "SIN CLIENTE"

    def _parse_fecha(raw):
        if hasattr(raw, "year"):
            return raw, f"{raw.year}-{raw.month:02d}-{raw.day:02d}"
        s = str(raw).strip()
        parts = s.replace("/", "-").split("-")
        if len(parts) != 3:
            return None, None
        try:
            if len(parts[0]) == 4:
                yr, mo, d = int(parts[0]), int(parts[1]), int(parts[2])
            else:
                yr, mo, d = int(parts[2]), int(parts[1]), int(parts[0])
            if not (2020 <= yr <= 2030 and 1 <= mo <= 12 and 1 <= d <= 31):
                return None, None
            return date(yr, mo, d), f"{yr}-{mo:02d}-{d:02d}"
        except Exception:
            return None, None

    rows_out = []
    last_ventas_date = "0"

    # ─ 1. Bosque salidas.xlsx ────────────────────────────────────────────────
    if os.path.exists(VENTAS_F):
        wb = openpyxl.load_workbook(VENTAS_F, read_only=True, data_only=True)
        ws = wb["SALIDAS"] if "SALIDAS" in wb.sheetnames else wb.active
        it = ws.iter_rows(values_only=True)
        next(it, None); next(it, None)   # saltar titulo + cabecera
        count = 0
        for row in it:
            if not row or not any(row): continue
            fecha_dt, fecha_str = _parse_fecha(row[0] if len(row) > 0 else None)
            if fecha_dt is None: continue
            cod_raw = row[2] if len(row) > 2 else None
            qty_raw = row[3] if len(row) > 3 else None
            if cod_raw is None or qty_raw is None: continue
            try: qty = float(str(qty_raw).replace(",", "."))
            except: continue
            if qty <= 0: continue
            try: cod = str(int(float(str(cod_raw).strip())))
            except: cod = str(cod_raw).strip()
            if not cod or cod in ("nan", "None", ""): continue
            rs_raw  = row[1] if len(row) > 1 else None
            dep_raw = row[4] if len(row) > 4 else None
            fan_raw = row[5] if len(row) > 5 else None
            rs      = str(rs_raw).strip().upper()  if rs_raw  not in (None, "", "nan", "None") else ""
            dep     = str(dep_raw).strip().upper() if dep_raw not in (None, "", "nan", "None") else ""
            fantasia = _fantasia(fan_raw, rs_raw)
            if fecha_str > last_ventas_date:
                last_ventas_date = fecha_str
            rows_out.append([fecha_dt, rs, cod, qty, dep, fantasia, "BOSQUE_SALIDAS"])
            count += 1
        wb.close()
        print(f"    Bosque salidas.xlsx: {count} filas (hasta {last_ventas_date})")
    else:
        print(f"    Advertencia: no se encontro {VENTAS_F}")

    # ─ 2. Remitos GC (solo fechas > last_ventas_date) ───────────────────────
    gc_files = sorted(glob.glob(os.path.join(GC_SALIDAS_DIR, "Remitos GC*.xlsx")))
    gc_count = 0
    for fpath in gc_files:
        fname = os.path.basename(fpath)
        try:
            wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
        except Exception as e:
            print(f"    Salteo {fname}: {e}"); continue
        ws   = wb["Datos"] if "Datos" in wb.sheetnames else wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if len(rows) < 2: continue
        for row in rows[1:]:
            if not row or len(row) < 19: continue
            fecha_dt, fecha_str = _parse_fecha(row[3])
            if fecha_dt is None: continue
            if fecha_str <= last_ventas_date: continue   # ya cubierto por Bosque salidas
            qty_raw = row[9]; cod_raw = row[16]; dep_raw = row[18]
            if qty_raw is None or cod_raw is None: continue
            try: qty = float(str(qty_raw).replace(",", "."))
            except: continue
            if qty <= 0: continue
            try: cod = str(int(float(str(cod_raw).strip())))
            except: cod = str(cod_raw).strip()
            if not cod or cod in ("nan", "None", ""): continue
            dep_str = str(dep_raw).strip().upper() if dep_raw else ""
            dep = DEP_MAP.get(dep_str, dep_str or "SIN DEPOSITO")
            # Nombre de fantasia y razon social desde GC (col 4 y 5)
            fan_raw = row[4] if len(row) > 4 else None
            rs_raw  = row[5] if len(row) > 5 else None
            fantasia = _fantasia(fan_raw, rs_raw)
            rs = str(rs_raw).strip().upper() if rs_raw not in (None, "", "nan", "None") else ""
            rows_out.append([fecha_dt, rs, cod, qty, dep, fantasia, "GC_REMITO"])
            gc_count += 1
    print(f"    Remitos GC: {gc_count} filas nuevas (despues de {last_ventas_date})")

    # ─ 3. Contabilium (desde CONTABILIUM_CUTOVER, reemplaza GC) ─────────────
    os.makedirs(CONTABILIUM_DIR, exist_ok=True)
    cont_files = sorted(glob.glob(os.path.join(CONTABILIUM_DIR, "_tmp_HistorialStock_*.xlsx")))
    cont_count = 0
    _OBS_EXCL = {"Transferencia Stock", "Anulación venta stock", "Anulacion venta stock"}
    for fpath in cont_files:
        fname = os.path.basename(fpath)
        try:
            wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
        except Exception as e:
            print(f"    Salteo {fname}: {e}"); continue
        ws = wb.active
        rows_c = list(ws.iter_rows(values_only=True))
        wb.close()
        if len(rows_c) < 2: continue
        for row in rows_c[1:]:
            if not row or len(row) < 6: continue
            # Columnas: Fecha(0)|Deposito(1)|Detalle(2)|Codigo(3)|Nombre(4)|Cantidad(5)|...|Observaciones(7)
            fecha_dt, fecha_str = _parse_fecha(row[0])
            if fecha_dt is None: continue
            if fecha_str <= last_ventas_date: continue      # cubierto por BOSQUE_SALIDAS
            if fecha_str < CONTABILIUM_CUTOVER: continue    # antes del corte — no mezclar con GC
            obs = str(row[7]).strip() if len(row) > 7 and row[7] not in (None, "", "nan", "None") else ""
            if obs in _OBS_EXCL: continue
            cant_raw = row[5]
            if cant_raw is None: continue
            try: cant = float(cant_raw)
            except: continue
            if cant >= 0: continue      # solo salidas (negativos)
            cod_raw = row[3]
            if cod_raw is None: continue
            try: cod = str(int(float(str(cod_raw).strip())))
            except: cod = str(cod_raw).strip()
            if not cod or cod in ("nan", "None", ""): continue
            dep_raw = row[1]
            dep_str = str(dep_raw).strip().upper() if dep_raw not in (None, "", "nan", "None") else ""
            dep = DEP_MAP.get(dep_str, dep_str or "SIN DEPOSITO")
            fantasia = _fantasia_obs(obs)
            rows_out.append([fecha_dt, "", cod, abs(cant), dep, fantasia, "CONTABILIUM"])
            cont_count += 1
    print(f"    Contabilium: {cont_count} filas nuevas (desde {CONTABILIUM_CUTOVER})")

    # ─ 4. Aplicar correcciones manuales (persisten a través de reconstrucciones) ─
    n_corr = _aplicar_correcciones_salidas(rows_out)
    if n_corr:
        print(f"    Correcciones aplicadas: {n_corr} fila(s) desde {os.path.basename(SALIDAS_CORR)}")

    # ─ 5. Escribir consolidado ───────────────────────────────────────────────
    if not rows_out:
        print("    Sin datos para el consolidado.")
        return "?"
    rows_out.sort(key=lambda r: str(r[0])[:10] if r[0] else "")
    wb3 = openpyxl.Workbook()
    ws3 = wb3.active; ws3.title = "SALIDAS"
    ws3.append(SALIDAS_CONS_HEADERS)
    hdr_fill = PatternFill("solid", start_color="1c2230", end_color="1c2230")
    hdr_font = Font(bold=True, color="4ADE80", name="Arial")
    for cell in ws3[1]:
        cell.font = hdr_font; cell.fill = hdr_fill
    for r in rows_out:
        ws3.append(r)
    ws3.freeze_panes = "A2"

    # Guardado seguro: si SALIDAS_CONS está abierto en Excel, wb3.save() falla con
    # PermissionError. Reintentamos unas veces; si sigue bloqueado, escribimos a un
    # archivo alternativo para NO perder los datos frescos silenciosamente.
    saved = False
    for intento in range(5):
        try:
            wb3.save(SALIDAS_CONS)
            saved = True
            break
        except PermissionError:
            if intento == 0:
                print(f"    Aviso: {os.path.basename(SALIDAS_CONS)} está abierto (¿Excel?). Reintentando...")
            time.sleep(3)
    if not saved:
        fallback = SALIDAS_CONS.replace(".xlsx", "_PENDIENTE_GUARDAR.xlsx")
        wb3.save(fallback)
        print(f"    *** ERROR: no se pudo guardar {os.path.basename(SALIDAS_CONS)} (archivo abierto). "
              f"Datos frescos guardados en {os.path.basename(fallback)} — "
              f"cerrar Excel y reemplazar manualmente, o correr de nuevo. ***")
    wb3.close()
    max_date = max((str(r[0])[:10] for r in rows_out if r[0]), default=last_ventas_date)
    print(f"    Consolidado salidas: {len(rows_out)} filas -> {SALIDAS_CONS} (hasta {max_date})")
    return max_date if saved else last_ventas_date


def fetch_insumos():
    """
    Lee la hoja 'Stock insumos' desde CSV local o Google Sheets.
    Estructura: Grupo(0) Insumo(1) Especificaciones(2) Comprador(3)
                + bloques mensuales (fila 0 = nombre mes, ultimo valor del bloque = stock cierre).
    Retorna {"items": [...], "meses": [...], "error": None|str}
    """
    import csv, io, urllib.request

    def _num(s):
        s = str(s).replace(",", "").replace(".", "").replace(" ", "").strip()
        try:
            return int(s)
        except Exception:
            return 0

    all_rows = None
    # 1. CSV local
    if os.path.exists(INSUMOS_CSV):
        try:
            with open(INSUMOS_CSV, encoding="utf-8-sig") as f:
                all_rows = list(csv.reader(f))
            print(f"  Insumos: CSV local leido ({len(all_rows)} filas)")
        except Exception as e:
            print(f"  Advertencia insumos CSV local: {e}")

    # 2. Google Sheets gviz publico (solo funciona si la hoja esta publicada)
    if all_rows is None:
        try:
            req = urllib.request.Request(INSUMOS_GVIZ, headers={"User-Agent": "Mozilla/5.0"})
            with urllib.request.urlopen(req, timeout=20) as resp:
                all_rows = list(csv.reader(io.StringIO(resp.read().decode("utf-8-sig"))))
            print(f"  Insumos: Google Sheets leido ({len(all_rows)} filas)")
        except Exception as e:
            print(f"  Advertencia insumos Google Sheets (publico): {e}")

    # 3. Google Sheets via CDP (usa la sesion de una pestana de Brave ya logueada,
    #    funciona aunque la hoja no este publicada)
    if all_rows is None:
        try:
            content = _download_costos_via_cdp(INSUMOS_GVIZ)
            if content and not content.lstrip().startswith(("<", "{")):
                all_rows = list(csv.reader(io.StringIO(content)))
                print(f"  Insumos: leido via CDP ({len(all_rows)} filas)")
        except Exception as e:
            print(f"  Advertencia insumos via CDP: {e}")

    if all_rows is None:
        return {"items": [], "meses": [], "error": "No se pudo leer (ni local, ni publico, ni CDP)"}

    if not all_rows or len(all_rows) < 3:
        return {"items": [], "meses": [], "error": "Hoja vacia o sin datos"}

    # Detectar bloques de meses desde fila 0
    r0 = all_rows[0]
    months = []
    for i, v in enumerate(r0):
        if v and v.strip():
            months.append({"name": v.strip(), "col": i})

    if not months:
        return {"items": [], "meses": [], "error": "Sin cabeceras de meses"}

    items = []
    last_grupo = ""
    for row in all_rows[2:]:
        if not row or not any(row):
            continue
        insumo = str(row[1]).strip() if len(row) > 1 else ""
        if not insumo or insumo in ("Insumo", ""):
            continue

        grupo_raw = str(row[0]).strip() if row[0] and row[0].strip() else ""
        if grupo_raw:
            last_grupo = grupo_raw
        grupo = last_grupo

        esp  = str(row[2]).strip() if len(row) > 2 else ""
        comp = str(row[3]).strip() if len(row) > 3 else ""

        # Stock por mes: último valor numérico no vacío en cada bloque
        stock_por_mes = {}
        for idx, m in enumerate(months):
            next_col = months[idx + 1]["col"] if idx + 1 < len(months) else len(row)
            last_val = ""
            for c in range(m["col"], min(next_col, len(row))):
                v = row[c].strip() if len(row) > c else ""
                if v:
                    last_val = v
            stock_por_mes[m["name"]] = _num(last_val)

        items.append({
            "grupo":   grupo,
            "insumo":  insumo,
            "esp":     esp,
            "comp":    comp,
            "stock":   stock_por_mes,
        })

    print(f"  Insumos: {len(items)} items, meses: {[m['name'] for m in months]}")
    return {
        "items": items,
        "meses": [m["name"] for m in months],
        "error": None,
    }


COSTOS_SHEET_ID  = "12Ln3lXaWDqpx5hYndJcGH1QDlPaTab3yyEX8j_ahDyY"
COSTOS_SHEET_GID = "173195948"   # pestaña con costos actualizados (3 filas header)
DEST_SHEET_GID   = "1954024258"  # pestaña RESUMEN PRODUCTOS X DESTILERIA
DEST_CSV         = os.path.join(DATA_DIR, "Costos y PVP", "Analisis_costos_resumen_dest.csv")
CERV_SHEET_ID    = "1ekfbqVEqgGBlB_1cD3pqxC60Mn64LgbwNdsJsGIiT58"
CERV_SHEET_GID   = "2130605343"
CERV_CSV         = os.path.join(DATA_DIR, "Costos y PVP", "Analisis_costos_cervezas.csv")
COSTOS_BASE_YEAR = 2024   # mes 1 del sheet = Ene 2024
COSTOS_FIXED_COLS   = 6   # ESTADO, RUBRO, SUB RUBRO, CODIGO, NOMBRE, LINEA
COSTOS_COLS_PER_MO  = 9   # columnas por bloque mensual
COSTOS_RESUMEN_COLS = 9   # columnas por bloque RESUMEN (un bloque por año)


def _costos_col_letter(n):
    """Convierte número de columna (1-based) a letra Excel (A, B, ..., AA, AB, ...)."""
    s = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


def _costos_end_col():
    """
    Calcula la letra de la última columna necesaria para cubrir
    todos los meses hasta el año/mes actual, más un margen.
    Estructura: FIXED_COLS + meses_transcurridos × 9 + bloques_resumen × 9 + buffer
    """
    today = date.today()
    months_elapsed = (today.year - COSTOS_BASE_YEAR) * 12 + today.month
    # Buffer: 2 meses adelante + 1 bloque RESUMEN por año desde 2024 hasta year+1
    resumen_years = today.year - COSTOS_BASE_YEAR + 2
    buffer_months = 2
    total = (COSTOS_FIXED_COLS
             + (months_elapsed + buffer_months) * COSTOS_COLS_PER_MO
             + resumen_years * COSTOS_RESUMEN_COLS)
    return _costos_col_letter(total)


def _download_costos_via_cdp(url, port=9222):
    """
    Usa CDP para ejecutar fetch() en una pestana abierta de Google y obtener el CSV.
    Requiere Brave/Chrome con --remote-debugging-port=9222 y una sesion Google activa.
    Retorna el texto del CSV o None si falla.
    """
    import urllib.request as _ur, json as _j, socket as _sock, struct as _st, base64 as _b64, re as _re

    try:
        with _ur.urlopen(f"http://localhost:{port}/json/list", timeout=3) as r:
            targets = _j.loads(r.read())
    except Exception:
        return None

    # Preferir una pestaña de Google Sheets (docs.google.com/spreadsheets) —
    # el fetch a gviz/tq falla por CORS si se ejecuta desde otro origen Google
    # (mail.google.com, chat.google.com, etc.)
    ws_url = None
    for t in targets:
        if t.get("type") == "page" and "docs.google.com/spreadsheets" in t.get("url", "") \
           and t.get("webSocketDebuggerUrl"):
            ws_url = t["webSocketDebuggerUrl"]; break
    if not ws_url:
        for t in targets:
            if t.get("type") == "page" and "google" in t.get("url", "") and t.get("webSocketDebuggerUrl"):
                ws_url = t["webSocketDebuggerUrl"]; break
    if not ws_url:
        ws_url = next((t.get("webSocketDebuggerUrl") for t in targets
                       if t.get("type") == "page" and t.get("webSocketDebuggerUrl")), None)
    if not ws_url:
        return None

    def _ws_send(s, text):
        payload = text.encode(); n = len(payload); mask = os.urandom(4)
        masked = bytes(b ^ mask[i % 4] for i, b in enumerate(payload))
        hdr = (_st.pack('BB', 0x81, 0x80 | n) if n < 126
               else _st.pack('>BBH', 0x81, 0xFE, n) if n < 65536
               else _st.pack('>BBQ', 0x81, 0xFF, n))
        s.sendall(hdr + mask + masked)

    def _ws_recv(s):
        def _rd(n):
            d = b""
            while len(d) < n:
                chunk = s.recv(n - len(d))
                if not chunk: raise ConnectionError
                d += chunk
            return d
        b1, b2 = _rd(2); is_masked = (b2 & 0x80) != 0; n = b2 & 0x7F
        if n == 126: n = _st.unpack('>H', _rd(2))[0]
        elif n == 127: n = _st.unpack('>Q', _rd(8))[0]
        mask = _rd(4) if is_masked else b""
        payload = _rd(n)
        if is_masked: payload = bytes(b ^ mask[i % 4] for i, b in enumerate(payload))
        return payload.decode('utf-8', errors='replace')

    try:
        m = _re.match(r'ws://([^/:]+)(?::(\d+))?(/.+)', ws_url)
        if not m: return None
        host, ws_port_str, path = m.group(1), m.group(2), m.group(3)
        ws_port = int(ws_port_str) if ws_port_str else port
        s = _sock.create_connection((host, ws_port), timeout=10)
        key = _b64.b64encode(os.urandom(16)).decode()
        s.sendall((
            f"GET {path} HTTP/1.1\r\nHost: {host}:{ws_port}\r\n"
            f"Upgrade: websocket\r\nConnection: Upgrade\r\n"
            f"Sec-WebSocket-Key: {key}\r\nSec-WebSocket-Version: 13\r\n\r\n"
        ).encode())
        resp_hdr = b""
        while b"\r\n\r\n" not in resp_hdr: resp_hdr += s.recv(4096)
        if b"101" not in resp_hdr[:20]: s.close(); return None

        js = (
            "(async () => {"
            "  const r = await fetch(" + _j.dumps(url) + ", {credentials:'include'});"
            "  if (!r.ok) return '__HTTP_' + r.status;"
            "  return await r.text();"
            "})()"
        )
        _ws_send(s, _j.dumps({"id": 99, "method": "Runtime.evaluate",
                               "params": {"expression": js, "awaitPromise": True, "returnByValue": True}}))

        full = ""; s.settimeout(30)
        for _ in range(300):
            try:
                full += _ws_recv(s)
                try:
                    obj = _j.loads(full)
                    if obj.get("id") == 99:
                        val = obj.get("result", {}).get("result", {}).get("value")
                        s.close()
                        if isinstance(val, str) and val.startswith("__HTTP_"):
                            print(f"  CDP costos HTTP error: {val}"); return None
                        return val
                except _j.JSONDecodeError:
                    pass
            except _sock.timeout:
                break
        s.close()
    except Exception as e:
        print(f"  CDP costos error: {e}")
    return None


def download_costos_csv(force=False):
    """
    Descarga la hoja de costos (gid=173195948) de Google Sheets via CDP.
    Solo omite la descarga si el CSV ya fue actualizado HOY (o si force=False y es del dia).
    Retorna True si se descargo, False si ya estaba vigente o hubo error.
    """
    today = date.today()

    if not force and os.path.exists(COSTOS_CSV):
        mtime     = os.path.getmtime(COSTOS_CSV)
        file_date = date.fromtimestamp(mtime)
        if file_date >= today:
            print(f"  Costos CSV al dia ({file_date:%Y-%m-%d}), sin descargar")
            return False

    end_col = _costos_end_col()
    csv_url = (
        f"https://docs.google.com/spreadsheets/d/{COSTOS_SHEET_ID}/"
        f"gviz/tq?tqx=out:csv&gid={COSTOS_SHEET_GID}&range=A1:{end_col}200"
    )
    print(f"  Costos: descargando {today:%Y-%m} via CDP (col hasta {end_col}) ...")

    content = _download_costos_via_cdp(csv_url)

    if not content or len(content.splitlines()) < 5:
        print("  Costos: CDP no disponible o respuesta invalida, usando CSV local existente")
        return False

    os.makedirs(os.path.dirname(COSTOS_CSV), exist_ok=True)
    with open(COSTOS_CSV, "w", encoding="utf-8-sig", newline="") as f:
        f.write(content)
    n_rows = len(content.splitlines())
    print(f"  Costos: CSV guardado ({n_rows} filas, hasta col {end_col})")
    return True


def fetch_costos_completo():
    """
    Descarga (si es necesario) y lee la hoja COSTOS desde el CSV local.
    Soporta dos estructuras de sheet:

    Estructura ANTIGUA (4 filas header, hoja original COSTOS):
      fila 0 = indices internos (-2,-1,0,1,...)
      fila 1 = etiquetas de anio/fecha
      fila 2 = etiquetas de mes (incluye "RESUMEN YYYY")
      fila 3 = nombres de columna (ESTADO, RUBRO, ..., COSTO FINAL, MB, PVP, ...)
      fila 4+ = datos de productos

    Estructura NUEVA (3 filas header, gid=173195948):
      fila 0 = indices internos
      fila 1 = marcadores de anio ("2024", "2025", "2026") y fecha de actualizacion
      fila 2 = encabezados base (ESTADO, RUBRO, SUB-RUBRO, CODIGO, DESCRIPCION, LINEA)
      fila 3+ = datos de productos

    Bloques mensuales de 9 cols: +0=COSTO FINAL, +2=MB, +8=PVP
    Bloques RESUMEN (9 cols): +0=avg_COSTO, +2=avg_MB, +5=avg_PVP
    Retorna: (costos_dict, costos_data)
    """
    import csv, io

    # ── 1. Descargar CSV si esta desactualizado ───────────────────────────────
    download_costos_csv()

    def _money(s):
        if not s: return None
        s = str(s).replace("$", "").replace(" ", "").replace(",", "")
        try: return round(float(s), 2)
        except: return None

    def _pct(s):
        if not s: return None
        s = str(s).replace("%", "").replace(" ", "").replace(",", ".")
        try: return round(float(s), 2)
        except: return None

    costos_dict = {}
    costos_data = {
        "periodos": ["2025", "2026"],
        "meses":    [],
        "productos": [],
        "actualizacion": "",
        "error": None,
    }

    # ── 2. Leer CSV local ─────────────────────────────────────────────────────
    all_rows = None
    source   = "local"

    if os.path.exists(COSTOS_CSV):
        try:
            with open(COSTOS_CSV, encoding="utf-8-sig") as f:
                all_rows = list(csv.reader(f))
            print(f"  Costos: CSV leido ({len(all_rows)} filas)")
        except Exception as e:
            print(f"  Advertencia leyendo CSV local: {e}")

    if all_rows is None:
        try:
            download_costos_csv(force=True)
            with open(COSTOS_CSV, encoding="utf-8-sig") as f:
                all_rows = list(csv.reader(f))
            source = "Google Sheets (forzado)"
        except Exception as e:
            print(f"  Advertencia costos fallback: {e}")
            costos_data["error"] = str(e)
            return costos_dict, costos_data

    if len(all_rows) < 4:
        costos_data["error"] = "CSV con menos de 4 filas"
        return costos_dict, costos_data

    # ── 3. Detectar estructura y mapear columnas ──────────────────────────────
    row1 = all_rows[1]
    row2 = all_rows[2]
    row3 = all_rows[3]
    month_labels = []   # "YYYY-MM" por cada bloque mensual en costo_cols

    # Estructura antigua: fila 3 tiene etiquetas "COSTO FINAL"
    costo_cols_chk = [i for i, h in enumerate(row3) if h.strip() == "COSTO FINAL"]

    if costo_cols_chk:
        # ── Estructura ANTIGUA ──────────────────────────────────────────────
        costo_cols = costo_cols_chk
        data_start = 4
        resumen_pos = {}
        for i, r2 in enumerate(row2):
            if "RESUMEN" in r2.upper():
                yr_match = r2.strip().split()[-1]   # "RESUMEN 2025" -> "2025"
                resumen_pos[yr_match] = i
        print(f"  Costos: estructura antigua ({len(costo_cols)} bloques mensuales)")
    else:
        # ── Estructura NUEVA: bloques derivados de marcadores de anio en fila 1 ──
        data_start = 3
        year_starts = []
        for i, v in enumerate(row1):
            v_s = str(v).strip()
            if re.match(r'^20\d{2}$', v_s):
                year_starts.append((int(v_s), i))
        year_starts.sort(key=lambda x: x[1])   # ordenar por posicion de columna

        costo_cols = []
        resumen_pos = {}


        if year_starts:
            max_col = max((len(r) for r in all_rows[data_start:] if r), default=300)
            year_starts.append((9999, max_col + 1))   # sentinel al final
            for k in range(len(year_starts) - 1):
                yr, start_col = year_starts[k]
                next_col      = year_starts[k + 1][1]
                n_blocks      = (next_col - start_col) // 9
                has_resumen   = n_blocks >= 13
                n_monthly     = n_blocks - 1 if has_resumen else n_blocks
                if has_resumen:
                    resumen_pos[str(yr)] = start_col + (n_blocks - 1) * 9
                # Solo incluir bloques mensuales hasta el mes actual del anio en curso
                # (los bloques futuros son proyecciones, no costos reales)
                if yr < date.today().year:
                    max_b = n_monthly          # anio pasado: todos los meses
                elif yr == date.today().year and n_monthly == 12:
                    max_b = date.today().month  # anio actual: hasta el mes de hoy
                elif yr == date.today().year:
                    max_b = n_monthly           # anio actual con pocos bloques (ej. solo Dic)
                else:
                    max_b = 0                   # anio futuro: omitir
                # Años pasados con <12 bloques: los bloques son los ÚLTIMOS meses del año
                # Año actual o años completos: los bloques arrancan desde Enero
                if yr < date.today().year and n_monthly < 12:
                    mes_inicio = 12 - n_monthly + 1
                else:
                    mes_inicio = 1
                for b in range(min(n_monthly, max_b)):
                    col = start_col + b * 9
                    costo_cols.append(col)
                    month_labels.append(f"{yr}-{mes_inicio + b:02d}")
        print(f"  Costos: estructura nueva ({len(costo_cols)} bloques, resumen: {list(resumen_pos.keys())})")

    def _get_val(row, col, offset, default=None):
        idx = col + offset
        if idx < len(row):
            return row[idx]
        return default

    # ── 4. Parsear productos ──────────────────────────────────────────────────
    for row in all_rows[data_start:]:
        if len(row) < 5: continue
        cod_raw = row[3].strip()
        desc    = row[4].strip()
        if not cod_raw or not desc: continue
        try:   cod = str(int(float(cod_raw)))
        except: continue

        # Costo mas reciente: escanear costo_cols de derecha a izquierda
        costo_actual = pvp_actual = mb_actual = None
        for ci in reversed(costo_cols):
            c = _money(_get_val(row, ci, 0))
            if c and c > 0:
                costo_actual = c
                pvp_actual   = _money(_get_val(row, ci, 8))
                mb_actual    = _pct(_get_val(row, ci, 2))
                break

        # Datos por mes (costo_cols + month_labels van en paralelo)
        meses_data = {}
        for ci, ml in zip(costo_cols, month_labels):
            c = _money(_get_val(row, ci, 0))
            if c is not None and c > 0:
                meses_data[ml] = {
                    "costo": c,
                    "pvp":   _money(_get_val(row, ci, 8)),
                    "mb":    _pct(_get_val(row, ci, 2)),
                }

        # Datos de RESUMEN 2025 y 2026
        periodos_data = {}
        for yr, base in resumen_pos.items():
            if yr not in ("2025", "2026"): continue
            if base < len(row):
                periodos_data[yr] = {
                    "costo": _money(_get_val(row, base, 0)),
                    "mb":    _pct(_get_val(row, base, 2)),
                    "cmv":   _pct(_get_val(row, base, 3)),
                    "pvp":   _money(_get_val(row, base, 5)),
                }

        costos_data["productos"].append({
            "cod":    cod,
            "desc":   desc,
            "rubro":  row[1].strip(),
            "sub":    row[2].strip(),
            "linea":  row[5].strip() if len(row) > 5 else "",
            "estado": row[0].strip(),
            "costo":  costo_actual,
            "pvp":    pvp_actual,
            "mb":     mb_actual,
            "periodos": periodos_data,
            "meses":  meses_data,
        })

        if costo_actual and costo_actual > 0:
            costos_dict[cod] = costo_actual

    costos_data["meses"] = month_labels
    costos_data["actualizacion"] = datetime.now(_AR).strftime("%Y-%m-%dT%H:%M:%S-03:00")
    n = len(costos_data["productos"])
    print(f"  Costos: {n} productos cargados desde {source}  |  {len(costos_dict)} con precio")

    return costos_dict, costos_data


def _build_ventas_output(monthly):
    """
    Convierte el dict monthly {yr -> mo_str -> {deps, prods}} a las estructuras
    VD_out / MONTHLY_DATA / PROD_DATA que usa el dashboard.
    Devuelve (result_dict, ventas_hasta_str, monthly).
    """
    VD_out       = {"annual": {}, "monthly": {}, "deps": {}, "rubros": {}}
    MONTHLY_DATA = {}
    PROD_DATA    = {}
    CLIENT_DATA  = {}
    VD           = {}

    for yr in sorted(monthly.keys()):
        yr_mo = monthly[yr]

        tot_c = sum(p["c"] for mo_s in yr_mo for p in yr_mo[mo_s]["prods"].values())
        tot_u = sum(p["u"] for mo_s in yr_mo for p in yr_mo[mo_s]["prods"].values())
        # m = monto = u × costo (misma magnitud que c ya que no hay PVP por transaccion)
        VD[yr] = {"m": round(tot_c, 2), "c": round(tot_c, 2), "u": round(tot_u, 2)}
        VD_out["annual"][yr] = {**VD[yr], "mg": 0.0}

        series = []
        for mo in range(1, 13):
            mo_s = str(mo)
            if mo_s not in yr_mo: continue
            prods_mo = list(yr_mo[mo_s]["prods"].values())
            tot_c_mo = sum(p["c"] for p in prods_mo)
            tot_u_mo = sum(p["u"] for p in prods_mo)
            if tot_u_mo > 0:
                series.append({"mes": mo, "label": MESES_NOM[mo],
                                "m": round(tot_c_mo, 2), "c": round(tot_c_mo, 2),
                                "u": round(tot_u_mo, 2), "mg": 0.0})
        VD_out["monthly"][yr] = series

        dep_tot = {}
        for mo_s, mo_data in yr_mo.items():
            for dep, dv in mo_data["deps"].items():
                if dep not in dep_tot: dep_tot[dep] = {"m": 0.0, "c": 0.0, "u": 0.0}
                dep_tot[dep]["c"] += dv["c"]; dep_tot[dep]["u"] += dv["u"]
        for v in dep_tot.values():
            v["m"] = v["c"]   # monto = costo
            v["mg"] = 0.0
        VD_out["deps"][yr] = dep_tot

        rub_dep = {}
        for mo_s, mo_data in yr_mo.items():
            for p in mo_data["prods"].values():
                rub = p["rub"]
                if rub not in rub_dep: rub_dep[rub] = {}
                for dep, dv in p["deps"].items():
                    if dep not in rub_dep[rub]: rub_dep[rub][dep] = {"m": 0.0, "c": 0.0, "u": 0.0}
                    rub_dep[rub][dep]["c"] += dv["c"]; rub_dep[rub][dep]["u"] += dv["u"]
        for subs in rub_dep.values():
            for v in subs.values(): v["m"] = v["c"]
        VD_out["rubros"][yr] = rub_dep

        _WK_RANGES = {"1":"1–7","2":"8–14","3":"15–21","4":"22–28","5":"29–31"}
        md_yr = {"deps": {}, "rubs": {}, "prods": {}, "weekly": {}}
        for mo_s, mo_data in yr_mo.items():
            md_yr["deps"][mo_s] = {dep: {**dv, "m": dv.get("c", 0.0), "mg": 0.0}
                                    for dep, dv in mo_data["deps"].items()}
            rubs_mo = {}
            for p in mo_data["prods"].values():
                rk = p["rub"]
                if rk not in rubs_mo: rubs_mo[rk] = {"m": 0.0, "c": 0.0, "u": 0.0}
                rubs_mo[rk]["c"] += p["c"]; rubs_mo[rk]["u"] += p["u"]
            for v in rubs_mo.values(): v["m"] = v["c"]
            md_yr["rubs"][mo_s] = rubs_mo
            prods_list = sorted(mo_data["prods"].values(), key=lambda p: -p["u"])
            for p in prods_list:
                p["mg"] = 0.0
                for wv in p.get("weekly", {}).values():
                    wv["u"] = round(wv["u"], 2); wv["c"] = round(wv["c"], 2)
            md_yr["prods"][mo_s] = prods_list
            # Weekly breakdown for this month
            wkly_raw = mo_data.get("weekly", {})
            mo_int_lbl = int(mo_s)
            mes_nom = MESES_NOM[mo_int_lbl] if 1 <= mo_int_lbl <= 12 else mo_s
            md_yr["weekly"][mo_s] = sorted(
                [{"w": wk,
                  "label": f"{_WK_RANGES.get(wk, wk)} {mes_nom}",
                  "u": round(v["u"], 2), "c": round(v["c"], 2)}
                 for wk, v in wkly_raw.items()],
                key=lambda x: int(x["w"])
            )
        MONTHLY_DATA[yr] = md_yr

        prod_agg = {}
        for mo_s, mo_data in yr_mo.items():
            for cod, p in mo_data["prods"].items():
                if cod not in prod_agg:
                    prod_agg[cod] = {"cod": cod, "art": p["art"], "rub": p["rub"], "sub": p["sub"],
                                     "m": 0.0, "c": 0.0, "u": 0.0, "deps": {}}
                pa = prod_agg[cod]
                if p["art"] and not pa["art"]: pa["art"] = p["art"]
                pa["c"] += p["c"]; pa["m"] += p["c"]; pa["u"] += p["u"]
                for dep, dv in p["deps"].items():
                    if dep not in pa["deps"]: pa["deps"][dep] = {"m": 0.0, "c": 0.0, "u": 0.0}
                    pa["deps"][dep]["c"] += dv["c"]; pa["deps"][dep]["m"] += dv["c"]; pa["deps"][dep]["u"] += dv["u"]

        total_u = sum(p["u"] for p in prod_agg.values())
        prods_sorted = sorted(prod_agg.values(), key=lambda p: -p["u"])
        acum = 0
        for p in prods_sorted:
            p["mg"]   = 0.0
            p["pct"]  = round(p["u"] / total_u * 100, 2) if total_u > 0 else 0
            acum     += p["pct"]
            p["acum"] = round(acum, 2)
            p["tk"]   = 0
        PROD_DATA[yr] = prods_sorted

        # ── CLIENT_DATA ──
        cl_yr = {}
        for mo_s, mo_data in yr_mo.items():
            for fantasia, cl_mo in (mo_data.get("clients") or {}).items():
                if fantasia not in cl_yr:
                    cl_yr[fantasia] = {"u": 0.0, "c": 0.0, "monthly": {}, "prods": {}, "deps": {}}
                cl = cl_yr[fantasia]
                cl["u"] += cl_mo["u"]; cl["c"] += cl_mo["c"]
                mo_int = int(mo_s)
                if mo_int not in cl["monthly"]:
                    cl["monthly"][mo_int] = {"u": 0.0, "c": 0.0, "prods": {}, "weekly": {}}
                mo_entry = cl["monthly"][mo_int]
                mo_entry["u"] += cl_mo["u"]
                mo_entry["c"] += cl_mo["c"]
                # productos del mes para este cliente
                for cod, pv in cl_mo["prods"].items():
                    if cod not in mo_entry["prods"]:
                        mo_entry["prods"][cod] = {**pv, "u": 0.0, "c": 0.0}
                    mo_entry["prods"][cod]["u"] += pv["u"]
                    mo_entry["prods"][cod]["c"] += pv["c"]
                # weekly del mes para este cliente
                for wk, wv in cl_mo.get("weekly", {}).items():
                    if wk not in mo_entry["weekly"]:
                        mo_entry["weekly"][wk] = {"u": 0.0, "c": 0.0}
                    mo_entry["weekly"][wk]["u"] += wv["u"]
                    mo_entry["weekly"][wk]["c"] += wv["c"]
                # productos anuales
                for cod, pv in cl_mo["prods"].items():
                    if cod not in cl["prods"]:
                        cl["prods"][cod] = {**pv, "u": 0.0, "c": 0.0}
                    cl["prods"][cod]["u"] += pv["u"]
                    cl["prods"][cod]["c"] += pv["c"]
                for dep, dv in cl_mo["deps"].items():
                    if dep not in cl["deps"]:
                        cl["deps"][dep] = {"u": 0.0, "c": 0.0}
                    cl["deps"][dep]["u"] += dv["u"]
                    cl["deps"][dep]["c"] += dv["c"]
        # Finalizar: prods como lista ordenada, monthly como serie con prods y weekly ordenados
        _WK_RANGES2 = {"1":"1–7","2":"8–14","3":"15–21","4":"22–28","5":"29–31"}
        for cl in cl_yr.values():
            cl["prods"] = sorted(cl["prods"].values(), key=lambda p: -p["u"])
            for v in cl["monthly"].values():
                for pv in v["prods"].values():
                    for wv in pv.get("weekly", {}).values():
                        wv["u"] = round(wv["u"], 2); wv["c"] = round(wv["c"], 2)
            cl["monthly"] = [
                {"mes": mo, "label": MESES_NOM[mo],
                 "u": round(v["u"], 2), "c": round(v["c"], 2),
                 "prods": sorted(v["prods"].values(), key=lambda p: -p["u"]),
                 "weekly": sorted(
                     [{"w": wk,
                       "label": f"{_WK_RANGES2.get(wk, wk)} {MESES_NOM[mo]}",
                       "u": round(wv["u"], 2), "c": round(wv["c"], 2)}
                      for wk, wv in v.get("weekly", {}).items()],
                     key=lambda x: int(x["w"])
                 )}
                for mo, v in sorted(cl["monthly"].items())
            ]
            cl["u"] = round(cl["u"], 2); cl["c"] = round(cl["c"], 2)
        CLIENT_DATA[yr] = cl_yr

    ventas_max_yr = max(monthly.keys()) if monthly else "2026"
    ventas_max_mo = max(int(k) for k in monthly.get(ventas_max_yr, {"0": None}).keys()) if monthly else 4
    ventas_hasta  = f"{ventas_max_yr}-{ventas_max_mo:02d}"

    return {"VD": VD_out, "MONTHLY_DATA": MONTHLY_DATA, "PROD_DATA": PROD_DATA,
            "CLIENT_DATA": CLIENT_DATA}, ventas_hasta, monthly


def parse_remitos_gc(folder=None, prod_lookup=None, costos=None):
    """
    Lee archivos 'Remitos GC*.xlsx' de Data/Salidas/GC/.
    Hoja 'Datos': fila 0 = headers camelCase, datos desde fila 1.
    Col 3=Fecha  Col 9=Cantidad  Col 14=Producto  Col 16=Codigo  Col 18=Deposito
    Devuelve misma estructura que parse_ventas().
    """
    if folder      is None: folder      = GC_SALIDAS_DIR
    if prod_lookup is None: prod_lookup = {}
    if costos      is None: costos      = {}

    try:
        import openpyxl
    except ImportError:
        import subprocess, sys
        subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl",
                        "--break-system-packages", "-q"])
        import openpyxl

    os.makedirs(folder, exist_ok=True)
    files = sorted(glob.glob(os.path.join(folder, "Remitos GC*.xlsx")))
    if not files:
        print(f"  Remitos GC: no se encontraron archivos en {folder}")
        return {"VD": {}, "MONTHLY_DATA": {}, "PROD_DATA": {}}, None, {}

    monthly    = {}
    total_rows = 0
    dup_total  = 0
    seen_rows  = set()   # (fecha_str, cod, dep, fantasia) — descarta duplicados

    for fpath in files:
        fname = os.path.basename(fpath)
        try:
            wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
        except Exception as e:
            print(f"  Salteo {fname}: {e}")
            continue

        ws   = wb["Datos"] if "Datos" in wb.sheetnames else wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        if len(rows) < 2:
            print(f"  Salteo {fname}: sin datos")
            continue

        file_rows = 0
        for row in rows[1:]:          # row 0 = headers
            if not row or len(row) < 19: continue

            fecha_raw = row[3]
            qty_raw   = row[9]
            prod_raw  = row[14]
            cod_raw   = row[16]
            dep_raw   = row[18]

            if fecha_raw is None or qty_raw is None or cod_raw is None:
                continue

            # Fecha
            try:
                if hasattr(fecha_raw, "year"):
                    yr_int, mo = fecha_raw.year, fecha_raw.month
                else:
                    s = str(fecha_raw).strip()
                    parts = s.replace("/", "-").split("-")
                    if len(parts) == 3:
                        yr_int, mo = (int(parts[0]), int(parts[1])) if len(parts[0]) == 4 \
                                     else (int(parts[2]), int(parts[1]))
                    else:
                        continue
                if not (2020 <= yr_int <= 2030 and 1 <= mo <= 12):
                    continue
                yr = str(yr_int)
            except Exception:
                continue

            # Cantidad
            try:
                u = float(str(qty_raw).replace(",", "."))
            except Exception:
                continue
            if u <= 0:
                continue

            # Código
            try:
                cod = str(int(float(str(cod_raw).strip())))
            except Exception:
                cod = str(cod_raw).strip()
            if not cod or cod in ("nan", "None", ""):
                continue

            # Depósito
            dep_str = str(dep_raw).strip().upper() if dep_raw else ""
            dep = DEP_MAP.get(dep_str, dep_str if dep_str else "SIN DEPOSITO")

            # Deduplicación: misma (fecha, cod, dep, fantasia) = error de carga
            fan_raw = row[4] if len(row) > 4 else None
            fan_str = str(fan_raw).strip().upper() if fan_raw not in (None, "", "nan", "None") else ""
            fecha_str = f"{yr}-{mo:02d}"
            if hasattr(fecha_raw, "year"):
                fecha_str = f"{fecha_raw.year}-{fecha_raw.month:02d}-{fecha_raw.day:02d}"
            row_key = (fecha_str, cod, dep, fan_str)
            if row_key in seen_rows:
                dup_total += 1; continue
            seen_rows.add(row_key)

            # Nombre del producto
            art_gc = str(prod_raw).strip() if prod_raw else ""
            if art_gc in ("nan", "None", ""): art_gc = ""

            # Enriquecer desde lookup
            pi  = prod_lookup.get(cod, {})
            rub = pi.get("rub", "SIN RUBRO") or "SIN RUBRO"
            sub = pi.get("sub", "") or ""
            art = pi.get("art", "") or art_gc or ""

            costo_unit = costos.get(cod, 0.0)
            c = round(u * costo_unit, 2)

            mo_str = str(mo)
            total_rows += 1; file_rows += 1

            if yr not in monthly:
                monthly[yr] = {}
            if mo_str not in monthly[yr]:
                monthly[yr][mo_str] = {"deps": {}, "prods": {}}
            mo_data = monthly[yr][mo_str]

            if dep not in mo_data["deps"]:
                mo_data["deps"][dep] = {"m": 0.0, "c": 0.0, "u": 0.0}
            mo_data["deps"][dep]["c"] += c
            mo_data["deps"][dep]["u"] += u

            if cod not in mo_data["prods"]:
                mo_data["prods"][cod] = {"cod": cod, "art": art, "rub": rub, "sub": sub,
                                         "m": 0.0, "c": 0.0, "u": 0.0, "deps": {}}
            p = mo_data["prods"][cod]
            if art and not p["art"]: p["art"] = art
            p["c"] += c; p["u"] += u
            if dep not in p["deps"]:
                p["deps"][dep] = {"m": 0.0, "c": 0.0, "u": 0.0}
            p["deps"][dep]["c"] += c
            p["deps"][dep]["u"] += u

        print(f"  Remitos GC {fname}: {file_rows} filas")

    print(f"  Remitos GC total: {total_rows} filas de {len(files)} archivo(s) ({dup_total} duplicados descartados)")

    if not monthly:
        return {"VD": {}, "MONTHLY_DATA": {}, "PROD_DATA": {}}, None, {}

    return _build_ventas_output(monthly)


def parse_ventas(ventas_path, prod_lookup=None, costos=None):
    """
    Lee el consolidado de salidas (Salidas_consolidado.xlsx) o, como fallback,
    Bosque salidas.xlsx directamente.
    Consolidado: cabecera en fila 1 (FECHA|RAZON SOCIAL|CODIGO|CANTIDAD|DEPOSITO|NOMBRE FANTASIA|FUENTE)
    Bosque salidas: titulo en fila 1, cabecera en fila 2 → saltar ambas.
    Nombre fantasia siempre en MAYUSCULAS.
    """
    if prod_lookup is None: prod_lookup = {}
    if costos      is None: costos      = {}

    try:
        import openpyxl
    except ImportError:
        import subprocess, sys
        subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl",
                        "--break-system-packages", "-q"])
        import openpyxl

    if not os.path.exists(ventas_path):
        print(f"  Archivo salidas no encontrado: {ventas_path}")
        return {"VD": {}, "MONTHLY_DATA": {}, "PROD_DATA": {}}, "?", {}

    is_consolidado = (os.path.normcase(ventas_path) == os.path.normcase(SALIDAS_CONS))

    wb = openpyxl.load_workbook(ventas_path, read_only=True, data_only=True)
    ws = wb["SALIDAS"] if "SALIDAS" in wb.sheetnames else wb.active
    rows_iter = ws.iter_rows(values_only=True)

    if is_consolidado:
        next(rows_iter, None)   # solo cabecera
    else:
        next(rows_iter, None)   # título vacío
        next(rows_iter, None)   # cabecera

    monthly = {}   # yr -> mo_str -> {deps:{}, prods:{cod:{}}}
    row_count = skip_count = 0

    for row in rows_iter:
        if not row or not any(row):
            continue

        fecha_raw  = row[0] if len(row) > 0 else None
        rs_raw     = row[1] if len(row) > 1 else None
        cod_raw    = row[2] if len(row) > 2 else None
        qty_raw    = row[3] if len(row) > 3 else None
        dep_raw    = row[4] if len(row) > 4 else None
        fan_raw    = row[5] if len(row) > 5 else None
        fuente_raw = row[6] if len(row) > 6 else None

        if fecha_raw is None or cod_raw is None or qty_raw is None:
            skip_count += 1; continue

        # ── Fecha ──
        try:
            if hasattr(fecha_raw, "year"):
                yr_int, mo, day = fecha_raw.year, fecha_raw.month, fecha_raw.day
            else:
                s = str(fecha_raw).strip()
                parts = s.replace("/", "-").split("-")
                if len(parts) == 3:
                    if len(parts[0]) == 4:
                        yr_int, mo, day = int(parts[0]), int(parts[1]), int(parts[2])
                    else:
                        yr_int, mo, day = int(parts[2]), int(parts[1]), int(parts[0])
                else:
                    skip_count += 1; continue
            if not (2020 <= yr_int <= 2030 and 1 <= mo <= 12 and 1 <= day <= 31):
                skip_count += 1; continue
            yr = str(yr_int)
        except Exception:
            skip_count += 1; continue
        wk_str = str((day - 1) // 7 + 1)   # semana dentro del mes: "1"–"5"

        # ── Cantidad ──
        try:
            u = float(str(qty_raw).replace(",", "."))
        except Exception:
            skip_count += 1; continue
        if u <= 0:
            continue

        # ── Código ──
        try:
            cod = str(int(float(str(cod_raw).strip())))
        except Exception:
            cod = str(cod_raw).strip()
        if not cod or cod in ("nan", "None", ""):
            skip_count += 1; continue

        # ── Depósito ──
        dep_str = str(dep_raw).strip().upper() if dep_raw else ""
        dep = DEP_MAP.get(dep_str, dep_str if dep_str else "SIN DEPOSITO")

        # ── Nombre de fantasía del cliente (siempre MAYUSCULAS) ──
        fantasia = str(fan_raw).strip().upper() if fan_raw not in (None, "", "nan", "None") else ""
        if not fantasia or fantasia in ("NAN", "NONE"):
            fantasia = str(rs_raw).strip().upper() if rs_raw not in (None, "", "nan", "None") else "SIN CLIENTE"
        if not fantasia or fantasia in ("NAN", "NONE", ""):
            fantasia = "SIN CLIENTE"

        # ── Enriquecer desde productos ──
        pi  = prod_lookup.get(cod, {})
        rub = pi.get("rub", "SIN RUBRO") or "SIN RUBRO"
        sub = pi.get("sub", "") or ""
        art = pi.get("art", "") or ""

        # ── Costo ──
        costo_unit = costos.get(cod, 0.0)
        c = round(u * costo_unit, 2)
        m = 0.0  # sin precio de venta en el archivo de salidas

        mo_str = str(mo)
        row_count += 1

        if yr not in monthly:
            monthly[yr] = {}
        if mo_str not in monthly[yr]:
            monthly[yr][mo_str] = {"deps": {}, "prods": {}, "weekly": {}}
        mo_data = monthly[yr][mo_str]

        # deps
        if dep not in mo_data["deps"]:
            mo_data["deps"][dep] = {"m": 0.0, "c": 0.0, "u": 0.0}
        mo_data["deps"][dep]["c"] += c
        mo_data["deps"][dep]["u"] += u

        # prods
        if cod not in mo_data["prods"]:
            mo_data["prods"][cod] = {"cod": cod, "art": art, "rub": rub, "sub": sub,
                                     "m": 0.0, "c": 0.0, "u": 0.0, "deps": {}}
        p = mo_data["prods"][cod]
        if art and not p["art"]: p["art"] = art
        p["c"] += c; p["u"] += u
        if dep not in p["deps"]:
            p["deps"][dep] = {"m": 0.0, "c": 0.0, "u": 0.0}
        p["deps"][dep]["c"] += c
        p["deps"][dep]["u"] += u
        if "weekly" not in p: p["weekly"] = {}
        if wk_str not in p["weekly"]: p["weekly"][wk_str] = {"u": 0.0, "c": 0.0}
        p["weekly"][wk_str]["u"] += u
        p["weekly"][wk_str]["c"] += c

        # ── Cliente ──
        if "clients" not in mo_data:
            mo_data["clients"] = {}
        if fantasia not in mo_data["clients"]:
            mo_data["clients"][fantasia] = {"u": 0.0, "c": 0.0, "prods": {}, "deps": {}}
        cl = mo_data["clients"][fantasia]
        cl["u"] += u; cl["c"] += c
        if cod not in cl["prods"]:
            cl["prods"][cod] = {"cod": cod, "art": art, "rub": rub, "sub": sub, "u": 0.0, "c": 0.0, "weekly": {}}
        cl["prods"][cod]["u"] += u
        cl["prods"][cod]["c"] += c
        if wk_str not in cl["prods"][cod]["weekly"]:
            cl["prods"][cod]["weekly"][wk_str] = {"u": 0.0, "c": 0.0}
        cl["prods"][cod]["weekly"][wk_str]["u"] += u
        cl["prods"][cod]["weekly"][wk_str]["c"] += c
        if dep not in cl["deps"]:
            cl["deps"][dep] = {"u": 0.0, "c": 0.0}
        cl["deps"][dep]["u"] += u
        cl["deps"][dep]["c"] += c

        # ── Semanal global ──
        if wk_str not in mo_data["weekly"]:
            mo_data["weekly"][wk_str] = {"u": 0.0, "c": 0.0}
        mo_data["weekly"][wk_str]["u"] += u
        mo_data["weekly"][wk_str]["c"] += c

        # ── Semanal por cliente ──
        if "weekly" not in cl:
            cl["weekly"] = {}
        if wk_str not in cl["weekly"]:
            cl["weekly"][wk_str] = {"u": 0.0, "c": 0.0}
        cl["weekly"][wk_str]["u"] += u
        cl["weekly"][wk_str]["c"] += c

    wb.close()
    print(f"  Salidas: {row_count} filas procesadas, {skip_count} saltadas")

    # Post-proceso: rellenar art/rub/sub vacíos desde prod_lookup
    filled = 0
    for yr_mo in monthly.values():
        for mo_data in yr_mo.values():
            for p in mo_data["prods"].values():
                if not p.get("art") or not p.get("rub") or p.get("rub") == "SIN RUBRO":
                    pi = prod_lookup.get(str(p["cod"]), {})
                    if pi.get("art") and not p.get("art"):
                        p["art"] = pi["art"]; filled += 1
                    if pi.get("rub") and (not p.get("rub") or p.get("rub") == "SIN RUBRO"):
                        p["rub"] = pi["rub"]
                    if pi.get("sub") and not p.get("sub"):
                        p["sub"] = pi["sub"]
    if filled:
        print(f"  Nombres enriquecidos desde PRODUCTOS.xlsx: {filled} productos actualizados")

    return _build_ventas_output(monthly)


# ─── 3. VELOCIDAD DESDE VENTAS ────────────────────────────────────────────────
def compute_velocity(monthly, ref_date):
    """
    Para cada (cod, dep_canonico) calcula:
      q90:  unidades vendidas en los ultimos ~90 dias (3 meses hacia atras)
      q365: unidades vendidas en los ultimos ~365 dias (12 meses hacia atras)
    Devuelve dict:  vel[cod][dep_canonico] = {"q90": x, "q365": y}
    DEP_MAP ya unifico KLOZER MKT->KLOZER, OFI/OFICINA->OFICINA en ventas.
    """
    import calendar as _cal

    cutoff_90  = ref_date - timedelta(days=90)
    cutoff_365 = ref_date - timedelta(days=365)

    vel = {}  # cod -> {dep -> {q90, q365}}

    for yr, yr_mo in monthly.items():
        for mo_s, mo_data in yr_mo.items():
            mo       = int(mo_s)
            yr_int   = int(yr)
            mo_start = date(yr_int, mo, 1)
            mo_days  = _cal.monthrange(yr_int, mo)[1]
            mo_end   = date(yr_int, mo, mo_days)

            def _frac(cutoff):
                s = max(mo_start, cutoff)
                e = min(mo_end, ref_date)
                if s > e: return 0.0
                return ((e - s).days + 1) / mo_days

            f365 = _frac(cutoff_365)
            f90  = _frac(cutoff_90)

            if f365 <= 0:
                continue

            for p in mo_data["prods"].values():
                cod = p["cod"]
                if cod not in vel:
                    vel[cod] = {}
                for dep, dv in p.get("deps", {}).items():
                    if dep not in vel[cod]:
                        vel[cod][dep] = {"q90": 0.0, "q365": 0.0}
                    u = dv.get("u", 0)
                    vel[cod][dep]["q365"] += u * f365
                    if f90 > 0:
                        vel[cod][dep]["q90"] += u * f90

    return vel


def apply_velocity(inv_data, vel):
    """Aplica velocidad calculada desde ventas a cada articulo de INV_DATA.
    Ventas de KLOZER y KLOZER MKT quedan separadas en el stock pero en ventas
    ambas se registran bajo el canal 'KLOZER', por lo que la velocidad se
    asigna al total klozer+klozer_mkt para calcular meses de stock combinados.
    """
    for item in inv_data:
        cod = str(item["cod"])
        cv = vel.get(cod, {})

        # Velocidad KLOZER depósito (sin MKT)
        kv = cv.get("KLOZER", {})
        k90  = round(kv.get("q90",  0.0), 2)
        k365 = round(kv.get("q365", 0.0), 2)

        # Velocidad KLOZER MKT (canal propio)
        kmv  = cv.get("KLOZER_MKT", {})
        km90_raw  = round(kmv.get("q90",  0.0), 2)
        km365_raw = round(kmv.get("q365", 0.0), 2)

        # Velocidad OFICINA
        ov = cv.get("OFICINA", {})
        o90  = round(ov.get("q90",  0.0), 2)
        o365 = round(ov.get("q365", 0.0), 2)

        # Velocidad combinada para meses de stock totales (klozer+mkt+ofi)
        b90  = round(k90 + km90_raw + o90,  2)
        b365 = round(k365 + km365_raw + o365, 2)

        # min/mes KLOZER (solo depósito KLOZER)
        min_k90  = round(k90  / 90  * 30, 1) if k90  > 0 else 0.0
        min_k365 = round(k365 / 365 * 30, 1) if k365 > 0 else 0.0
        _cnt_k   = (1 if min_k90 else 0) + (1 if min_k365 else 0)
        min_k    = round((min_k90 + min_k365) / _cnt_k, 1) if _cnt_k else 0.0

        # min/mes KLOZER MKT (propio)
        min_km90  = round(km90_raw  / 90  * 30, 1) if km90_raw  > 0 else 0.0
        min_km365 = round(km365_raw / 365 * 30, 1) if km365_raw > 0 else 0.0
        _cnt_km   = (1 if min_km90 else 0) + (1 if min_km365 else 0)
        min_km    = round((min_km90 + min_km365) / _cnt_km, 1) if _cnt_km else 0.0

        min_o90  = round(o90  / 90  * 30, 1) if o90  > 0 else 0.0
        min_o365 = round(o365 / 365 * 30, 1) if o365 > 0 else 0.0
        _cnt_o   = (1 if min_o90 else 0) + (1 if min_o365 else 0)
        min_o    = round((min_o90 + min_o365) / _cnt_o, 1) if _cnt_o else 0.0

        _mb90    = round(b90  / 90  * 30, 1) if b90  > 0 else 0.0
        _mb365   = round(b365 / 365 * 30, 1) if b365 > 0 else 0.0
        _cnt_b   = (1 if _mb90 else 0) + (1 if _mb365 else 0)
        min_b    = round((_mb90 + _mb365) / _cnt_b, 1) if _cnt_b else 0.0

        # Meses de stock KLOZER depósito
        k_total = item["klozer"] + item.get("klozer_mkt", 0)
        km      = item.get("klozer_mkt", 0)
        pk      = round(k_total / min_k,    2) if min_k    > 0 else 0.0
        pk90    = round(k_total / min_k90,  2) if min_k90  > 0 else 0.0
        pk365   = round(k_total / min_k365, 2) if min_k365 > 0 else 0.0
        pk_dep  = round(item["klozer"] / min_k, 2) if min_k > 0 else 0.0
        # Meses de stock KLOZER MKT usando velocidad propia
        pk_mkt    = round(km / min_km,    2) if min_km    > 0 else 0.0
        pk_mkt90  = round(km / min_km90,  2) if min_km90  > 0 else 0.0
        pk_mkt365 = round(km / min_km365, 2) if min_km365 > 0 else 0.0
        po      = round(item["ofi"]  / min_o,    2) if min_o    > 0 else 0.0
        po90    = round(item["ofi"]  / min_o90,  2) if min_o90  > 0 else 0.0
        po365   = round(item["ofi"]  / min_o365, 2) if min_o365 > 0 else 0.0
        pb      = round(item["both"] / min_b,    2) if min_b    > 0 else 0.0

        item.update({
            "k90": k90, "k365": k365,
            "min_k": min_k, "min_k90": min_k90, "min_k365": min_k365,
            "km90": km90_raw, "km365": km365_raw,
            "min_km": min_km, "min_km90": min_km90, "min_km365": min_km365,
            "o90": o90, "o365": o365,
            "min_o": min_o, "min_o90": min_o90, "min_o365": min_o365,
            "b90": b90, "b365": b365, "min_b": min_b,
            "pk": pk, "pk90": pk90, "pk365": pk365,
            "pk_dep": pk_dep,
            "pk_mkt": pk_mkt, "pk_mkt90": pk_mkt90, "pk_mkt365": pk_mkt365,
            "po": po, "po90": po90, "po365": po365, "pb": pb,
        })
    return inv_data



# ─── 5. LOOKUP DESDE INVENTARIO GENERAL ──────────────────────────────────────
PROD_F = os.path.join(DATA_DIR, "Productos", "PRODUCTOS.xlsx")

def load_productos():
    """
    Lee Inventario/Productos.xlsx y devuelve lookup:
      {cod_str: {"art": NOMBRE, "rub": RUBRO, "sub": SUB_RUBRO}}
    Detecta automaticamente el orden de columnas desde la fila de encabezado.
    Columnas soportadas: CODIGO/RUBRO/SUB RUBRO/NOMBRE (cualquier orden).
    """
    lookup = {}
    if not os.path.exists(PROD_F):
        print("  Advertencia: no se encontro Productos.xlsx")
        return lookup
    try:
        import openpyxl
    except ImportError:
        import subprocess, sys
        subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl",
                        "--break-system-packages", "-q"])
        import openpyxl
    try:
        wb = openpyxl.load_workbook(PROD_F, read_only=True, data_only=True)
        sh_name = "DESCRIPCION" if "DESCRIPCION" in wb.sheetnames else wb.sheetnames[0]
        ws = wb[sh_name]
        rows = list(ws.iter_rows(values_only=True))
        # Encontrar fila de encabezado (contiene CODIGO o NOMBRE)
        hdr_idx = None
        col_cod = col_nom = col_rub = col_sub = None
        for i, r in enumerate(rows):
            row_up = [str(c).upper().strip() if c else "" for c in r]
            if any(x in row_up for x in ("CODIGO", "NOMBRE", "CODE")):
                hdr_idx = i
                for j, h in enumerate(row_up):
                    if h in ("CODIGO", "CODE", "COD"): col_cod = j
                    elif h in ("NOMBRE", "DESCRIPCION", "NAME"): col_nom = j
                    elif "SUB" in h: col_sub = j
                    elif h == "RUBRO": col_rub = j
                break
        if hdr_idx is None or col_cod is None:
            print("  Advertencia Productos.xlsx: no se encontro encabezado con CODIGO")
            return lookup
        # Defaults si alguna columna no existe
        if col_nom is None: col_nom = col_cod + 1
        if col_rub is None: col_rub = -1
        if col_sub is None: col_sub = -1
        for r in rows[hdr_idx + 1:]:
            if not r or r[col_cod] is None: continue
            cod_raw = str(r[col_cod]).strip()
            nom_raw = str(r[col_nom]).strip() if col_nom < len(r) and r[col_nom] else ""
            rub_raw = str(r[col_rub]).strip().upper() if col_rub >= 0 and col_rub < len(r) and r[col_rub] else ""
            sub_raw = str(r[col_sub]).strip().upper() if col_sub >= 0 and col_sub < len(r) and r[col_sub] else ""
            if not cod_raw or cod_raw in ("nan","None",""): continue
            # Normalizar codigos numericos
            try:
                cod = str(int(float(cod_raw)))
            except Exception:
                cod = cod_raw
            if not nom_raw or nom_raw in ("nan","None",""): continue
            lookup[cod] = {"art": nom_raw, "rub": rub_raw or "OTROS", "sub": sub_raw}
        wb.close()
        print(f"  Productos: {len(lookup)} articulos cargados desde Productos.xlsx")
    except Exception as e:
        print(f"  Advertencia Productos.xlsx: {e}")
    return lookup


# ─── 6. PROYECCIONES DE COMPRA ───────────────────────────────────────────────
PROY_DIR  = os.path.join(BASE, "Data", "Supply Chain", "proyecciones")
PROY_FILE = os.path.join(PROY_DIR, "proyecciones.xlsx")

def generate_proyecciones(inv_data):
    """
    Genera Data/Supply Chain/proyecciones/proyecciones.xlsx con los productos
    que tienen menos de 3 meses de stock (segun velocidad KLOZER).
    Preserva las columnas de aprobacion si el archivo ya existe.
    Columnas: Codigo | Descripcion | Rubro | Subrubro | Stock | Min/mes | Meses |
              Sug_30d | Sug_60d | Sug_90d | Cant_aprobada | Aprobado | Fecha_aprobacion
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        import subprocess, sys
        subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl",
                        "--break-system-packages", "-q"])
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

    os.makedirs(PROY_DIR, exist_ok=True)

    # Leer aprobaciones existentes para no perderlas al regenerar
    existing = {}   # cod -> {cant_aprobada, aprobado, fecha}
    if os.path.exists(PROY_FILE):
        try:
            wb_old = openpyxl.load_workbook(PROY_FILE, data_only=True)
            ws_old = wb_old.active
            hdrs   = [str(c.value).strip().lower() if c.value else "" for c in next(ws_old.iter_rows())]
            i_cod  = next((i for i, h in enumerate(hdrs) if "cod" in h), None)
            i_cant = next((i for i, h in enumerate(hdrs) if "cant" in h), None)
            i_ap   = next((i for i, h in enumerate(hdrs) if h == "aprobado"), None)
            i_fec  = next((i for i, h in enumerate(hdrs) if "fecha" in h), None)
            if i_cod is not None:
                for row in ws_old.iter_rows(min_row=2, values_only=True):
                    if not row or row[i_cod] is None: continue
                    try:    cod = str(int(float(str(row[i_cod]))))
                    except: cod = str(row[i_cod])
                    existing[cod] = {
                        "cant": row[i_cant] if i_cant is not None else None,
                        "ap":   row[i_ap]   if i_ap   is not None else None,
                        "fec":  row[i_fec]  if i_fec  is not None else None,
                    }
            wb_old.close()
        except Exception as e:
            print(f"    Advertencia proyecciones existentes: {e}")

    # Filtrar productos con < 3 meses de stock KLOZER (solo deposito KLOZER)
    alerta = [it for it in inv_data if it.get("min_k", 0) > 0 and it.get("pk_dep", 0) < 3]
    alerta.sort(key=lambda x: x.get("pk_dep", 0))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Proyecciones"

    HEADERS = ["Codigo", "Descripcion", "Rubro", "Subrubro", "Stock",
               "Min/mes", "Meses", "Sug 30d", "Sug 60d", "Sug 90d",
               "Cant aprobada", "Aprobado", "Fecha aprobacion"]
    WIDTHS  = [10, 38, 16, 16, 10, 10, 9, 10, 10, 10, 14, 11, 20]

    hdr_fill = PatternFill("solid", start_color="1c2230", end_color="1c2230")
    hdr_font = Font(bold=True, color="4ADE80", name="Arial")
    cen = Alignment(horizontal="center")

    for col, (h, w) in enumerate(zip(HEADERS, WIDTHS), 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = hdr_font; c.fill = hdr_fill; c.alignment = cen
        ws.column_dimensions[c.column_letter].width = w

    red_font  = Font(color="EF4444", bold=True, name="Arial")
    yel_font  = Font(color="EAB308", bold=True, name="Arial")
    norm_font = Font(name="Arial")

    for rn, it in enumerate(alerta, 2):
        cod   = str(it["cod"])
        stock = it.get("klozer", 0)           # solo deposito KLOZER
        min_k = round(it.get("min_k", 0), 1)
        meses = round(it.get("pk_dep", 0), 1) # meses de stock KLOZER
        sug30 = max(0, round(min_k * 1 - stock))
        sug60 = max(0, round(min_k * 2 - stock))
        sug90 = max(0, round(min_k * 3 - stock))
        ex    = existing.get(cod, {})

        vals = [int(cod) if cod.isdigit() else cod,
                it.get("art", ""), it.get("rub", ""), it.get("sub", ""),
                stock, min_k, meses,
                sug30, sug60, sug90,
                ex.get("cant"), ex.get("ap"), ex.get("fec")]

        for col, v in enumerate(vals, 1):
            c = ws.cell(row=rn, column=col, value=v)
            c.font = norm_font
        # Color en columna Meses según criticidad
        meses_cell = ws.cell(row=rn, column=7)
        if meses < 1:   meses_cell.font = red_font
        elif meses < 2: meses_cell.font = yel_font

    ws.freeze_panes = "A2"
    wb.save(PROY_FILE)
    print(f"    Proyecciones: {len(alerta)} productos con < 3 meses de stock KLOZER -> {PROY_FILE}")


# ─── 6a. RESUMEN DESTILERÍA / CERVEZAS ────────────────────────────────────────
def fetch_resumen_destileria():
    """
    Descarga la hoja RESUMEN PRODUCTOS X DESTILERIA (gid=1954024258) del mismo
    spreadsheet de costos via CDP y parsea los datos.

    Estructura de la hoja:
      Fila 0 : marcadores de año  (col 3 = "2024", col 9 = "2025", ...)
      Fila 1 : nombre del primer mes de cada año (celdas fusionadas → solo la 1ª visible)
      Fila 2 : DESTILERIA | LINEA | PRODUCTO (cabeceras base)
      Filas 3+: datos de productos; última fila = totales (sin nombre de producto)
      Cada período = 6 cols: [costo_dest, var_dest%, costo_cerv, var_cerv%, total, var_total%]
      Primer período (col 3-8) = Dic 2024; luego Jan 2025, Feb 2025 … hasta mes actual.
    """
    import csv as _csv

    today = date.today()

    # Descargar si no está actualizado hoy
    if os.path.exists(DEST_CSV) and date.fromtimestamp(os.path.getmtime(DEST_CSV)) >= today:
        print("  Destilería resumen al día, usando local")
    else:
        url = (f"https://docs.google.com/spreadsheets/d/{COSTOS_SHEET_ID}/"
               f"gviz/tq?tqx=out:csv&gid={DEST_SHEET_GID}")
        content = _download_costos_via_cdp(url)
        if content and len(content.splitlines()) >= 4:
            os.makedirs(os.path.dirname(DEST_CSV), exist_ok=True)
            with open(DEST_CSV, "w", encoding="utf-8-sig", newline="") as f:
                f.write(content)
            print(f"  Destilería resumen descargado ({len(content.splitlines())} filas)")
        else:
            print("  Destilería resumen: CDP no disponible, usando local si existe")

    if not os.path.exists(DEST_CSV):
        print("  Destilería resumen: sin datos")
        return {"periods": [], "products": [], "error": "Sin datos"}

    with open(DEST_CSV, encoding="utf-8-sig") as f:
        rows = list(_csv.reader(f))

    if len(rows) < 4:
        return {"periods": [], "products": [], "error": "CSV incompleto"}

    n_cols = max(len(r) for r in rows)
    n_periods = (n_cols - 3) // 6

    MONTH_NAMES = ['', 'Ene', 'Feb', 'Mar', 'Abr', 'May', 'Jun',
                   'Jul', 'Ago', 'Sep', 'Oct', 'Nov', 'Dic']

    # Reconstruir secuencia de meses: primer período = Dic 2024
    yr, mo = 2024, 12
    periods = []
    for i in range(n_periods):
        periods.append({"key": f"{yr}_{mo}",
                        "label": f"{MONTH_NAMES[mo]} '{str(yr)[2:]}",
                        "yr": yr, "m": mo})
        mo += 1
        if mo > 12:
            mo = 1
            yr += 1

    def _money(s):
        if not s or not s.strip(): return None
        try: return round(float(s.replace("$","").replace(" ","").replace(",","")), 2)
        except: return None

    def _pct(s):
        if not s or not s.strip(): return None
        try: return round(float(s.replace("%","").replace(" ","")) / 100, 6)
        except: return None

    products = []
    for row in rows[3:]:
        if not row: continue
        dest_name = row[0].strip() if row[0] else ""
        linea     = row[1].strip() if len(row) > 1 else ""
        producto  = row[2].strip() if len(row) > 2 else ""
        if not dest_name or not producto: continue   # fila de totales u otra vacía

        data = {}
        for i, p in enumerate(periods):
            c = 3 + i * 6
            if c + 5 >= len(row): break
            cd = _money(row[c]);   dv = _pct(row[c+1])
            cc = _money(row[c+2]); cv = _pct(row[c+3])
            tt = _money(row[c+4]); tv = _pct(row[c+5])
            if cd is not None or cc is not None or tt is not None:
                data[p["key"]] = {"dest": cd, "dest_var": dv,
                                  "cerv": cc, "cerv_var": cv,
                                  "total": tt, "total_var": tv}
        products.append({"destileria": dest_name, "linea": linea,
                         "producto": producto, "data": data})

    # Quedarse solo con períodos que tengan datos
    used = set()
    for p in products:
        used.update(p["data"].keys())
    periods = [p for p in periods if p["key"] in used]

    last_lbl = periods[-1]["label"] if periods else "—"
    print(f"  Destilería resumen: {len(products)} productos, {len(periods)} períodos (hasta {last_lbl})")
    return {"periods": periods, "products": products, "error": None}


# ─── 6b. COSTOS CERVEZAS ──────────────────────────────────────────────────────
def fetch_cervezas():
    """
    Lee una hoja por mes ("06/26", "05/26", …) del spreadsheet de cervezas.
    De cada hoja extrae col K (índice 10) = Costo Mes ACT para cada lata.
    Construye un array de períodos por producto, ordenado cronológicamente.
    """
    import csv as _csv
    from urllib.parse import quote

    today = date.today()
    CERV_DIR = os.path.join(DATA_DIR, "Costos y PVP", "cervezas_meses")
    os.makedirs(CERV_DIR, exist_ok=True)

    _LATA_ARTS = {
        110027: "CERVEZA TEMPLE WOLF IPA 0% ALCOHOL 473 ML",
        110038: "CERVEZA TEMPLE COSMICA 473 ML",
        110005: "CERVEZA TEMPLE BLACK SOUL STOUT 473 ML",
        110008: "CERVEZA TEMPLE FLOW APA 473 ML",
        110025: "CERVEZA TEMPLE SCOTISH 473 ML",
        110040: "CERVEZA TEMPLE GOLDEN LAGER MUNDIAL 473 ML",
        110012: "CERVEZA TEMPLE INDIE GOLDEN 473 ML",
        110030: "CERVEZA TEMPLE WOLF IPA 473 ML",
    }
    _LATA_CODES = {
        "wolf ipa 0%": 110027, "cosmica": 110038, "black soul stout": 110005,
        "flow apa": 110008, "scottish": 110025, "golden lager mundial": 110040,
        "indie golden": 110012, "wolf ipa": 110030,
    }
    _MESES_ES = ["ENE","FEB","MAR","ABR","MAY","JUN","JUL","AGO","SEP","OCT","NOV","DIC"]

    def _money(s):
        if not s or not s.strip(): return None
        s = s.strip().replace("$","").replace(" ","").replace(".","").replace(",",".")
        try: return round(float(s), 2)
        except: return None

    def _is_csv(text):
        """True si el texto parece CSV real (no HTML ni JSON de error)."""
        if not text: return False
        s = text.lstrip()[:80]
        return not (s.startswith('<') or s.startswith('{') or s.startswith('/*'))

    def _get_all_gids():
        """
        Pide la API gviz JSON del spreadsheet y extrae title->gid de todas las hojas.
        El campo 'p.allSheets' contiene la lista completa cuando hay sesión autenticada.
        """
        import json as _json, re as _re
        url = (f"https://docs.google.com/spreadsheets/d/{CERV_SHEET_ID}/"
               f"gviz/tq?tqx=out:json&headers=0")
        raw = _download_costos_via_cdp(url)
        if not raw: return {}
        m = _re.search(r'setResponse\((.+)\)\s*;?\s*$', raw, _re.DOTALL)
        if not m: return {}
        try:
            data = _json.loads(m.group(1))
        except Exception:
            return {}
        gid_map = {}
        for s in data.get("p", {}).get("allSheets", []):
            sid   = s.get("sheetId")
            title = s.get("title", "")
            if sid and _re.match(r'^\d{2}/\d{2}$', title):
                gid_map[title] = str(sid)
        # Fallback: parse HTML if allSheets not available
        if not gid_map:
            raw2 = _download_costos_via_cdp(
                f"https://docs.google.com/spreadsheets/d/{CERV_SHEET_ID}/edit")
            if raw2:
                for m2 in _re.finditer(r'"(\d{2}(?:/|\\u002F|\\/)(\d{2}))"', raw2):
                    name = m2.group(0).strip('"').replace('\\u002F','/').replace('\\/', '/')
                    ctx_s = max(0, m2.start() - 400)
                    ctx_e = min(len(raw2), m2.end() + 400)
                    gids  = _re.findall(r'\b(\d{7,10})\b', raw2[ctx_s:ctx_e])
                    for g in gids:
                        if not g.startswith('202') and not g.startswith('201') and len(g) >= 7:
                            gid_map.setdefault(name, g)
                            break
        print(f"    GIDs encontrados: {gid_map}")
        return gid_map

    # Descubrir GIDs de hojas mensuales
    _gid_map = _get_all_gids()

    def _get_sheet_rows(sheet_name, gid=None):
        """Descarga (o usa caché) una hoja mensual y retorna filas CSV.
        Preferencia: GID (más confiable) > sheet=NAME (solo planillas públicas)."""
        safe  = sheet_name.replace("/", "")
        cache = os.path.join(CERV_DIR, f"cerv_{safe}.csv")
        try:
            m2, y2 = int(sheet_name[:2]), 2000 + int(sheet_name[3:])
            is_cur = (m2 == today.month and y2 == today.year)
        except:
            is_cur = False
        stale = not os.path.exists(cache) or (is_cur and date.fromtimestamp(os.path.getmtime(cache)) < today)
        # Invalidar cache con HTML guardado por error
        if not stale and os.path.exists(cache):
            with open(cache, encoding="utf-8-sig") as _f:
                if not _is_csv(_f.read(200)):
                    stale = True
        if stale:
            content = None
            # 1) Por GID (el más confiable)
            if gid:
                url = (f"https://docs.google.com/spreadsheets/d/{CERV_SHEET_ID}/"
                       f"gviz/tq?tqx=out:csv&gid={gid}")
                content = _download_costos_via_cdp(url)
                if not _is_csv(content): content = None
            # 2) Por nombre sin encodear (fallback)
            if not content:
                url = (f"https://docs.google.com/spreadsheets/d/{CERV_SHEET_ID}/"
                       f"gviz/tq?tqx=out:csv&sheet={sheet_name}")
                content = _download_costos_via_cdp(url)
                if not _is_csv(content): content = None
            # 3) Por nombre URL-encodeado (otro fallback)
            if not content:
                url = (f"https://docs.google.com/spreadsheets/d/{CERV_SHEET_ID}/"
                       f"gviz/tq?tqx=out:csv&sheet={quote(sheet_name, safe='')}")
                content = _download_costos_via_cdp(url)
                if not _is_csv(content): content = None
            if content and len(content.splitlines()) >= 5:
                with open(cache, "w", encoding="utf-8-sig", newline="") as f:
                    f.write(content)
            else:
                return None
        if not os.path.exists(cache): return None
        with open(cache, encoding="utf-8-sig") as f:
            return list(_csv.reader(f))

    def _find_cost_col(row):
        """Retorna el índice de la celda que diga 'Costo Mes ACT' (case-insensitive)."""
        for i, cell in enumerate(row):
            if "costo mes act" in str(cell).lower().replace(".", "").strip():
                return i
        return None

    def _extract_lata_col_k(rows):
        """
        Busca la columna 'Costo Mes ACT' por nombre de encabezado en la sección LATA.
        Si no la encuentra por nombre, usa col 10 como fallback (posición conocida).
        """
        costs, fasons = {}, {}
        in_lata  = False
        cost_col = None     # se setea al encontrar el header "Costo Mes ACT"

        # Pre-scan: buscar "Costo Mes ACT" en las primeras 20 filas (headers globales)
        for row in rows[:20]:
            col = _find_cost_col(row)
            if col is not None:
                cost_col = col
                break

        for row in rows:
            c0 = row[0].strip() if row else ""
            c1 = row[1].strip() if len(row) > 1 else ""

            if "LATA" in c0.upper() and "RESUMEN" in c0.upper():
                if c1 and c1 != "Fason":
                    if costs: break
                in_lata = True
                # Puede que el header "Costo Mes ACT" esté en esta misma fila
                col = _find_cost_col(row)
                if col is not None:
                    cost_col = col
                continue

            if not in_lata: continue

            # Si aún no encontramos el header, chequearlo en la siguiente fila
            if cost_col is None:
                col = _find_cost_col(row)
                if col is not None:
                    cost_col = col
                    continue   # era fila de encabezado, no de datos

            if not c0: continue

            idx   = cost_col if cost_col is not None else 10
            costo = _money(row[idx]) if idx < len(row) else None
            if costo is None: continue

            norm = c0.lower().strip()
            for key in sorted(_LATA_CODES, key=len, reverse=True):
                if key in norm:
                    # Primer match gana: evita sobrescribir con filas de referencia/resumen
                    if key not in costs:
                        costs[key]  = costo
                        fasons[key] = c1.upper()
                    break
        return costs, fasons

    # ── Descargar hojas mensuales ─────────────────────────────────────────────
    month_costs  = {}
    month_fasons = {}
    yr, mo = today.year, today.month
    fails  = 0

    known_sheets = set(_gid_map.keys())
    # Calcular el mes más antiguo descubierto vía GID (para saber desde dónde explorar más)
    if known_sheets:
        oldest = min(known_sheets, key=lambda s: (2000 + int(s[3:]), int(s[:2])))
        oldest_yr  = 2000 + int(oldest[3:])
        oldest_mo  = int(oldest[:2])
    else:
        oldest_yr, oldest_mo = yr, mo

    for _ in range(48):  # hasta 4 años atrás
        sheet = f"{mo:02d}/{str(yr)[2:]}"
        gid   = _gid_map.get(sheet)

        # Fase 1: mientras haya meses en el mapa GID, usarlos; saltar los que no están
        in_gid_range = (yr > oldest_yr) or (yr == oldest_yr and mo >= oldest_mo)
        if known_sheets and in_gid_range and not gid:
            mo -= 1
            if mo == 0: mo, yr = 12, yr - 1
            continue

        rows = _get_sheet_rows(sheet, gid=gid)
        if rows:
            c, f = _extract_lata_col_k(rows)
            if c:
                month_costs[sheet]  = c
                month_fasons[sheet] = f
                fails = 0
                print(f"    {sheet}: {len(c)} productos")
            else:
                fails += 1
                print(f"    {sheet}: sin datos lata")
        else:
            fails += 1
        # Fuera del rango GID conocido, parar tras 3 fallos consecutivos
        if not in_gid_range and fails >= 3:
            break
        mo -= 1
        if mo == 0: mo, yr = 12, yr - 1

    # Fallback: si todo falló, usar el CSV del GID conocido (Analisis_costos_cervezas.csv)
    if not month_costs and os.path.exists(CERV_CSV):
        print("  Cervezas: fallback → CSV existente (gid conocido)")
        with open(CERV_CSV, encoding="utf-8-sig") as f:
            fb_rows = list(_csv.reader(f))
        c, faz = _extract_lata_col_k(fb_rows)
        if c:
            sheet_label = f"{today.month:02d}/{str(today.year)[2:]}"
            month_costs[sheet_label]  = c
            month_fasons[sheet_label] = faz

    if not month_costs:
        print("  Cervezas: sin datos")
        return {"barril": [], "lata": [], "error": "Sin datos"}

    # Ordenar cronológicamente (más antiguo primero)
    sorted_sheets = sorted(month_costs, key=lambda s: (2000 + int(s[3:]), int(s[:2])))
    last_sheet    = sorted_sheets[-1]
    fasons        = month_fasons.get(last_sheet, {})

    # ── Construir items de lata ───────────────────────────────────────────────
    lata = []
    for key, cod in sorted(_LATA_CODES.items(), key=lambda x: x[1]):
        periods = []
        prev    = None
        for sheet in sorted_sheets:
            costo = month_costs[sheet].get(key)
            if costo is None: continue
            m  = int(sheet[:2])
            yy = sheet[3:]
            label    = f"{_MESES_ES[m-1]} {yy}"
            costo_var = round((costo - prev) / prev, 4) if prev else None
            periods.append({"label": label, "costo": costo, "costo_var": costo_var})
            prev = costo
        if not periods: continue
        lata.append({
            "producto":  key,
            "fason":     fasons.get(key, ""),
            "art":       _LATA_ARTS[cod],
            "cod":       cod,
            "costo_fab": periods[-1]["costo"],
            "costo_ant": periods[-2]["costo"] if len(periods) >= 2 else None,
            "costo_var": periods[-1]["costo_var"],
            "periods":   periods,
        })

    n = len(sorted_sheets)
    print(f"  Cervezas: {len(lata)} latas · {n} períodos ({sorted_sheets[0]} → {sorted_sheets[-1]})")
    return {"barril": [], "lata": lata, "total_barril": None, "total_lata": None, "error": None}


# ─── 6c. ACTUALIZAR STOCK_CIERRE_MES EN DASHBOARD HTML ────────────────────────
def update_stock_cierre_mes():
    """
    Lee Stock_consolidado_por_deposito_y_dia.xlsx (KLOZER+OFI), encuentra el
    último día disponible por mes, suma stock por código y reemplaza el bloque
    STOCK_CIERRE_MES en bosquegin_dashboard.html.
    """
    import re as _re
    try:
        import openpyxl
    except ImportError:
        import subprocess, sys
        subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl",
                        "--break-system-packages", "-q"])
        import openpyxl

    DASHBOARD = os.path.join(BASE, "bosquegin_dashboard.html")
    if not os.path.exists(DASHBOARD):
        print("  Dashboard HTML no encontrado — omitiendo (modo cloud?)")
        return
    DEPS_OK   = {"KLOZER", "OFI", "OFICINA"}

    wb = openpyxl.load_workbook(CONS_FILE, read_only=True, data_only=True)
    ws = wb.active

    raw_rows = []
    for row in ws.iter_rows(values_only=True):
        if not row or row[0] == "fecha": continue
        dep = str(row[1]).strip().upper() if row[1] else ""
        if dep not in DEPS_OK: continue
        fecha_raw = row[0]
        cod_raw   = row[3]
        qty_raw   = row[5]
        if cod_raw is None or qty_raw is None: continue
        try:
            cod = str(int(float(str(cod_raw))))
            if not (5 <= len(cod) <= 6): continue
        except Exception:
            continue
        try:
            qty = float(str(qty_raw).replace(",", "."))
        except Exception:
            continue
        if qty <= 0: continue
        fecha = str(fecha_raw)[:10] if fecha_raw else ""
        if not fecha: continue
        raw_rows.append((fecha, cod, int(qty)))

    wb.close()

    # ── Contabilium stock EN VIVO (vía API REST) ────────────────────────────
    try:
        import contabilium_api as _cont_api
        _hoy = date.today().strftime("%Y-%m-%d")
        _stock_vivo = _cont_api.get_stock_todos_depositos(["KLOZER", "OFICINA"])
        _cont_count = 0
        for dep_nombre, items in _stock_vivo.items():
            for cod, qty in items.items():
                if qty <= 0: continue
                if not (5 <= len(cod) <= 6): continue
                raw_rows.append((_hoy, cod, int(qty)))
                _cont_count += 1
        print(f"  Contabilium stock EN VIVO: {_cont_count} filas ({_hoy})")
    except Exception as e:
        print(f"  Advertencia Contabilium API stock: {e}")

    if not raw_rows:
        print("  update_stock_cierre_mes: sin filas KLOZER+OFI válidas")
        return

    # Último día disponible por mes
    last_day = {}
    for fecha, cod, qty in raw_rows:
        parts = fecha.split("-")
        key = "%s_%s" % (int(parts[0]), int(parts[1]))
        if key not in last_day or fecha > last_day[key]:
            last_day[key] = fecha

    # Sumar stock por código en el último día de cada mes
    stock = {}
    for fecha, cod, qty in raw_rows:
        parts = fecha.split("-")
        key = "%s_%s" % (int(parts[0]), int(parts[1]))
        if fecha != last_day[key]: continue
        stock.setdefault(key, {})
        stock[key][cod] = stock[key].get(cod, 0) + qty

    month_keys = sorted(stock.keys(), key=lambda k: (int(k.split("_")[0]), int(k.split("_")[1])))

    # ── Leer meses existentes de data_stock_cierre.js para preservar historicos ──
    scm_path = os.path.join(BASE, "data_stock_cierre.js")
    existing_months = {}   # key -> raw JSON block string  (ej: '{"100001": 802, ...}')
    if os.path.exists(scm_path):
        try:
            old_js = open(scm_path, encoding="utf-8").read()
            for m_key, m_body in _re.findall(r'"(\d{4}_\d{1,2})"\s*:\s*(\{[^}]*\})', old_js):
                existing_months[m_key] = m_body
        except Exception as e:
            print("  Advertencia leyendo stock_cierre existente: %s" % e)

    # Los meses del Excel reemplazan los existentes; el resto se conserva
    all_keys = sorted(
        set(list(existing_months.keys()) + list(stock.keys())),
        key=lambda k: (int(k.split("_")[0]), int(k.split("_")[1]))
    )

    # Generar bloque JS (meses frescos del Excel o conservados del archivo)
    js_lines = []
    for i, key in enumerate(all_keys):
        comma_after = "," if i < len(all_keys) - 1 else ""
        if key in stock:
            # Datos frescos del Excel
            day  = last_day[key]
            data = stock[key]
            js_lines.append("  // %s: Stock_consolidado_por_deposito_y_dia.xlsx — KLOZER+OFI — %s" % (key, day))
            sorted_cods = sorted(data.keys())
            entries = ['"%s": %d' % (c, data[c]) for c in sorted_cods]
            chunk = 4
            entry_lines = []
            for j in range(0, len(entries), chunk):
                sl = entries[j:j+chunk]
                entry_lines.append("    %s%s" % (", ".join(sl), "," if j + chunk < len(entries) else ""))
            js_lines.append('  "%s": {\n%s\n  }%s' % (key, "\n".join(entry_lines), comma_after))
        else:
            # Mes histórico conservado tal cual
            js_lines.append('  "%s": %s%s' % (key, existing_months[key], comma_after))

    new_block = "\n".join(js_lines)

    # Escribir data_stock_cierre.js
    scm_js = "window.STOCK_CIERRE_MES={\n" + new_block + "\n};"
    with open(scm_path, "w", encoding="utf-8") as f:
        f.write(scm_js)
    fresh = [k for k in all_keys if k in stock]
    kept  = [k for k in all_keys if k not in stock]
    print("  data_stock_cierre.js: %d meses frescos (%s), %d historicos preservados (%s)" % (
        len(fresh), ", ".join(fresh), len(kept), ", ".join(kept)))
    print("  Último snapshot: %s -> %s" % (month_keys[-1], last_day[month_keys[-1]]))


# ─── 7. MAIN ──────────────────────────────────────────────────────────────────
def main():
    print("=" * 60)
    print(f"Actualizando Bosque Gin Dashboard — {datetime.now(_AR):%Y-%m-%d %H:%M} (AR)")
    print("=" * 60)

    # Resumen de salud: cada paso se registra acá para que los problemas se vean
    # de inmediato al final, en vez de quedar enterrados en 100 líneas de log.
    _health = []
    def _ok(nombre, detalle=""):
        _health.append((nombre, True, detalle))
    def _fail(nombre, err):
        _health.append((nombre, False, str(err)))

    # Configurar CDP en Brave (una vez; si ya está configurado no hace nada)
    try:
        import subprocess
        ps1 = os.path.join(BASE, "setup_brave_cdp.ps1")
        subprocess.run(
            ["powershell", "-NonInteractive", "-ExecutionPolicy", "Bypass",
             "-File", ps1],
            capture_output=True, timeout=15,
            creationflags=subprocess.CREATE_NO_WINDOW
        )
    except Exception:
        pass

    # Gestión Cervecera (GC) fue reemplazada por Contabilium desde 2026-06-29.
    # La descarga automática de GC falla siempre (sesión/acceso dado de baja) —
    # deshabilitada para no perder tiempo ni ensuciar el log en cada corrida.
    # Los archivos GC históricos en Data/Salidas/GC siguen usándose igual
    # (ver update_salidas_consolidado()). Para reactivar: GC_DOWNLOAD_ENABLED = True.
    GC_DOWNLOAD_ENABLED = False
    if GC_DOWNLOAD_ENABLED:
        print("\n[0/5] Descargando archivos de Gestión Cervecera...")
        try:
            from gc_downloader import descargar_todo
            _last_d = _last_ventas_date()
            if _last_d:
                _gc_desde = date.fromisoformat(_last_d) + timedelta(days=1)
                print(f"  Bosque salidas hasta {_last_d} -> GC descarga desde {_gc_desde}")
            else:
                _gc_desde = None
            descargar_todo(verbose=True, fecha_desde_date=_gc_desde)
        except Exception as e:
            print(f"  Advertencia GC downloader: {e}")
            print("  Continuando con los archivos existentes...")

    print("\n[0d/5] Descargando historial de salidas de Contabilium...")
    try:
        from contabilium_downloader import descargar_historial as _cont_descargar_historial
        _fpath = _cont_descargar_historial(verbose=True)
        _ok("Historial Contabilium", os.path.basename(_fpath) if _fpath else "sin novedades")
    except Exception as e:
        print(f"  Advertencia Contabilium downloader: {e}")
        print("  Continuando con los archivos existentes...")
        _fail("Historial Contabilium", e)
    # Nota: el stock ya NO se descarga por CDP/export — update_stock_cierre_mes()
    # lo trae en vivo desde contabilium_api.get_stock_todos_depositos()

    # Destilería desde hoja RESUMEN PRODUCTOS X DESTILERIA (Google Sheets)
    print("\n[0b/5] Descargando resumen destilería/cervezas...")
    destileria = {}
    try:
        destileria = fetch_resumen_destileria()
        _n_per = len(destileria.get("periods", destileria.get("months", [])))
        _ok("Destilería resumen", f"{_n_per} períodos")
    except Exception as e:
        print("  Advertencia destilería resumen: %s" % e)
        _fail("Destilería resumen", e)
        # Fallback: conservar datos anteriores si existen
        if os.path.exists(OUT_JS):
            try:
                raw     = open(OUT_JS, encoding="utf-8").read()
                js_body = raw.strip()[len("var BOSQUE_DATA="):-1]
                old     = json.loads(js_body)
                destileria = old.get("destileria", {})
            except Exception:
                pass

    # Cervezas (costos por producto BARRIL y LATA)
    print("\n[0c/5] Descargando costos cervezas...")
    cervezas = {}
    try:
        cervezas = fetch_cervezas()
        _ok("Costos cervezas", f"{len(cervezas.get('periodos', cervezas.get('periods', [])))} períodos")
    except Exception as e:
        print(f"  Advertencia cervezas: {e}")
        _fail("Costos cervezas", e)
        if os.path.exists(OUT_JS):
            try:
                raw = open(OUT_JS, encoding="utf-8").read()
                js_body = raw.strip()[len("var BOSQUE_DATA="):-1]
                old = json.loads(js_body)
                cervezas = old.get("cervezas", {})
            except Exception:
                pass

    print("\n[1/5] Actualizando consolidado de inventario...")
    try:
        update_consolidado()
        _ok("Consolidado inventario")
    except Exception as e:
        print("  Advertencia consolidado: %s" % e)
        _fail("Consolidado inventario", e)

    print("\n[1b/5] Actualizando STOCK_CIERRE_MES en dashboard...")
    try:
        update_stock_cierre_mes()
        _ok("Stock cierre mes")
    except Exception as e:
        print("  Advertencia stock cierre mes: %s" % e)
        _fail("Stock cierre mes", e)

    print("\n[2/5] Cargando productos (rubro/subrubro)...")
    try:
        inv_gen = load_productos()
        print("  -> %d productos cargados desde PRODUCTOS.xlsx" % len(inv_gen))
        _ok("Productos", f"{len(inv_gen)} productos")
    except Exception as e:
        print("  Advertencia productos: %s" % e)
        _fail("Productos", e)
        inv_gen = {}

    print("\n[3/5] Leyendo inventario desde consolidado...")
    try:
        inv_data, stock_hasta = parse_stock(INV_DIR)
        print("  -> %d articulos al %s" % (len(inv_data), stock_hasta))
        if inv_gen:
            for item in inv_data:
                gen = inv_gen.get(str(item["cod"]))
                if gen:
                    if gen["art"]: item["art"] = gen["art"]
                    if gen["rub"]: item["rub"] = gen["rub"]
                    if gen["sub"]: item["sub"] = gen["sub"]
        _ok("Stock consolidado (histórico)", f"{len(inv_data)} artículos al {stock_hasta}")
    except Exception as e:
        print("  ERROR stock: %s" % e)
        _fail("Stock consolidado (histórico)", e)
        inv_data, stock_hasta = [], date.today().strftime("%Y-%m-%d")

    print("\n[3b/5] Sobreescribiendo stock EN VIVO desde Contabilium...")
    try:
        inv_data, stock_hasta = apply_stock_contabilium_vivo(inv_data, inv_gen)
        print("  -> stock actualizado al %s (vía API Contabilium)" % stock_hasta)
        _ok("Stock EN VIVO (API)", f"al {stock_hasta}")
    except Exception as e:
        print("  Advertencia stock vivo Contabilium: %s" % e)
        _fail("Stock EN VIVO (API)", e)

    print("\n[4/5] Actualizando costos desde Google Sheets...")
    try:
        costos, costos_data = fetch_costos_completo()   # descarga automática incluida
        _ok("Costos")
    except Exception as e:
        print("  Advertencia costos: %s" % e)
        _fail("Costos", e)
        costos, costos_data = {}, {"periodos": [], "productos": [], "actualizacion": "", "error": str(e)}

    print("\n[4b/5] Actualizando consolidado de salidas...")
    try:
        _max_fecha_salidas = update_salidas_consolidado()
        if _max_fecha_salidas and _max_fecha_salidas != "0":
            _dias_atraso = (date.today() - date.fromisoformat(_max_fecha_salidas)).days
            if _dias_atraso > 1:
                _fail("Consolidado salidas", f"desactualizado — última fecha {_max_fecha_salidas} ({_dias_atraso} días atrás, ¿archivo bloqueado?)")
            else:
                _ok("Consolidado salidas", f"hasta {_max_fecha_salidas}")
        else:
            _ok("Consolidado salidas")
    except Exception as e:
        print("  Advertencia consolidado salidas: %s" % e)
        _fail("Consolidado salidas", e)

    print("\n[4c/5] Leyendo salidas desde consolidado...")
    salidas_src = SALIDAS_CONS if os.path.exists(SALIDAS_CONS) else VENTAS_F
    try:
        ventas, ventas_hasta, monthly_raw = parse_ventas(salidas_src, prod_lookup=inv_gen, costos=costos)
        print("  -> Datos hasta %s (fuente: %s)" % (ventas_hasta, os.path.basename(salidas_src)))
        _ok("Ventas/salidas", f"hasta {ventas_hasta}")
    except Exception as e:
        print("  ERROR ventas: %s" % e)
        _fail("Ventas/salidas", e)
        ventas, ventas_hasta, monthly_raw = {"VD": {}, "MONTHLY_DATA": {}, "PROD_DATA": {}}, "?", {}

    print("\n[4d/5] Cargando inventario insumos...")
    try:
        insumos_data = fetch_insumos()
        if insumos_data.get("error"):
            _fail("Insumos", insumos_data["error"])
        else:
            _ok("Insumos", f"{len(insumos_data.get('items', []))} items")
    except Exception as e:
        print(f"  Advertencia insumos: {e}")
        _fail("Insumos", e)
        insumos_data = {"items": [], "meses": [], "error": str(e)}

    print("\n[5/5] Calculando velocidad de ventas por deposito...")
    try:
        ref = date.today()
        vel = compute_velocity(monthly_raw, ref)
        inv_data = apply_velocity(inv_data, vel)
        con_vel = sum(1 for x in inv_data if x["min_k"] > 0 or x["min_o"] > 0)
        print("  -> %d articulos con velocidad calculada" % con_vel)
        _ok("Velocidad de ventas", f"{con_vel} artículos")
    except Exception as e:
        print("  ERROR velocidad: %s" % e)
        _fail("Velocidad de ventas", e)

    new_data = {
        "meta": {
            "generado":         datetime.now(_AR).strftime("%Y-%m-%dT%H:%M:%S-03:00"),
            "ventas_hasta":     ventas_hasta,
            "stock_hasta":      stock_hasta,
            "destileria_hasta": max((m["key"] for m in destileria.get("months", [])), default="?"),
        },
        "ventas":    ventas,
        "stock":     {"INV_DATA": inv_data},
        "destileria": destileria,
        "cervezas":  cervezas,
        "costos":    costos_data,
        "insumos":   insumos_data,
    }

    js_str = "var BOSQUE_DATA=" + json.dumps(new_data, ensure_ascii=False) + ";"
    with open(OUT_JS, "w", encoding="utf-8") as f:
        f.write(js_str)
    print("\nOK: bosquegin_data.js actualizado (%d chars)" % len(js_str))
    print("  Stock:   %d articulos al %s" % (len(inv_data), stock_hasta))
    print("  Ventas:  hasta %s" % ventas_hasta)

    # ── Archivos por sección (lazy loading) ───────────────────────────────────
    # CLIENT_DATA se separa de data_ventas.js: pesa ~85% del archivo (4+ MB) y solo
    # lo usa la pestaña Ventas al filtrar por cliente — overview/calendario/logistica
    # no lo necesitan nunca. Separarlo evita bajar 4+ MB de más en cada visita.
    _ventas_liviano = {k: v for k, v in new_data["ventas"].items() if k != "CLIENT_DATA"}
    _client_data    = new_data["ventas"].get("CLIENT_DATA", {})

    _b = "window.BOSQUE_DATA=window.BOSQUE_DATA||{};"
    _bv = _b + "window.BOSQUE_DATA.ventas=window.BOSQUE_DATA.ventas||{};"
    _section_files = {
        "data_meta.js":       _b  + "window.BOSQUE_DATA.meta="       + json.dumps(new_data["meta"],       ensure_ascii=False) + ";",
        "data_ventas.js":     _bv + "Object.assign(window.BOSQUE_DATA.ventas," + json.dumps(_ventas_liviano, ensure_ascii=False) + ");",
        "data_clientes.js":   _bv + "window.BOSQUE_DATA.ventas.CLIENT_DATA=" + json.dumps(_client_data, ensure_ascii=False) + ";",
        "data_stock.js":      _b  + "window.BOSQUE_DATA.stock="      + json.dumps(new_data["stock"],      ensure_ascii=False) + ";",
        "data_destileria.js": _b  + "window.BOSQUE_DATA.destileria=" + json.dumps(new_data["destileria"], ensure_ascii=False) + ";window.BOSQUE_DATA.cervezas=" + json.dumps(new_data["cervezas"], ensure_ascii=False) + ";",
        "data_costos.js":     _b  + "window.BOSQUE_DATA.costos="     + json.dumps(new_data["costos"],     ensure_ascii=False) + ";",
        "data_insumos.js":    _b  + "window.BOSQUE_DATA.insumos="    + json.dumps(new_data["insumos"],    ensure_ascii=False) + ";",
    }
    print("\nArchivos por sección:")
    for fname, content in _section_files.items():
        fpath = os.path.join(BASE, fname)
        with open(fpath, "w", encoding="utf-8") as f:
            f.write(content)
        print("  %s (%d KB)" % (fname, len(content) // 1024))

    print("\n[6/6] Generando proyecciones de compra...")
    try:
        generate_proyecciones(inv_data)
    except Exception as e:
        print("  Advertencia proyecciones: %s" % e)

    print("Recarga el dashboard para ver los cambios.")

    def _print_health_summary():
        fails = [h for h in _health if not h[1]]
        oks   = [h for h in _health if h[1]]
        print("\n" + "=" * 60)
        if fails:
            print("RESUMEN: %d de %d pasos OK — %d con problemas:" % (len(oks), len(_health), len(fails)))
            for nombre, _, detalle in fails:
                print("  ❌ %s — %s" % (nombre, str(detalle)[:150]))
        else:
            print("RESUMEN: %d/%d pasos OK — todo en orden." % (len(oks), len(_health)))
        print("=" * 60)

    # ── Auto-publicar en bosquegin.com via GitHub Pages ──────────────────────
    print("\n[7/7] Publicando datos en bosquegin.com...")

    if _SKIP_GIT_PUSH:
        print("  [cloud] Git push omitido — lo maneja el servidor cloud")
        _print_health_summary()
        return

    try:
        import subprocess

        def _run(cmd):
            r = subprocess.run(cmd, cwd=BASE, capture_output=True, text=True)
            return r.returncode, (r.stdout + r.stderr)

        for cmd in [
            ["git", "add", "bosquegin_data.js", "data_meta.js", "data_ventas.js", "data_clientes.js",
              "data_stock.js", "data_stock_cierre.js", "data_destileria.js", "data_costos.js", "data_insumos.js",
              "auth_static.js", "bosquegin_dashboard.html"],
            ["git", "commit", "-m", "data: actualizar " + datetime.now(_AR).strftime("%Y-%m-%d %H:%M")],
        ]:
            code, out = _run(cmd)
            if code != 0 and "nothing to commit" not in out:
                print("  Advertencia git (%s): %s" % (" ".join(cmd[1:2]), out.strip()[:120]))
            else:
                print("  OK: %s" % " ".join(cmd[1:2]))

        # Push directo primero (repo de un solo operador — casi nunca diverge).
        # Sólo si el remoto avanzó de verdad se intenta pull --rebase y reintenta.
        # Antes se hacía pull --rebase SIEMPRE, y fallaba cada vez por los cambios
        # sin commitear de Data/*.xlsx (que este script no agrega) — puro ruido.
        code, out = _run(["git", "push", "origin", "main"])
        if code != 0 and ("rejected" in out or "non-fast-forward" in out):
            print("  Remoto avanzó — haciendo pull --rebase y reintentando push...")
            code2, out2 = _run(["git", "pull", "--rebase", "origin", "main"])
            if code2 != 0:
                print("  Advertencia git (pull --rebase): %s" % out2.strip()[:200])
            code, out = _run(["git", "push", "origin", "main"])

        if code != 0:
            print("  Advertencia git (push): %s" % out.strip()[:200])
            _fail("Publicación (git push)", out.strip()[:200])
        else:
            print("  OK: push")
            _ok("Publicación (git push)")
        print("  -> bosquegin.com se actualiza en ~60 segundos")
    except Exception as e:
        print("  Advertencia publicacion: %s" % e)
        _fail("Publicación (git push)", e)

    # ── Sincronizar código con Google Drive ───────────────────────────────────
    _sync_to_drive()

    _print_health_summary()


def _sync_to_drive():
    """Copia los archivos clave del tablero a la carpeta de Google Drive
    (G:/Mi unidad/Claude/Tablero operativo/) para que Drive Desktop los suba
    automáticamente a la nube. Silencioso — no interrumpe si Drive no está."""
    import shutil as _sh

    DRIVE_DEST = r"G:\Mi unidad\Claude\Tablero operativo"
    if not os.path.isdir(DRIVE_DEST):
        return   # Drive Desktop no montado — saltar silenciosamente

    FILES = [
        "bosquegin_dashboard.html",
        "actualizar_bosquegin.py",
        "gc_downloader.py",
        "servidor_render.py",
        "actualizar_cloud.py",
        "servidor_bosquegin.py",
        "servidor_bosquegin_bg.bat",
        "inicio_silencioso.vbs",
        "iniciar_tablero.bat",
        "iniciar_tablero_silencioso.vbs",
        "cloud_config.js",
        "render.yaml",
        "requirements.txt",
        "auth_static.js",
        "index.html",
        "oauth_tokens.json",
        "client_secret.json",
        "get_oauth_token.py",
        "setup_brave_cdp.ps1",
    ]

    synced = 0
    for fname in FILES:
        src = os.path.join(BASE, fname)
        dst = os.path.join(DRIVE_DEST, fname)
        if not os.path.exists(src):
            continue
        try:
            # Solo copiar si cambió (por tamaño o fecha)
            if (not os.path.exists(dst)
                    or os.path.getsize(src) != os.path.getsize(dst)
                    or os.path.getmtime(src) > os.path.getmtime(dst)):
                _sh.copy2(src, dst)
                synced += 1
        except Exception:
            pass

    # Sincronizar carpeta Versiones
    vers_src = os.path.join(BASE, "Versiones")
    vers_dst = os.path.join(DRIVE_DEST, "Versiones")
    if os.path.isdir(vers_src):
        try:
            _sh.copytree(vers_src, vers_dst, dirs_exist_ok=True)
        except Exception:
            pass

    if synced > 0:
        print(f"\n[Drive] {synced} archivo(s) sincronizados -> {DRIVE_DEST}")


if __name__ == "__main__":
    main()
