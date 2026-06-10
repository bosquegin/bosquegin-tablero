#!/usr/bin/env python3
"""
actualizar_bosquegin.py
Reconstruye bosquegin_data.js leyendo:
  - Stock:  Tablero operativo/Inventario/*.xlsx (archivos diarios)
  - Ventas: Tablero operativo/Ventas/Bosque ventas.xlsx (hoja DATOS)
  - Destileria: conserva la seccion existente en bosquegin_data.js
"""
import os, json, re, glob
from datetime import datetime, date, timedelta

import platform
if platform.system() == "Windows":
    BASE = r"C:\Users\SupplyDestileria\Documents\Bosque\Claude\Tablero operativo"
else:
    BASE = "/sessions/hopeful-quirky-archimedes/mnt/Tablero operativo"

DATA_DIR = os.path.join(BASE, "Data")
INV_DIR  = os.path.join(DATA_DIR, "Inventario")
VENTAS_F = os.path.join(DATA_DIR, "Salidas", "Bosque salidas.xlsx")
OUT_JS   = os.path.join(BASE, "bosquegin_data.js")

MESES_NOM = ['','Ene','Feb','Mar','Abr','May','Jun','Jul','Ago','Sep','Oct','Nov','Dic']

def do_mg(m, c):
    return round((m - c) / m * 100, 1) if m > 0 else 0.0

def parse_file_date(fname):
    """
    Extrae fecha del nombre. Formatos soportados:
      - d-m-yyyy  o  d/m/yyyy  (ej: 8-5-2026)
      - yyyy-mm-dd              (ej: 2026-05-08)
    Devuelve string 'yyyy-mm-dd' o date.today() como fallback.
    """
    # Intenta d-m-yyyy (formato local argentino)
    m = re.search(r'(\d{1,2})[-./](\d{1,2})[-./](\d{4})', fname)
    if m:
        d_n, mo_n, yr_n = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if 1 <= mo_n <= 12 and 1 <= d_n <= 31 and 2020 <= yr_n <= 2030:
            return f"{yr_n}-{mo_n:02d}-{d_n:02d}"
    # Intenta yyyy-mm-dd
    m = re.search(r'(\d{4})[-./](\d{2})[-./](\d{2})', fname)
    if m:
        return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
    return date.today().strftime("%Y-%m-%d")


# ─── 1. ACTUALIZAR CONSOLIDADO ────────────────────────────────────────────────
CONS_FILE = os.path.join(INV_DIR, "Stock_consolidado_por_deposito_y_dia.xlsx")

def update_consolidado():
    """
    Agrega al archivo consolidado las filas de los archivos diarios
    que sean mas recientes que la ultima fecha ya registrada.
    """
    try:
        import openpyxl
    except ImportError:
        import subprocess, sys
        subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl",
                        "--break-system-packages", "-q"])
        import openpyxl

    # Leer ultima fecha, lookup y fechas ya presentes (para evitar duplicados)
    wb0 = openpyxl.load_workbook(CONS_FILE, read_only=True, data_only=True)
    ws0 = wb0.active
    lookup        = {}
    last_date     = "0"
    existing_keys = set()   # (fecha, dep, cod) ya en el consolidado
    for r in ws0.iter_rows(values_only=True):
        if not r or r[0] == "fecha": continue
        if r[3] is not None:
            try:
                cod = str(int(float(str(r[3]))))
                if cod not in lookup:
                    lookup[cod] = (r[6] or "", r[7] or "", r[2] or "")
                fecha_k = str(r[0])[:10] if r[0] else ""
                dep_k   = str(r[1]).strip().upper() if r[1] else ""
                if fecha_k and dep_k:
                    existing_keys.add((fecha_k, dep_k, cod))
            except Exception:
                pass
        if r[0]:
            d = str(r[0])[:10]
            if d > last_date: last_date = d
    wb0.close()
    print(f"    Consolidado al {last_date} | {len(lookup)} codigos conocidos")

    # Archivos diarios mas nuevos
    all_daily = (sorted(glob.glob(os.path.join(INV_DIR, "Stock productos*.xlsx"))) +
                 sorted(glob.glob(os.path.join(INV_DIR, "Stock Productos Comp*.xlsx"))))
    new_files = []
    for fp in all_daily:
        fd = parse_file_date(os.path.basename(fp))
        if fd and fd > last_date:
            new_files.append((fd, fp))
    new_files.sort()

    if not new_files:
        print("    Consolidado ya esta al dia.")
        return last_date

    print(f"    Agregando {len(new_files)} archivo(s) nuevos...")
    new_rows = []

    for file_date, fpath in new_files:
        fname = os.path.basename(fpath)
        try:
            wb2 = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
        except Exception as e:
            print(f"      Salteo {fname}: {e}")
            continue
        ws2 = wb2.active
        rows = list(ws2.iter_rows(values_only=True))

        # Encontrar fila de cabeceras reales
        hdr_idx = None
        for i, r in enumerate(rows[:5]):
            if r and r[0] and str(r[0]).strip().upper().replace("O","O") in ("DEPOSITO", "DEP\xd3SITO", "DEPÓSITO"):
                hdr_idx = i; break
        if hdr_idx is None:
            wb2.close(); continue

        hdr_str = " ".join(str(c).upper() if c else "" for c in rows[hdr_idx])
        is_main = "CANTIDAD" in hdr_str or "ENVASE" in hdr_str
        current_dep = None

        for row in rows[hdr_idx + 1:]:
            if not row: continue
            dep_cell = row[0] if len(row) > 0 else None
            if dep_cell and str(dep_cell).strip() not in ("", "None"):
                raw = str(dep_cell).strip().upper()
                if not re.match(r'^[WwLl]\d', raw) and not re.match(r'^\d+\.?\d*$', raw):
                    current_dep = raw

            if is_main:
                cod_cell = row[2] if len(row) > 2 else None
                qty_cell = row[6] if len(row) > 6 else None
                art_cell = row[1] if len(row) > 1 else None
                prv_cell = None
            else:  # COMP
                cod_cell = row[3] if len(row) > 3 else None
                qty_cell = row[4] if len(row) > 4 else None
                art_cell = row[1] if len(row) > 1 else None
                prv_cell = row[2] if len(row) > 2 else None

            if cod_cell is None or qty_cell is None: continue
            try:
                cod = str(int(float(str(cod_cell).strip())))
            except Exception:
                continue
            try:
                qty = float(str(qty_cell).replace(",", "."))
            except Exception:
                continue
            if qty <= 0: continue

            art = str(art_cell).strip() if art_cell else ""
            if art in ("nan", "None", ""): art = lookup.get(cod, ("", "", ""))[2]
            prv = str(prv_cell).strip() if prv_cell else None
            if prv in ("nan", "None", ""): prv = None

            ri = lookup.get(cod, ("", "", art))
            rub = ri[0]; subrub = ri[1]
            if not art: art = ri[2]

            key = (file_date, current_dep, cod)
            if key not in existing_keys:
                new_rows.append((file_date, current_dep, art, int(cod), prv, qty, rub, subrub))
                existing_keys.add(key)

        wb2.close()

    if new_rows:
        wb3 = openpyxl.load_workbook(CONS_FILE)
        ws3 = wb3.active
        for row in new_rows:
            ws3.append(list(row))
        wb3.save(CONS_FILE)
        wb3.close()
        print(f"    +{len(new_rows)} filas agregadas al consolidado.")
        return max(r[0] for r in new_rows)
    return last_date


# ─── 2. STOCK desde consolidado ───────────────────────────────────────────────
def parse_stock(inv_dir):
    """
    Lee el archivo Stock_consolidado_por_deposito_y_dia.xlsx.
    Para cada producto toma el stock del dia mas reciente por deposito.
    Columnas: fecha | deposito | articulo | codigo | proveedor | cantidad | rubro | subrubro
    Devuelve (inv_data_list, last_date_str).
    """
    try:
        import openpyxl
    except ImportError:
        import subprocess, sys
        subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl",
                        "--break-system-packages", "-q"])
        import openpyxl

    K_DEPS  = {"KLOZER"}
    KM_DEPS = {"KLOZER MKT"}
    O_DEPS  = {"OFI", "OFICINA"}
    SG_DEPS = {"SHOP GALLERY"}
    AV_DEPS = {"AVOLTA"}

    prod_info = {}   # cod -> {art, rub, sub}
    all_rows  = []   # lista de (file_date, dep_grp, cod, qty)
    dep_max   = {}   # dep_grp -> fecha maxima disponible

    wb = openpyxl.load_workbook(CONS_FILE, read_only=True, data_only=True)
    ws = wb.active

    for row in ws.iter_rows(values_only=True):
        if not row or row[0] == "fecha": continue
        # Columnas: fecha(0) deposito(1) articulo(2) codigo(3) proveedor(4) cantidad(5) rubro(6) subrubro(7)
        fecha_raw = row[0]
        dep_raw   = str(row[1]).strip().upper() if row[1] else ""
        art_raw   = str(row[2]).strip()         if row[2] else ""
        cod_raw   = row[3]
        qty_raw   = row[5]
        rub_raw   = str(row[6]).strip()         if row[6] else ""
        sub_raw   = str(row[7]).strip()         if row[7] else ""

        if cod_raw is None or qty_raw is None: continue
        try:
            cod = str(int(float(str(cod_raw))))
        except Exception:
            continue
        try:
            qty = float(str(qty_raw).replace(",", "."))
        except Exception:
            continue
        if qty < 0: qty = 0

        file_date = str(fecha_raw)[:10] if fecha_raw else ""
        if not file_date: continue

        if cod not in prod_info:
            prod_info[cod] = {"art": art_raw, "rub": rub_raw or "OTROS", "sub": sub_raw}
        elif art_raw and len(art_raw) > len(prod_info[cod]["art"]):
            prod_info[cod]["art"] = art_raw

        dep_grp = None
        if dep_raw in K_DEPS:    dep_grp = "klozer"
        elif dep_raw in KM_DEPS: dep_grp = "klozer_mkt"
        elif dep_raw in O_DEPS:  dep_grp = "ofi"
        elif dep_raw in SG_DEPS: dep_grp = "shop_gallery"
        elif dep_raw in AV_DEPS: dep_grp = "avolta"

        if dep_grp:
            all_rows.append((file_date, dep_grp, cod, qty))
            if dep_grp not in dep_max or file_date > dep_max[dep_grp]:
                dep_max[dep_grp] = file_date

    wb.close()

    # Solo usar datos de la fecha MAS RECIENTE de cada deposito.
    # Si un producto no aparece en esa fecha → stock = 0 para ese deposito.
    prod_best = {}   # cod -> {dep_grp -> qty}
    for file_date, dep_grp, cod, qty in all_rows:
        if file_date != dep_max.get(dep_grp): continue
        if cod not in prod_best: prod_best[cod] = {}
        # Usar max para tolerar filas duplicadas en el consolidado
        prod_best[cod][dep_grp] = max(prod_best[cod].get(dep_grp, 0), qty)

    print(f"    Fecha max por deposito: { {k: v for k,v in sorted(dep_max.items())} }")

    # Build INV_DATA (sin velocidad — se aplica despues desde ventas)
    # Construir lista de todos los codigos conocidos (con o sin stock hoy)
    all_cods = set(prod_info.keys())
    # También incluir productos que aparecieron en prod_best
    all_cods |= set(prod_best.keys())

    inv_data = []
    for cod in all_cods:
        deps    = prod_best.get(cod, {})
        klozer       = round(deps.get("klozer",       0))
        klozer_mkt   = round(deps.get("klozer_mkt",   0))
        ofi          = round(deps.get("ofi",           0))
        shop_gallery = round(deps.get("shop_gallery",  0))
        avolta       = round(deps.get("avolta",        0))
        both         = klozer + klozer_mkt + ofi + shop_gallery + avolta

        # Ignorar articulos sin stock y sin codigo numerico
        if klozer == 0 and klozer_mkt == 0 and ofi == 0 and shop_gallery == 0 and avolta == 0:
            try:
                int(cod)
            except (ValueError, TypeError):
                continue

        # Fecha del deposito = la fecha maxima de ese deposito (si tiene datos)
        k_fecha  = dep_max.get("klozer",       "") if klozer       > 0 else ""
        km_fecha = dep_max.get("klozer_mkt",   "") if klozer_mkt   > 0 else ""
        o_fecha  = dep_max.get("ofi",          "") if ofi          > 0 else ""
        sg_fecha = dep_max.get("shop_gallery", "") if shop_gallery > 0 else ""
        av_fecha = dep_max.get("avolta",       "") if avolta       > 0 else ""

        info = prod_info.get(cod, {})
        inv_data.append({
            "cod": cod, "art": info.get("art", ""), "rub": info.get("rub", ""),
            "sub": info.get("sub", ""),
            "klozer": klozer, "klozer_mkt": klozer_mkt, "ofi": ofi,
            "shop_gallery": shop_gallery, "avolta": avolta, "both": both,
            "k_fecha": k_fecha, "km_fecha": km_fecha, "o_fecha": o_fecha,
            "sg_fecha": sg_fecha, "av_fecha": av_fecha,
            # velocidad se rellena luego
            "k90": 0.0, "k365": 0.0, "min_k": 0.0, "min_k90": 0.0, "min_k365": 0.0,
            "o90": 0.0, "o365": 0.0, "min_o": 0.0, "min_o90": 0.0, "min_o365": 0.0,
            "b90": 0.0, "b365": 0.0, "min_b": 0.0,
            "pk": 0.0, "pk90": 0.0, "pk365": 0.0, "pk_dep": 0.0, "pk_mkt": 0.0, "pk_mkt90": 0.0, "pk_mkt365": 0.0,
            "po": 0.0, "po90": 0.0, "po365": 0.0, "pb": 0.0,
        })

    stock_hasta = max(dep_max.values()) if dep_max else date.today().strftime("%Y-%m-%d")
    inv_data.sort(key=lambda x: x["art"])
    return inv_data, stock_hasta


# ─── 2. VENTAS ────────────────────────────────────────────────────────────────
DEP_MAP = {
    "KLOZER": "KLOZER", "KLOZER MKT": "KLOZER_MKT",
    "OFI": "OFICINA", "OFICINA": "OFICINA",
    "AVOLTA": "AVOLTA", "FILIDORO": "FILIDORO",
    "SHOP GALLERY": "SHOP GALLERY",
    "DIRECTORO EXPORTACIONES": "DIRECTORO EXPORTACIONES",
    "SUR BONIFICACIONES": "SUR BONIFICACIONES",
    "BIERHAUS (FRIO)": "BIERHAUS",
    "QUIRINO": "QUIRINO",
    "POLDITOS/MERLE": "POLDITOS/MERLE",
}


def fetch_costos():
    """Descarga CODIGO + COSTO FINAL de Google Sheets (pestaña con gid=173195948)."""
    url = ("https://docs.google.com/spreadsheets/d/"
           "12Ln3lXaWDqpx5hYndJcGH1QDlPaTab3yyEX8j_ahDyY/"
           "export?format=csv&gid=173195948")
    costos = {}
    try:
        import urllib.request, csv, io
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=20) as resp:
            text = resp.read().decode("utf-8-sig")
        reader = csv.reader(io.StringIO(text))
        headers = None
        col_cod = col_cost = None
        for row in reader:
            if headers is None:
                headers = [h.strip().upper() for h in row]
                print(f"  Costos headers: {headers[:10]}")
                for i, h in enumerate(headers):
                    if h in ("CODIGO", "COD", "CODE", "CÓDIGO"): col_cod = i
                    if "COSTO FINAL" in h: col_cost = i
                if col_cod is None or col_cost is None:
                    print(f"  Costos: no se encontraron columnas CODIGO/COSTO FINAL")
                    break
                continue
            if len(row) <= max(col_cod, col_cost): continue
            cod_raw = row[col_cod].strip()
            cost_raw = row[col_cost].strip().replace("$", "").replace(".", "").replace(",", ".")
            if not cod_raw or not cost_raw: continue
            try:
                cod  = str(int(float(cod_raw)))
                cost = float(cost_raw)
                if cost > 0: costos[cod] = cost
            except Exception:
                pass
        print(f"  Costos: {len(costos)} precios cargados de Google Sheets")
    except Exception as e:
        print(f"  Advertencia costos Google Sheets: {e}")
    return costos


def parse_ventas(ventas_path, prod_lookup=None, costos=None):
    """
    Lee Salidas/Bosque salidas.xlsx (hoja SALIDAS).
    Columnas: FECHA(dd/mm/yyyy) | RAZON SOCIAL | CODIGO | CANTIDAD | DEPOSITO | NOMBRE FANTASIA
    Rubro/subrubro se enriquece desde prod_lookup (PRODUCTOS.xlsx).
    Costo = CANTIDAD * costo_unitario (de costos dict si disponible).
    """
    if prod_lookup is None: prod_lookup = {}
    if costos      is None: costos      = {}

    try:
        import openpyxl
    except ImportError:
        import subprocess, sys
        subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl",
                        "--break-system-packages", "-q"])
        import openpyxl

    if not os.path.exists(ventas_path):
        print(f"  Archivo salidas no encontrado: {ventas_path}")
        return {"VD": {}, "MONTHLY_DATA": {}, "PROD_DATA": {}}, "?", {}

    wb = openpyxl.load_workbook(ventas_path, read_only=True, data_only=True)
    ws = wb["SALIDAS"] if "SALIDAS" in wb.sheetnames else wb.active
    rows_iter = ws.iter_rows(values_only=True)

    # Fila 1 = vacía/título, fila 2 = cabecera — saltar ambas
    next(rows_iter, None)
    next(rows_iter, None)

    monthly = {}   # yr -> mo_str -> {deps:{}, prods:{cod:{}}}
    row_count = skip_count = 0

    for row in rows_iter:
        if not row or not any(row):
            continue

        fecha_raw = row[0] if len(row) > 0 else None
        cod_raw   = row[2] if len(row) > 2 else None
        qty_raw   = row[3] if len(row) > 3 else None
        dep_raw   = row[4] if len(row) > 4 else None

        if fecha_raw is None or cod_raw is None or qty_raw is None:
            skip_count += 1; continue

        # ── Fecha ──
        try:
            if hasattr(fecha_raw, "year"):
                yr_int, mo = fecha_raw.year, fecha_raw.month
            else:
                s = str(fecha_raw).strip()
                parts = s.replace("/", "-").split("-")
                if len(parts) == 3:
                    if len(parts[0]) == 4:
                        yr_int, mo = int(parts[0]), int(parts[1])
                    else:
                        yr_int, mo = int(parts[2]), int(parts[1])
                else:
                    skip_count += 1; continue
            if not (2020 <= yr_int <= 2030 and 1 <= mo <= 12):
                skip_count += 1; continue
            yr = str(yr_int)
        except Exception:
            skip_count += 1; continue

        # ── Cantidad ──
        try:
            u = float(str(qty_raw).replace(",", "."))
        except Exception:
            skip_count += 1; continue
        if u <= 0:
            continue

        # ── Código ──
        try:
            cod = str(int(float(str(cod_raw).strip())))
        except Exception:
            cod = str(cod_raw).strip()
        if not cod or cod in ("nan", "None", ""):
            skip_count += 1; continue

        # ── Depósito ──
        dep_str = str(dep_raw).strip().upper() if dep_raw else ""
        dep = DEP_MAP.get(dep_str, dep_str if dep_str else "SIN DEPOSITO")

        # ── Enriquecer desde productos ──
        pi  = prod_lookup.get(cod, {})
        rub = pi.get("rub", "SIN RUBRO") or "SIN RUBRO"
        sub = pi.get("sub", "") or ""
        art = pi.get("art", "") or ""

        # ── Costo ──
        costo_unit = costos.get(cod, 0.0)
        c = round(u * costo_unit, 2)
        m = 0.0  # sin precio de venta en el archivo de salidas

        mo_str = str(mo)
        row_count += 1

        if yr not in monthly:
            monthly[yr] = {}
        if mo_str not in monthly[yr]:
            monthly[yr][mo_str] = {"deps": {}, "prods": {}}
        mo_data = monthly[yr][mo_str]

        # deps
        if dep not in mo_data["deps"]:
            mo_data["deps"][dep] = {"m": 0.0, "c": 0.0, "u": 0.0}
        mo_data["deps"][dep]["c"] += c
        mo_data["deps"][dep]["u"] += u

        # prods
        if cod not in mo_data["prods"]:
            mo_data["prods"][cod] = {"cod": cod, "art": art, "rub": rub, "sub": sub,
                                     "m": 0.0, "c": 0.0, "u": 0.0, "deps": {}}
        p = mo_data["prods"][cod]
        if art and not p["art"]: p["art"] = art
        p["c"] += c; p["u"] += u
        if dep not in p["deps"]:
            p["deps"][dep] = {"m": 0.0, "c": 0.0, "u": 0.0}
        p["deps"][dep]["c"] += c
        p["deps"][dep]["u"] += u

    wb.close()
    print(f"  Salidas: {row_count} filas procesadas, {skip_count} saltadas")

    # ── Build final structures ────────────────────────────────────────────────
    VD      = {}
    VD_out  = {"annual": {}, "monthly": {}, "deps": {}, "rubros": {}}
    MONTHLY_DATA = {}
    PROD_DATA    = {}

    for yr in sorted(monthly.keys()):
        yr_mo = monthly[yr]

        # Totales anuales
        tot_c = sum(p["c"] for mo_s in yr_mo for p in yr_mo[mo_s]["prods"].values())
        tot_u = sum(p["u"] for mo_s in yr_mo for p in yr_mo[mo_s]["prods"].values())
        VD[yr] = {"m": 0.0, "c": round(tot_c, 2), "u": round(tot_u, 2)}
        VD_out["annual"][yr] = {**VD[yr], "mg": 0.0}

        # Monthly series
        series = []
        for mo in range(1, 13):
            mo_s = str(mo)
            if mo_s not in yr_mo: continue
            prods_mo = list(yr_mo[mo_s]["prods"].values())
            tot_c_mo = sum(p["c"] for p in prods_mo)
            tot_u_mo = sum(p["u"] for p in prods_mo)
            if tot_u_mo > 0:
                series.append({"mes": mo, "label": MESES_NOM[mo],
                                "m": 0, "c": round(tot_c_mo, 2), "u": round(tot_u_mo, 2),
                                "mg": 0.0})
        VD_out["monthly"][yr] = series

        # Deps aggregate
        dep_tot = {}
        for mo_s, mo_data in yr_mo.items():
            for dep, dv in mo_data["deps"].items():
                if dep not in dep_tot: dep_tot[dep] = {"m": 0.0, "c": 0.0, "u": 0.0}
                dep_tot[dep]["c"] += dv["c"]; dep_tot[dep]["u"] += dv["u"]
        for v in dep_tot.values(): v["mg"] = 0.0
        VD_out["deps"][yr] = dep_tot

        # Rubros: {rub: {dep: {m,c,u}}}
        rub_dep = {}
        for mo_s, mo_data in yr_mo.items():
            for p in mo_data["prods"].values():
                rub = p["rub"]
                if rub not in rub_dep: rub_dep[rub] = {}
                for dep, dv in p["deps"].items():
                    if dep not in rub_dep[rub]: rub_dep[rub][dep] = {"m": 0.0, "c": 0.0, "u": 0.0}
                    rub_dep[rub][dep]["c"] += dv["c"]; rub_dep[rub][dep]["u"] += dv["u"]
        VD_out["rubros"][yr] = rub_dep

        # MONTHLY_DATA
        md_yr = {"deps": {}, "rubs": {}, "prods": {}}
        for mo_s, mo_data in yr_mo.items():
            md_yr["deps"][mo_s] = {dep: {**dv, "mg": 0.0}
                                    for dep, dv in mo_data["deps"].items()}
            # rubs desde prods (para que tenga sub correcto)
            rubs_mo = {}
            for p in mo_data["prods"].values():
                rk = p["rub"]
                if rk not in rubs_mo: rubs_mo[rk] = {"m": 0.0, "c": 0.0, "u": 0.0}
                rubs_mo[rk]["c"] += p["c"]; rubs_mo[rk]["u"] += p["u"]
            md_yr["rubs"][mo_s] = rubs_mo
            prods_list = sorted(mo_data["prods"].values(), key=lambda p: -p["u"])
            for p in prods_list: p["mg"] = 0.0
            md_yr["prods"][mo_s] = prods_list
        MONTHLY_DATA[yr] = md_yr

        # PROD_DATA
        prod_agg = {}
        for mo_s, mo_data in yr_mo.items():
            for cod, p in mo_data["prods"].items():
                if cod not in prod_agg:
                    prod_agg[cod] = {"cod": cod, "art": p["art"], "rub": p["rub"], "sub": p["sub"],
                                     "m": 0.0, "c": 0.0, "u": 0.0, "deps": {}}
                pa = prod_agg[cod]
                if p["art"] and not pa["art"]: pa["art"] = p["art"]
                pa["c"] += p["c"]; pa["u"] += p["u"]
                for dep, dv in p["deps"].items():
                    if dep not in pa["deps"]: pa["deps"][dep] = {"m": 0.0, "c": 0.0, "u": 0.0}
                    pa["deps"][dep]["c"] += dv["c"]; pa["deps"][dep]["u"] += dv["u"]

        total_u = sum(p["u"] for p in prod_agg.values())
        prods_sorted = sorted(prod_agg.values(), key=lambda p: -p["u"])
        acum = 0
        for p in prods_sorted:
            p["mg"]   = 0.0
            p["pct"]  = round(p["u"] / total_u * 100, 2) if total_u > 0 else 0
            acum     += p["pct"]
            p["acum"] = round(acum, 2)
            p["tk"]   = 0
        PROD_DATA[yr] = prods_sorted

    ventas_max_yr = max(monthly.keys()) if monthly else "2026"
    ventas_max_mo = max(int(k) for k in monthly.get(ventas_max_yr, {"0": None}).keys()) if monthly else 4
    ventas_hasta  = f"{ventas_max_yr}-{ventas_max_mo:02d}"

    return {"VD": VD_out, "MONTHLY_DATA": MONTHLY_DATA, "PROD_DATA": PROD_DATA}, ventas_hasta, monthly


# ─── 3. VELOCIDAD DESDE VENTAS ────────────────────────────────────────────────
def compute_velocity(monthly, ref_date):
    """
    Para cada (cod, dep_canonico) calcula:
      q90:  unidades vendidas en los ultimos ~90 dias (3 meses hacia atras)
      q365: unidades vendidas en los ultimos ~365 dias (12 meses hacia atras)
    Devuelve dict:  vel[cod][dep_canonico] = {"q90": x, "q365": y}
    DEP_MAP ya unifico KLOZER MKT->KLOZER, OFI/OFICINA->OFICINA en ventas.
    """
    cutoff_90  = ref_date - timedelta(days=90)
    cutoff_365 = ref_date - timedelta(days=365)

    vel = {}  # cod -> {dep -> {q90, q365}}

    for yr, yr_mo in monthly.items():
        for mo_s, mo_data in yr_mo.items():
            mo = int(mo_s)
            # Usamos el ultimo dia del mes para comparar
            if mo == 12:
                month_end = date(int(yr) + 1, 1, 1) - timedelta(days=1)
            else:
                month_end = date(int(yr), mo + 1, 1) - timedelta(days=1)

            in_365 = month_end >= cutoff_365
            in_90  = month_end >= cutoff_90

            if not in_365:
                continue

            for p in mo_data["prods"].values():
                cod = p["cod"]
                if cod not in vel:
                    vel[cod] = {}
                for dep, dv in p.get("deps", {}).items():
                    if dep not in vel[cod]:
                        vel[cod][dep] = {"q90": 0.0, "q365": 0.0}
                    u = dv.get("u", 0)
                    vel[cod][dep]["q365"] += u
                    if in_90:
                        vel[cod][dep]["q90"] += u

    return vel


def apply_velocity(inv_data, vel):
    """Aplica velocidad calculada desde ventas a cada articulo de INV_DATA.
    Ventas de KLOZER y KLOZER MKT quedan separadas en el stock pero en ventas
    ambas se registran bajo el canal 'KLOZER', por lo que la velocidad se
    asigna al total klozer+klozer_mkt para calcular meses de stock combinados.
    """
    for item in inv_data:
        cod = str(item["cod"])
        cv = vel.get(cod, {})

        # Velocidad KLOZER depósito (sin MKT)
        kv = cv.get("KLOZER", {})
        k90  = round(kv.get("q90",  0.0), 2)
        k365 = round(kv.get("q365", 0.0), 2)

        # Velocidad KLOZER MKT (canal propio)
        kmv  = cv.get("KLOZER_MKT", {})
        km90_raw  = round(kmv.get("q90",  0.0), 2)
        km365_raw = round(kmv.get("q365", 0.0), 2)

        # Velocidad OFICINA
        ov = cv.get("OFICINA", {})
        o90  = round(ov.get("q90",  0.0), 2)
        o365 = round(ov.get("q365", 0.0), 2)

        # Velocidad combinada para meses de stock totales (klozer+mkt+ofi)
        b90  = round(k90 + km90_raw + o90,  2)
        b365 = round(k365 + km365_raw + o365, 2)

        # min/mes KLOZER (solo depósito KLOZER)
        min_k90  = round(k90  / 90  * 30, 1) if k90  > 0 else 0.0
        min_k365 = round(k365 / 365 * 30, 1) if k365 > 0 else 0.0
        _cnt_k   = (1 if min_k90 else 0) + (1 if min_k365 else 0)
        min_k    = round((min_k90 + min_k365) / _cnt_k, 1) if _cnt_k else 0.0

        # min/mes KLOZER MKT (propio)
        min_km90  = round(km90_raw  / 90  * 30, 1) if km90_raw  > 0 else 0.0
        min_km365 = round(km365_raw / 365 * 30, 1) if km365_raw > 0 else 0.0
        _cnt_km   = (1 if min_km90 else 0) + (1 if min_km365 else 0)
        min_km    = round((min_km90 + min_km365) / _cnt_km, 1) if _cnt_km else 0.0

        min_o90  = round(o90  / 90  * 30, 1) if o90  > 0 else 0.0
        min_o365 = round(o365 / 365 * 30, 1) if o365 > 0 else 0.0
        _cnt_o   = (1 if min_o90 else 0) + (1 if min_o365 else 0)
        min_o    = round((min_o90 + min_o365) / _cnt_o, 1) if _cnt_o else 0.0

        _mb90    = round(b90  / 90  * 30, 1) if b90  > 0 else 0.0
        _mb365   = round(b365 / 365 * 30, 1) if b365 > 0 else 0.0
        _cnt_b   = (1 if _mb90 else 0) + (1 if _mb365 else 0)
        min_b    = round((_mb90 + _mb365) / _cnt_b, 1) if _cnt_b else 0.0

        # Meses de stock KLOZER depósito
        k_total = item["klozer"] + item.get("klozer_mkt", 0)
        km      = item.get("klozer_mkt", 0)
        pk      = round(k_total / min_k,    2) if min_k    > 0 else 0.0
        pk90    = round(k_total / min_k90,  2) if min_k90  > 0 else 0.0
        pk365   = round(k_total / min_k365, 2) if min_k365 > 0 else 0.0
        pk_dep  = round(item["klozer"] / min_k, 2) if min_k > 0 else 0.0
        # Meses de stock KLOZER MKT usando velocidad propia
        pk_mkt    = round(km / min_km,    2) if min_km    > 0 else 0.0
        pk_mkt90  = round(km / min_km90,  2) if min_km90  > 0 else 0.0
        pk_mkt365 = round(km / min_km365, 2) if min_km365 > 0 else 0.0
        po      = round(item["ofi"]  / min_o,    2) if min_o    > 0 else 0.0
        po90    = round(item["ofi"]  / min_o90,  2) if min_o90  > 0 else 0.0
        po365   = round(item["ofi"]  / min_o365, 2) if min_o365 > 0 else 0.0
        pb      = round(item["both"] / min_b,    2) if min_b    > 0 else 0.0

        item.update({
            "k90": k90, "k365": k365,
            "min_k": min_k, "min_k90": min_k90, "min_k365": min_k365,
            "km90": km90_raw, "km365": km365_raw,
            "min_km": min_km, "min_km90": min_km90, "min_km365": min_km365,
            "o90": o90, "o365": o365,
            "min_o": min_o, "min_o90": min_o90, "min_o365": min_o365,
            "b90": b90, "b365": b365, "min_b": min_b,
            "pk": pk, "pk90": pk90, "pk365": pk365,
            "pk_dep": pk_dep,
            "pk_mkt": pk_mkt, "pk_mkt90": pk_mkt90, "pk_mkt365": pk_mkt365,
            "po": po, "po90": po90, "po365": po365, "pb": pb,
        })
    return inv_data



# ─── 5. LOOKUP DESDE INVENTARIO GENERAL ──────────────────────────────────────
PROD_F = os.path.join(DATA_DIR, "Productos", "PRODUCTOS.xlsx")

def load_productos():
    """
    Lee Inventario/Productos.xlsx y devuelve lookup:
      {cod_str: {"art": NOMBRE, "rub": RUBRO, "sub": SUB_RUBRO}}
    Detecta automaticamente el orden de columnas desde la fila de encabezado.
    Columnas soportadas: CODIGO/RUBRO/SUB RUBRO/NOMBRE (cualquier orden).
    """
    lookup = {}
    if not os.path.exists(PROD_F):
        print("  Advertencia: no se encontro Productos.xlsx")
        return lookup
    try:
        import openpyxl
    except ImportError:
        import subprocess, sys
        subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl",
                        "--break-system-packages", "-q"])
        import openpyxl
    try:
        wb = openpyxl.load_workbook(PROD_F, read_only=True, data_only=True)
        sh_name = "DESCRIPCION" if "DESCRIPCION" in wb.sheetnames else wb.sheetnames[0]
        ws = wb[sh_name]
        rows = list(ws.iter_rows(values_only=True))
        # Encontrar fila de encabezado (contiene CODIGO o NOMBRE)
        hdr_idx = None
        col_cod = col_nom = col_rub = col_sub = None
        for i, r in enumerate(rows):
            row_up = [str(c).upper().strip() if c else "" for c in r]
            if any(x in row_up for x in ("CODIGO", "NOMBRE", "CODE")):
                hdr_idx = i
                for j, h in enumerate(row_up):
                    if h in ("CODIGO", "CODE", "COD"): col_cod = j
                    elif h in ("NOMBRE", "DESCRIPCION", "NAME"): col_nom = j
                    elif "SUB" in h: col_sub = j
                    elif h == "RUBRO": col_rub = j
                break
        if hdr_idx is None or col_cod is None:
            print("  Advertencia Productos.xlsx: no se encontro encabezado con CODIGO")
            return lookup
        # Defaults si alguna columna no existe
        if col_nom is None: col_nom = col_cod + 1
        if col_rub is None: col_rub = -1
        if col_sub is None: col_sub = -1
        for r in rows[hdr_idx + 1:]:
            if not r or r[col_cod] is None: continue
            cod_raw = str(r[col_cod]).strip()
            nom_raw = str(r[col_nom]).strip() if col_nom < len(r) and r[col_nom] else ""
            rub_raw = str(r[col_rub]).strip().upper() if col_rub >= 0 and col_rub < len(r) and r[col_rub] else ""
            sub_raw = str(r[col_sub]).strip().upper() if col_sub >= 0 and col_sub < len(r) and r[col_sub] else ""
            if not cod_raw or cod_raw in ("nan","None",""): continue
            # Normalizar codigos numericos
            try:
                cod = str(int(float(cod_raw)))
            except Exception:
                cod = cod_raw
            if not nom_raw or nom_raw in ("nan","None",""): continue
            lookup[cod] = {"art": nom_raw, "rub": rub_raw or "OTROS", "sub": sub_raw}
        wb.close()
        print(f"  Productos: {len(lookup)} articulos cargados desde Productos.xlsx")
    except Exception as e:
        print(f"  Advertencia Productos.xlsx: {e}")
    return lookup


# ─── 6. PROYECCIONES DE COMPRA ───────────────────────────────────────────────
PROY_DIR  = os.path.join(BASE, "Data", "Supply Chain", "proyecciones")
PROY_FILE = os.path.join(PROY_DIR, "proyecciones.xlsx")

def generate_proyecciones(inv_data):
    """
    Genera Data/Supply Chain/proyecciones/proyecciones.xlsx con los productos
    que tienen menos de 3 meses de stock (segun velocidad KLOZER).
    Preserva las columnas de aprobacion si el archivo ya existe.
    Columnas: Codigo | Descripcion | Rubro | Subrubro | Stock | Min/mes | Meses |
              Sug_30d | Sug_60d | Sug_90d | Cant_aprobada | Aprobado | Fecha_aprobacion
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        import subprocess, sys
        subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl",
                        "--break-system-packages", "-q"])
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

    os.makedirs(PROY_DIR, exist_ok=True)

    # Leer aprobaciones existentes para no perderlas al regenerar
    existing = {}   # cod -> {cant_aprobada, aprobado, fecha}
    if os.path.exists(PROY_FILE):
        try:
            wb_old = openpyxl.load_workbook(PROY_FILE, data_only=True)
            ws_old = wb_old.active
            hdrs   = [str(c.value).strip().lower() if c.value else "" for c in next(ws_old.iter_rows())]
            i_cod  = next((i for i, h in enumerate(hdrs) if "cod" in h), None)
            i_cant = next((i for i, h in enumerate(hdrs) if "cant" in h), None)
            i_ap   = next((i for i, h in enumerate(hdrs) if h == "aprobado"), None)
            i_fec  = next((i for i, h in enumerate(hdrs) if "fecha" in h), None)
            if i_cod is not None:
                for row in ws_old.iter_rows(min_row=2, values_only=True):
                    if not row or row[i_cod] is None: continue
                    try:    cod = str(int(float(str(row[i_cod]))))
                    except: cod = str(row[i_cod])
                    existing[cod] = {
                        "cant": row[i_cant] if i_cant is not None else None,
                        "ap":   row[i_ap]   if i_ap   is not None else None,
                        "fec":  row[i_fec]  if i_fec  is not None else None,
                    }
            wb_old.close()
        except Exception as e:
            print(f"    Advertencia proyecciones existentes: {e}")

    # Filtrar productos con < 3 meses de stock y velocidad conocida
    alerta = [it for it in inv_data if it.get("min_k", 0) > 0 and it.get("pk", 0) < 3]
    alerta.sort(key=lambda x: x.get("pk", 0))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Proyecciones"

    HEADERS = ["Codigo", "Descripcion", "Rubro", "Subrubro", "Stock",
               "Min/mes", "Meses", "Sug 30d", "Sug 60d", "Sug 90d",
               "Cant aprobada", "Aprobado", "Fecha aprobacion"]
    WIDTHS  = [10, 38, 16, 16, 10, 10, 9, 10, 10, 10, 14, 11, 20]

    hdr_fill = PatternFill("solid", start_color="1c2230", end_color="1c2230")
    hdr_font = Font(bold=True, color="4ADE80", name="Arial")
    cen = Alignment(horizontal="center")

    for col, (h, w) in enumerate(zip(HEADERS, WIDTHS), 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = hdr_font; c.fill = hdr_fill; c.alignment = cen
        ws.column_dimensions[c.column_letter].width = w

    red_font  = Font(color="EF4444", bold=True, name="Arial")
    yel_font  = Font(color="EAB308", bold=True, name="Arial")
    norm_font = Font(name="Arial")

    for rn, it in enumerate(alerta, 2):
        cod   = str(it["cod"])
        stock = it.get("klozer", 0) + it.get("klozer_mkt", 0)
        min_k = round(it.get("min_k", 0), 1)
        meses = round(it.get("pk", 0),    1)
        sug30 = max(0, round(min_k * 1 - stock))
        sug60 = max(0, round(min_k * 2 - stock))
        sug90 = max(0, round(min_k * 3 - stock))
        ex    = existing.get(cod, {})

        vals = [int(cod) if cod.isdigit() else cod,
                it.get("art", ""), it.get("rub", ""), it.get("sub", ""),
                stock, min_k, meses,
                sug30, sug60, sug90,
                ex.get("cant"), ex.get("ap"), ex.get("fec")]

        for col, v in enumerate(vals, 1):
            c = ws.cell(row=rn, column=col, value=v)
            c.font = norm_font
        # Color en columna Meses según criticidad
        meses_cell = ws.cell(row=rn, column=7)
        if meses < 1:   meses_cell.font = red_font
        elif meses < 2: meses_cell.font = yel_font

    ws.freeze_panes = "A2"
    wb.save(PROY_FILE)
    print(f"    Proyecciones: {len(alerta)} productos con < 3 meses de stock -> {PROY_FILE}")


# ─── 7. MAIN ──────────────────────────────────────────────────────────────────
def main():
    print("=" * 60)
    print(f"Actualizando Bosque Gin Dashboard — {datetime.now():%Y-%m-%d %H:%M}")
    print("=" * 60)

    # Destileria existente
    destileria = {}
    if os.path.exists(OUT_JS):
        try:
            raw     = open(OUT_JS, encoding="utf-8").read()
            js_body = raw.strip()[len("var BOSQUE_DATA="):-1]
            old     = json.loads(js_body)
            destileria = old.get("destileria", {})
            print("  Destileria conservada: %d meses" % len(destileria.get("months", [])))
        except Exception as e:
            print("  Advertencia destileria: %s" % e)

    print("\n[1/5] Actualizando consolidado de inventario...")
    try:
        update_consolidado()
    except Exception as e:
        print("  Advertencia consolidado: %s" % e)

    print("\n[2/5] Cargando productos (rubro/subrubro)...")
    try:
        inv_gen = load_productos()
        print("  -> %d productos cargados" % len(inv_gen))
    except Exception as e:
        print("  Advertencia productos: %s" % e)
        inv_gen = {}

    print("\n[3/5] Leyendo inventario desde consolidado...")
    try:
        inv_data, stock_hasta = parse_stock(INV_DIR)
        print("  -> %d articulos al %s" % (len(inv_data), stock_hasta))
        if inv_gen:
            for item in inv_data:
                gen = inv_gen.get(str(item["cod"]))
                if gen:
                    if gen["art"]: item["art"] = gen["art"]
                    if gen["rub"]: item["rub"] = gen["rub"]
                    if gen["sub"]: item["sub"] = gen["sub"]
    except Exception as e:
        print("  ERROR stock: %s" % e)
        inv_data, stock_hasta = [], date.today().strftime("%Y-%m-%d")

    print("\n[4/5] Descargando costos desde Google Sheets...")
    try:
        costos = fetch_costos()
    except Exception as e:
        print("  Advertencia costos: %s" % e)
        costos = {}

    print("\n[4b/5] Leyendo salidas...")
    try:
        ventas, ventas_hasta, monthly_raw = parse_ventas(VENTAS_F, prod_lookup=inv_gen, costos=costos)
        print("  -> Datos hasta %s" % ventas_hasta)
    except Exception as e:
        print("  ERROR ventas: %s" % e)
        ventas, ventas_hasta, monthly_raw = {"VD": {}, "MONTHLY_DATA": {}, "PROD_DATA": {}}, "?", {}

    print("\n[5/5] Calculando velocidad de ventas por deposito...")
    try:
        ref = date.today()
        vel = compute_velocity(monthly_raw, ref)
        inv_data = apply_velocity(inv_data, vel)
        con_vel = sum(1 for x in inv_data if x["min_k"] > 0 or x["min_o"] > 0)
        print("  -> %d articulos con velocidad calculada" % con_vel)
    except Exception as e:
        print("  ERROR velocidad: %s" % e)

    new_data = {
        "meta": {
            "generado":         datetime.now().strftime("%Y-%m-%dT%H:%M:%S"),
            "ventas_hasta":     ventas_hasta,
            "stock_hasta":      stock_hasta,
            "destileria_hasta": max((m["key"] for m in destileria.get("months", [])), default="?"),
        },
        "ventas":    ventas,
        "stock":     {"INV_DATA": inv_data},
        "destileria": destileria,
    }

    js_str = "var BOSQUE_DATA=" + json.dumps(new_data, ensure_ascii=False) + ";"
    with open(OUT_JS, "w", encoding="utf-8") as f:
        f.write(js_str)
    print("\nOK: bosquegin_data.js actualizado (%d chars)" % len(js_str))
    print("  Stock:   %d articulos al %s" % (len(inv_data), stock_hasta))
    print("  Ventas:  hasta %s" % ventas_hasta)

    print("\n[6/6] Generando proyecciones de compra...")
    try:
        generate_proyecciones(inv_data)
    except Exception as e:
        print("  Advertencia proyecciones: %s" % e)

    print("Recarga el dashboard para ver los cambios.")

if __name__ == "__main__":
    main()
